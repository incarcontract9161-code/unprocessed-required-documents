import streamlit as st
import pandas as pd
import os
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, PageBreak, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime, timedelta
import io
import numpy as np

# ==========================================
# 0. 페이지 설정
# ==========================================
st.set_page_config(page_title="M스캔 전용 서류 처리 대시보드", layout="wide", page_icon="📊")

# ==========================================
# 1. 전역 설정 & 가이드 내용
# ==========================================
EXCEL_FILE = "insurance_data.xlsx"
TARGET_FILE = "target_settings.xlsx"
APP_PASSWORD = os.environ.get("APP_PASSWORD", "incar961")
MANUAL_FILES = ["모바일동의_독려_안내.pdf", "모바일_보험가입확인서_장기_계피동일건발송절차_v2.pdf"]

GUIDANCE_DOCS = [
    ["No.", "서류명", "법적근거", "목적 및 주요내용"],
    ["1", "개인정보동의서", "개인정보보호법 15조\n대리점 표준 내부통제기준 27조", "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에 따른 개인정보 처리로 신계약시 필수 징구"],
    ["2", "비교설명확인서", "보험업감독규정 별표 5-6", "유사 상품 3개 이상 비교·설명 이행 사실 고객 확인 서명"],
    ["3", "고지의무확인서", "금융소비자보호법 26조와 동법시행령 24조", "판매자 권한·책임·보상 관련 핵심 사항 고지, 소비자 오인 예방"],
    ["4", "완전판매확인서\n(대상: 종신, CI, CEO정기, 고액)", "금융소비자보호법 제17·19조\n영업지원기준안", "약관,청약서 부본 제공, 중요 상품 이해 및 자발적 가입 확인, 설명 의무 이행 증빙력 확보"]
]

MOBILE_GUIDE = {
    "title": "📱 모바일동의(M스캔) 집중 관리 안내",
    "reasons": [
        "**자동 매칭** : 실적 입력 시 서류 자동 연결, 수작업 업로드 불필요, 오류 감소",
        "**타임스탬프** : 전자서명 시점 실시간 기록, 계약 전 작성 객관적 증빙, 서명 위·변조 불가",
        "**누락 방지** : 필수 항목 입력후 다음 단계 진행, 불완전판매 리스크 최소화",
        "**비용 절감** : 비교확인서 스캔 시 5년 원본 보관 의무 해소"
    ],
    "faq": [
        ("스캔 업로드도 가능한가요?", "네, 가능합니다. 다만 자동매칭·타임스탬프·누락방지 기능으로 업무 효율과 법적 보호가 강력한 모바일동의를 권장합니다."),
        ("완전판매확인서는 모든 계약에 필수인가요?", "영업지원기준안대로 종신, CI, CEO정기보험 및 고액 계약(저축성 300만원, 비저축성 100만원 이상)은 필수이며, 이외에는 금소법 취지에 따라 모든 계약에서 활용을 권장합니다."),
        ("모바일동의 절차는 어떻게 진행되나요?", "고객 동의 본인인증 → 설계사 동의 본인인증 → 전자서명 → 타임스탬프 기록 → 실적 입력 매칭\n*(법인, 미성년 계약 외 모바일 동의 전건 가능)*")
    ],
    "do_list": [
        "✅ 계약 체결 전 필수 서류 4종 100% 완비 (미완비 시 불완전판매 및 리스크 계약 간주)",
        "✅ 모바일동의 적극 활용 (자동매칭·타임스탬프 확보)",
        "✅ 사후징구 금지 (2026년 5월부터 서류 미비 시 내부 통제 미충족 조직으로 관리)",
        "✅ 고객 정보 확인 기록 및 적합성 판단 근거 남기기",
        "✅ 대리·중개 고지사항(권한·책임·보상) 사전 안내"
    ]
}

PROCESS_FLOW = [
    {"step": "1", "title": "보험구분 및 계/피동일 설정", "desc": "손보/생보 선택 후 계약자와 피보험자가 동일할 경우 '계/피동일' 체크박스 선택. 체크 시 피보험자 입력칸 자동 생략."},
    {"step": "2", "title": "완전판매확인서 발송 기준 설정", "desc": "필수발송대상: 생보보장성(변액/종신/100만↑), 생보저축성(300만↑), 손보보장/저축성(100만↑). 조건 외 선택적 발송 가능."},
    {"step": "3", "title": "계약자 결재", "desc": "페이퍼 화면 변경 → 동의함/부동의함 선택 → 서명 입력 → 다음 클릭 → 마지막 페이지 상단 저장 클릭."},
    {"step": "4", "title": "피보험자 결재 (계피상이 시)", "desc": "계약자 결재 완료 시 자동 문자 발송 → 화면 확대 가능 → 계약자 결재와 동일한 순서로 진행."},
    {"step": "5", "title": "설계사 결재", "desc": "계약자/피보험자 결재 완료 시 자동 문자 발송 → 글씨 진하게 선택(필요시) → 다음 클릭 → 저장 완료."}
]

# ==========================================
# 2. 데이터 로딩 (안전 연산 적용)
# ==========================================
def safe_rate(num, den):
    return (num / den.replace(0, float('nan')) * 100).round(1).fillna(0.0)

@st.cache_data(ttl=300)
def load_data():
    if not os.path.exists(EXCEL_FILE):
        st.error(f"⚠️ '{EXCEL_FILE}' 파일이 없습니다.")
        return pd.DataFrame()
    try:
        df = pd.read_excel(EXCEL_FILE)
        if df.empty: return pd.DataFrame()

        df["보험시작일_dt"] = pd.to_datetime(df["보험시작일"], errors="coerce")
        df["월_피리어드"] = df["보험시작일_dt"].dt.to_period("M").astype(str)
        
        for col in ["FA고지", "비교설명", "완전판매"]:
            df[f"{col}_c"] = df[col].fillna(" ").astype(str).str.strip()

        def is_total_scanned(val):
            if pd.isna(val) or val == " ": return False
            return str(val).strip() in ["스캔", "M스캔", "보험사스캔"]

        def is_m_scanned(val):
            if pd.isna(val) or val == " ": return False
            return str(val).strip() == "M스캔"

        def is_cs_target(val):
            if pd.isna(val) or val == " ": return False
            return str(val).strip() in ["스캔", "M스캔", "미스캔"]

        df["FA_전체스캔"] = df["FA고지_c"].apply(is_total_scanned).astype(int)
        df["비교_전체스캔"] = df["비교설명_c"].apply(is_total_scanned).astype(int)
        df["완판_전체스캔"] = df["완전판매_c"].apply(is_total_scanned).astype(int)

        df["FA_M스캔"] = df["FA고지_c"].apply(is_m_scanned).astype(int)
        df["비교_M스캔"] = df["비교설명_c"].apply(is_m_scanned).astype(int)
        df["완판_M스캔"] = df["완전판매_c"].apply(is_m_scanned).astype(int)

        df["완판_대상"] = df["완전판매_c"].apply(is_cs_target).astype(int)
        df["FA_target"] = 1
        df["비교_target"] = 1
        df["완판_target"] = df["완판_대상"]

        df["대상건"] = df[["FA_target", "비교_target", "완판_target"]].sum(axis=1).astype(int)
        df["전체스캔건"] = df[["FA_전체스캔", "비교_전체스캔", "완판_전체스캔"]].sum(axis=1).astype(int)
        df["M스캔건"] = df[["FA_M스캔", "비교_M스캔", "완판_M스캔"]].sum(axis=1).astype(int)

        return df
    except Exception as e:
        st.error(f"❌ 엑셀 파일 읽기 오류: {e}")
        return pd.DataFrame()

def get_file_update_time():
    if os.path.exists(EXCEL_FILE):
        return datetime.fromtimestamp(os.path.getmtime(EXCEL_FILE)).strftime("%Y-%m-%d %H:%M:%S")
    return "알 수 없음"

# ==========================================
# 3. 목표 관리 함수 (KeyError 해결 적용)
# ==========================================
@st.cache_data(ttl=300)
def load_targets():
    if not os.path.exists(TARGET_FILE):
        return pd.DataFrame(columns=["영업가족", "M스캔율_목표", "대상건_목표", "배분사유", "특이사항"])
    try:
        df = pd.read_excel(TARGET_FILE)
        # 필수 컬럼 확인 및 누락 시 추가
        required_cols = ["영업가족", "M스캔율_목표", "대상건_목표", "배분사유", "특이사항"]
        for col in required_cols:
            if col not in df.columns:
                df[col] = "" if col != "M스캔율_목표" and col != "대상건_목표" else 0
        return df
    except:
        return pd.DataFrame(columns=["영업가족", "M스캔율_목표", "대상건_목표", "배분사유", "특이사항"])

def save_targets(df_targets):
    wb = Workbook()
    ws = wb.active
    ws.title = "목표설정"
    
    headers = ["영업가족", "M스캔율_목표", "대상건_목표", "배분사유", "특이사항"]
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h).font = Font(bold=True)
    
    for ri, (_, row) in enumerate(df_targets.iterrows(), 2):
        ws.cell(ri, 1, row.get("영업가족", ""))
        ws.cell(ri, 2, row.get("M스캔율_목표", 0))
        ws.cell(ri, 3, row.get("대상건_목표", 0))
        ws.cell(ri, 4, row.get("배분사유", ""))
        ws.cell(ri, 5, row.get("특이사항", ""))
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    with open(TARGET_FILE, "wb") as f:
        f.write(buf.getvalue())

@st.cache_data(ttl=300)
def build_org_stats(df, months=None, group_cols=["영업가족"], view_mode="누적"):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()

    keys = group_cols.copy()
    if view_mode == "월별":
        keys = ["월_피리어드"] + keys

    agg_df = src.groupby(keys).agg(
        대상건=("대상건", "sum"),
        전체스캔건=("전체스캔건", "sum"),
        M스캔건=("M스캔건", "sum")
    ).reset_index()

    agg_df["M스캔율_대상"] = safe_rate(agg_df["M스캔건"], agg_df["대상건"])
    agg_df["M스캔율_완료"] = safe_rate(agg_df["M스캔건"], agg_df["전체스캔건"])

    if view_mode == "월별":
        agg_df = agg_df.rename(columns={"월_피리어드": "월"})
        agg_df["월_표시"] = agg_df["월"].apply(lambda x: f"{x.replace('-', '.')[:7]}월" if pd.notna(x) else "")
        agg_df["표시명"] = agg_df["월_표시"] + " | " + agg_df[group_cols[-1]].astype(str)
    else:
        agg_df["월_표시"] = ""
        agg_df["표시명"] = agg_df[group_cols[-1]].astype(str)

    return agg_df

# ==========================================
# 4. 목표 자동배분 함수 (KeyError 해결)
# ==========================================
def auto_allocate_targets(df_actual, df_existing, increase_rate=0.10):
    if df_actual.empty or "영업가족" not in df_actual.columns:
        return pd.DataFrame(columns=["영업가족", "M스캔율_목표", "대상건_목표", "배분사유", "특이사항"])
    
    actual_stats = df_actual.groupby("영업가족").agg({
        "대상건": "sum", "M스캔건": "sum", "전체스캔건": "sum"
    }).reset_index()
    
    if actual_stats.empty:
        return pd.DataFrame(columns=["영업가족", "M스캔율_목표", "대상건_목표", "배분사유", "특이사항"])
    
    actual_stats["현재_실적율"] = safe_rate(actual_stats["M스캔건"], actual_stats["대상건"])
    
    p30 = actual_stats["대상건"].quantile(0.3)
    p70 = actual_stats["대상건"].quantile(0.7)
    max_vol = actual_stats["대상건"].max()
    
    def calc_target(row):
        volume = row["대상건"]
        current_rate = row["현재_실적율"]
        
        if volume < p30:
            base_adj = 1.15 + min(0.05, (p30 - volume) / max(p30, 1) * 0.05)
            group_label = "소규모(성장유도)"
        elif volume < p70:
            base_adj = 1.10
            group_label = "중규모(기준)"
        else:
            ratio = (volume - p70) / max(max_vol - p70, 1)
            base_adj = 1.05 + ratio * 0.03
            group_label = "대규모(현실유지)"
        
        target_rate = min(max(current_rate * base_adj, 30.0), 95.0)
        target_rate = round(target_rate, 1)
        
        vol_adj = 1.10 if volume < p30 else 1.05
        target_volume = int(row["대상건"] * vol_adj)
        
        reason = f"{group_label} | 기본+{int((base_adj-1)*100)}%"
        
        return pd.Series({
            "M스캔율_목표": target_rate,
            "대상건_목표": target_volume,
            "배분사유": reason,
            "특이사항": ""  # ✅ 기본값 명시
        })
    
    result = actual_stats.apply(calc_target, axis=1)
    final_df = pd.concat([actual_stats[["영업가족", "대상건", "M스캔건", "현재_실적율"]], result], axis=1)
    
    # ✅ KeyError 해결: 기존 데이터와 병합 시 '특이사항' 컬럼 안전하게 처리
    if not df_existing.empty and "영업가족" in df_existing.columns:
        if "특이사항" in df_existing.columns:
            final_df = final_df.merge(df_existing[["영업가족", "특이사항"]], on="영업가족", how="left")
            final_df["특이사항"] = final_df["특이사항"].fillna("")
        else:
            final_df["특이사항"] = ""
    else:
        final_df["특이사항"] = ""
    
    return final_df[["영업가족", "M스캔율_목표", "대상건_목표", "배분사유", "특이사항"]]

# ==========================================
# 5. 목표대비 현황 시뮬레이션 (중복 그래프 제거)
# ==========================================
def show_target_comparison(df_targets, df_actual, sel_months):
    """목표 vs 실제 통합 차트 (월별/누적)"""
    if df_targets.empty or df_actual.empty:
        st.info("데이터가 없습니다.")
        return
    
    actual_stats = build_org_stats(df_actual, sel_months, ["영업가족"], "누적")
    merged = actual_stats.merge(df_targets, on="영업가족", how="left")
    merged["M스캔율_목표"] = merged["M스캔율_목표"].fillna(50.0)
    merged["대상건_목표"] = merged["대상건_목표"].fillna(100)
    merged["M스캔율_달성율"] = ((merged["M스캔율_대상"] / merged["M스캔율_목표"].replace(0, float('nan'))) * 100).round(1).fillna(0)
    
    # KPI
    col1, col2, col3 = st.columns(3)
    with col1:
        avg_ach = merged["M스캔율_달성율"].mean()
        st.metric("📈 평균 달성율", f"{avg_ach:.1f}%")
    with col2:
        achieved = len(merged[merged["M스캔율_달성율"] >= 100])
        st.metric("✅ 목표달성 조직", f"{achieved}개")
    with col3:
        total_orgs = len(merged)
        st.metric("📊 전체 조직", f"{total_orgs}개")
    
    st.divider()
    
    # 통합 차트 (목표 vs 실제)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=merged["영업가족"],
        y=merged["M스캔율_대상"],
        name="실제 스캔율",
        marker_color="#3498DB",
        text=[f"{v:.1f}%" for v in merged["M스캔율_대상"]],
        textposition="outside"
    ))
    fig.add_trace(go.Scatter(
        x=merged["영업가족"],
        y=merged["M스캔율_목표"],
        name="목표 스캔율",
        mode="lines+markers",
        line=dict(color="#E74C3C", width=3),
        marker=dict(size=8),
        text=[f"{v:.1f}%" for v in merged["M스캔율_목표"]],
        textposition="top center"
    ))
    fig.update_layout(
        title="영업가족별 목표 vs 실제 스캔율",
        xaxis_tickangle=-45,
        yaxis_title="M스캔율(%)",
        height=400,
        legend=dict(orientation="h", y=1.15)
    )
    st.plotly_chart(fig, use_container_width=True)
    
    # 달성율 분포
    fig2 = px.histogram(
        merged, x="M스캔율_달성율", nbins=20,
        color_discrete_sequence=["#3498DB"],
        title="목표 달성율 분포",
        labels={"M스캔율_달성율": "달성율(%)"}
    )
    fig2.add_vline(x=100, line_dash="dash", line_color="red", annotation_text="목표")
    fig2.update_layout(height=300, xaxis_range=[0,200])
    st.plotly_chart(fig2, use_container_width=True)

# ==========================================
# 6. UI – 로그인 & 대시보드
# ==========================================
def login_page():
    st.title("🔐 시스템 접속")
    pwd = st.text_input("접속 비밀번호를 입력하세요", type="password")
    if st.button("접속하기", use_container_width=True, type="primary"):
        if pwd == APP_PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")

def dashboard_page():
    st.title("📱 M스캔 전용 서류 처리 현황 대시보드")
    
    with st.expander("📌 책임판매 필수서류 4종 & 모바일동의 집중 관리 안내 (2026.05 시행)", expanded=False):
        st.markdown("✅ **체결 전 완비 원칙** : 4종 중 단 1개라도 미완비 시 불완전판매 및 리스크 계약 간주\n"
                    "📱 **모바일동의 표준화** : 자동매칭·타임스탬프·누락방지 기능으로 업무 효율과 법적 증빙력 확보\n"
                    "⚠️ **사후징구 금지** : 2026년 5월부터 서류 미비 시 내부 통제 미충족 조직으로 관리")

    df = load_data()
    if df.empty:
        st.warning("데이터가 없습니다. insurance_data.xlsx 파일을 확인해주세요.")
        return

    col1, col2 = st.columns([2, 1])
    with col1: st.success(f"총 {len(df):,}건의 데이터 로드 완료")
    with col2: st.info(f"기준: {get_file_update_time()}")

    month_col = "월_피리어드"
    all_months = sorted(df[month_col].dropna().unique())
    st.subheader("분석 기간 선택")
    sel_months = st.multiselect("월 선택", all_months, default=[all_months[-1]] if all_months else [])
    if not sel_months: st.warning("최소 1개 이상의 월을 선택해주세요."); return

    df_sel = df[df[month_col].isin(sel_months)].copy()
    if df_sel.empty: st.info("선택한 기간에 데이터가 없습니다."); return

    total_docs = int(df_sel["대상건"].sum())
    total_scanned = int(df_sel["전체스캔건"].sum())
    m_scanned = int(df_sel["M스캔건"].sum())
    avg_rate_target = safe_rate(pd.Series([m_scanned]), pd.Series([total_docs])).iloc[0]
    avg_rate_scan = safe_rate(pd.Series([m_scanned]), pd.Series([total_scanned])).iloc[0]

    st.divider()

    tab_dash, tab_target, tab_map, tab_guide, tab_manual = st.tabs([
        "📊 현황 대시보드", "🎯 목표관리 & 공문출력", "🗺️ M스캔 활용 현황", 
        "📱 가이드 & 프로세스", "📚 매뉴얼 다운로드"
    ])

    # ==========================================
    # 탭 1: 현황 대시보드
    # ==========================================
    with tab_dash:
        ctrl1, ctrl2, ctrl3, ctrl4 = st.columns([1, 1, 1, 1])
        with ctrl1:
            agg_group = st.selectbox("집계 기준", ["부문", "총괄", "부서", "영업가족"], key="agg_group")
        with ctrl2:
            view_mode_ui = st.radio("보기 방식", ["누적 통합", "월별 비교"], horizontal=True, key="view_mode")
            view_mode = "월별" if view_mode_ui == "월별 비교" else "누적"
        with ctrl3:
            min_target = st.number_input("🔻 최소 대상건수 필터", min_value=0, step=10, value=10, key="min_target_dash")
        with ctrl4:
            # ✅ 하위 표시 개수로 변경
            bottom_n_filter = st.number_input("🔽 하위 표시 개수", min_value=5, step=5, value=20, key="bottom_n_dash")

        hierarchy = ["부문", "총괄", "부서", "영업가족"]
        idx = hierarchy.index(agg_group) + 1
        group_cols = hierarchy[:idx]

        agg = build_org_stats(df_sel, sel_months, group_cols, view_mode)

        rate_type = st.radio("📊 지표 선", ["M스캔율 (대상대비)", "M스캔율 (완료대비)"], horizontal=True, index=0, key="rate_type")
        compare_type = st.radio("🎯 비교 기준 선택", ["전사 평균대비", "목표치(+10%) 대비"], horizontal=True, index=0, key="compare_type")

        is_target = "대상대비" in rate_type
        rate_col = "M스캔율_대상" if is_target else "M스캔율_완료"
        avg_val = avg_rate_target if is_target else avg_rate_scan
        target_val = round(avg_val * 1.1, 1)
        baseline_val = avg_val if "평균" in compare_type else target_val
        baseline_label = "전사 평균" if "평균" in compare_type else "목표치(+10%)"

        if min_target > 0: agg = agg[agg["대상건"] >= min_target].copy()
        
        # ✅ 하위 N개 보기: 오름차순 정렬 후 하위 추출
        agg = agg.sort_values(rate_col, ascending=True).reset_index(drop=True)
        
        if view_mode == "월별":
            agg["순위"] = agg.groupby("월").cumcount() + 1
            if bottom_n_filter: agg = agg[agg["순위"] <= bottom_n_filter].reset_index(drop=True)
        else:
            agg = agg.head(bottom_n_filter).reset_index(drop=True)
            agg["순위"] = range(1, len(agg) + 1)

        agg["기준치"] = baseline_val
        agg["대비_격차"] = (agg[rate_col] - baseline_val).round(1)

        st.markdown(f"### 📈 **{rate_type}** 현황 (하위 {bottom_n_filter}개) | 비교 기준: **{baseline_label}** ({baseline_val:.1f}%)")
        met1, met2, met3, met4 = st.columns(4)
        with met1: st.metric("🏢 전사 평균", f"{avg_val:.1f}%")
        with met2: st.metric("🎯 목표치 (+10%)", f"{target_val:.1f}%")
        with met3: st.metric("📊 선택 조직 평균", f"{agg[rate_col].mean():.1f}%")
        with met4: 
            delta = agg[rate_col].mean() - baseline_val
            st.metric("📐 기준 대비 격차", f"{delta:+.1f}%", delta=round(delta, 1))
        
        st.divider()

        display_cols = ["순위"]
        if view_mode == "월별": display_cols.append("월_표시")
        display_cols.extend(group_cols + ["대상건", "전체스캔건", "M스캔건", rate_col, "기준치", "대비_격차"])
        display_cols = [c for c in display_cols if c in agg.columns]

        st.dataframe(
            agg[display_cols].style.format({
                "대상건":"{:,}", "전체스캔건":"{:,}", "M스캔건":"{:,}",
                rate_col:"{:.1f}%", "기준치":"{:.1f}%", "대비_격차":"{:+.1f}%"
            })
            .highlight_max(subset=[rate_col], color="#d1f2eb")
            .highlight_min(subset=["대비_격차"], color="#fadbd8"),
            use_container_width=True, hide_index=True, height=350
        )

        # 차트 렌더링
        top = agg
        if view_mode == "월별":
            months_order = sorted(top["월_표시"].dropna().unique())
            fig = px.bar(
                top, x=agg_group, y=rate_col, color="월_표시", barmode="group",
                title=f"월별 {rate_type} 비교 (조직별)", text_auto=".1f%",
                category_orders={"월_표시": months_order},
                hover_data={**{col: True for col in group_cols}, "대상건":":,", "전체스캔건":":,", "M스캔건":":,"}
            )
            y_max_bar = max(top[rate_col].max() * 1.2, 5)
            fig.update_layout(yaxis_range=[0, y_max_bar], yaxis_title="M스캔율(%)", legend_title="월", xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
            
            st.subheader("📉 조직별 월간 추이 분석")
            all_orgs = sorted(agg[agg_group].unique())
            trend_orgs = st.multiselect("추이 분석할 조직 선택 (검색 가능)", all_orgs, default=all_orgs[:5], max_selections=10, key="trend_org_select")
            
            show_data_labels = st.checkbox("📊 데이터 라벨 표시", value=True, key="show_trend_labels")
            
            if trend_orgs:
                trend_data = agg[agg[agg_group].isin(trend_orgs)].copy()
                max_val = trend_data[rate_col].max()
                y_upper = max(max_val * 1.25, 5)
                if max_val < 10: y_upper = min(y_upper, 20)
                elif max_val < 30: y_upper = min(y_upper, 40)
                else: y_upper = 100

                fig_line = go.Figure()
                months_sorted = sorted(trend_data["월_표시"].dropna().unique())
                for org in trend_orgs:
                    org_data = trend_data[trend_data[agg_group] == org].sort_values("월")
                    hover_text = []
                    for _, row in org_data.iterrows():
                        hover_info = f"월: {row['월_표시']}<br>"
                        for col in group_cols:
                            if col in row: hover_info += f"{col}: {row[col]}<br>"
                        hover_info += f"M스캔율: {row[rate_col]:.1f}%<br>대상건: {row['대상건']:,}"
                        hover_text.append(hover_info)
                    
                    # ✅ 데이터 라벨 조건부 표시 수정
                    text_vals = [f"{v:.1f}%" for v in org_data[rate_col]] if show_data_labels else None
                    
                    fig_line.add_trace(go.Scatter(
                        x=org_data["월_표시"], y=org_data[rate_col],
                        name=org, mode="lines+markers", 
                        text=text_vals, 
                        textposition="top center" if show_data_labels else None,
                        hovertext=hover_text, hoverinfo="text"
                    ))
                fig_line.add_hline(y=baseline_val, line_dash="dash", line_color="red", line_width=2,
                                   annotation_text=f"{baseline_label} {baseline_val:.1f}%")
                fig_line.update_layout(
                    title="조직별 월간 추이",
                    yaxis_range=[0, y_upper], yaxis_title="M스캔율(%)",
                    xaxis=dict(title="월", tickangle=-45, categoryorder="array", categoryarray=months_sorted),
                    legend=dict(orientation="h", y=1.15)
                )
                st.plotly_chart(fig_line, use_container_width=True)
        else:
            fig = go.Figure()
            hover_texts = []
            for _, row in top.iterrows():
                hover_info = f"조직: {row['표시명']}<br>"
                for col in group_cols:
                    if col in row: hover_info += f"{col}: {row[col]}<br>"
                hover_info += f"M스캔율: {row[rate_col]:.1f}%<br>대상건: {row['대상건']:,}<br>전체스캔건: {row['전체스캔건']:,}<br>M스캔건: {row['M스캔건']:,}"
                hover_texts.append(hover_info)
            
            fig.add_trace(go.Bar(x=top["표시명"], y=top[rate_col], text=[f"{v:.1f}%" for v in top[rate_col]], 
                                 textposition="outside", marker_color=top[rate_col], hovertext=hover_texts, hoverinfo="text"))
            fig.add_hline(y=baseline_val, line_dash="dash", line_color="red", line_width=2, annotation_text=f"{baseline_label} {baseline_val:.1f}%")
            fig.update_layout(title=f"조직별 {rate_type} (정렬: 하위순)", xaxis_tickangle=-45, 
                              yaxis_range=[0, max(top[rate_col].max() * 1.2, 5)], yaxis_title="M스캔율(%)", height=420)
            st.plotly_chart(fig, use_container_width=True)

    # ==========================================
    # 탭 2: 목표관리 & 공문출력
    # ==========================================
    with tab_target:
        st.subheader("🎯 영업가족별 목표 설정 및 공문 출력")
        
        df_targets = load_targets()
        
        if "영업가족" in df_sel.columns:
            all_families = sorted(df_sel["영업가족"].dropna().unique())
        else:
            all_families = []
        
        st.markdown("### ① 목표 자동배분 (실적규모 기반 현실적 차등)")
        st.caption("소규모: +15~20% | 중규모: +10% | 대규모: +5~8% (건수 감안 현실적 목표)")
        
        if st.button("🔄 자동배분 실행", use_container_width=True, type="primary"):
            with st.spinner("🎯 실적규모 분석 및 목표 배분 중..."):
                # ✅ 오류 해결: 반환값 구조 안전하게 처리
                new_targets = auto_allocate_targets(df_sel, df_targets, increase_rate=0.10)
                
                if not new_targets.empty:
                    st.session_state["auto_targets"] = new_targets
                    st.success(f"✅ {len(new_targets)}개 영업가족 목표 자동배분 완료!")
                    st.rerun()
        
        if "auto_targets" in st.session_state:
            st.markdown("#### 📋 자동배분 결과 미리보기")
            preview_df = st.session_state["auto_targets"].copy()
            st.dataframe(
                preview_df[["영업가족", "M스캔율_목표", "대상건_목표", "배분사유"]]
                .style.format({"M스캔율_목표": "{:.1f}%", "대상건_목표": "{:,}"})
                .highlight_max(subset=["M스캔율_목표"], color="#d4edda")
                .highlight_min(subset=["M스캔율_목표"], color="#f8d7da"),
                use_container_width=True, height=250
            )
            
            if st.button("💾 이 결과로 목표 저장", key="save_auto"):
                save_targets(st.session_state["auto_targets"])
                st.success("✅ 목표가 저장되었습니다!")
                del st.session_state["auto_targets"]
                st.rerun()
        
        st.divider()
        
        st.markdown("### ② 목표 수정 (출력자가 영업가족 특성 반영)")
        st.caption("자동배분 또는 기존 목표를 수정할 수 있습니다. 수정 후 '저장'을 눌러주세요.")
        
        if df_targets.empty and all_families:
            df_targets = pd.DataFrame({
                "영업가족": all_families,
                "M스캔율_목표": 50.0,
                "대상건_목표": 100,
                "배분사유": "신규등록",
                "특이사항": ""
            })
        
        if not df_targets.empty:
            edited_df = st.data_editor(
                df_targets,
                column_config={
                    "영업가족": st.column_config.TextColumn("영업가족", disabled=True),
                    "M스캔율_목표": st.column_config.NumberColumn("M스캔율 목표(%)", min_value=0, max_value=100, step=0.5, format="%.1f%%"),
                    "대상건_목표": st.column_config.NumberColumn("대상건 목표", min_value=0, step=10, format="%d 건"),
                    "배분사유": st.column_config.TextColumn("배분사유", disabled=True),
                    "특이사항": st.column_config.TextColumn("특이사항 (공문 출력 시 포함)", max_chars=50)
                },
                hide_index=True,
                use_container_width=True,
                height=300,
                key="target_editor"
            )
            
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 수정사항 저장", use_container_width=True):
                    save_targets(edited_df)
                    st.success("✅ 목표가 저장되었습니다!")
                    st.rerun()
            with col2:
                if st.button("📊 달성율 시뮬레이션", use_container_width=True):
                    show_target_comparison(edited_df, df_sel, sel_months)
        
        st.divider()
        
        st.markdown("### ③ 목표대비 현황 그래프")
        st.caption("목표대비 현재 스캔율을 확인하세요.")
        
        if not df_targets.empty and "영업가족" in df_sel.columns:
            show_target_comparison(df_targets, df_sel, sel_months)
        
        st.divider()
        
        st.markdown("### ④ 공문형식 리포트 출력")
        st.markdown("목표관리 현황과 모바일동의 안내자료를 결합한 공문형식 리포트를 생성합니다.")
        
        col1, col2 = st.columns(2)
        with col1:
            report_title = st.text_input("문서 제목", "M스캔 목표관리 현황 및 모바일동의 독려 안내")
            report_dept = st.text_input("발급 부서", "영업지원팀")
        with col2:
            report_date = st.date_input("문서 일자", datetime.now())
            report_recipient = st.text_input("수신", "전 영업가족")
        
        if st.button("🔍 출력 전 최종 확인", use_container_width=True):
            with st.expander("📋 출력 내용 미리보기", expanded=True):
                st.markdown(f"**문서제목**: {report_title}")
                st.markdown(f"**발급부서**: {report_dept} | **일자**: {report_date.strftime('%Y년 %m월 %d일')}")
                st.markdown(f"**수신**: {report_recipient}")
                st.divider()
                st.markdown("📊 **포함될 목표 데이터**")
                if not df_targets.empty:
                    st.dataframe(df_targets.head(5)[["영업가족", "M스캔율_목표", "대상건_목표", "특이사항"]].style.format({
                        "M스캔율_목표": "{:.1f}%", "대상건_목표": "{:,}"
                    }), use_container_width=True, height=150)
                    st.caption(f"총 {len(df_targets)}개 영업가족 목표 포함")
                else:
                    st.warning("⚠️ 목표 데이터가 없습니다. 먼저 목표를 설정해주세요.")
        
        if st.button("🖨️ 공문형식 리포트 생성 (PDF)", use_container_width=True, type="primary"):
            with st.spinner("📄 리포트 생성 중..."):
                try:
                    buf = generate_official_report(
                        df_sel, df_targets, sel_months, 
                        report_title, report_dept, 
                        report_date.strftime("%Y년 %m월 %d일"), 
                        report_recipient
                    )
                    st.download_button(
                        label="📥 PDF 다운로드",
                        data=buf,
                        file_name=f"공문_{report_title}_{report_date.strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                    st.success("✅ 리포트가 생성되었습니다. 다운로드 버튼을 눌러주세요.")
                except Exception as e:
                    st.error(f"❌ 리포트 생성 오류: {e}")

    # ==========================================
    # 탭 3: M스캔 활용 현황
    # ==========================================
    with tab_map:
        st.subheader("조직별 M스캔 활용도 분포 (M스캔율 기준 내림차순)")
        mc1, mc2, mc3, mc4, mc5 = st.columns([1, 1, 1, 1, 1])
        with mc1: map_level = st.selectbox("집계 단위", ["부문", "총괄", "부서", "영업가족"], key="map_level")
        with mc2: map_type = st.radio("차트 유형", ["가로막대", "파이 차트"], horizontal=True, key="map_type")
        with mc3:
            map_min_target = st.number_input("🔻 최소 대상건수 필터", min_value=0, step=10, value=10, key="min_target_map")
        with mc4:
            map_top_n = st.number_input("🔝 상위 표시 개수", min_value=5, step=5, value=20, key="top_n_map")
        with mc5:
            map_compare = st.radio("🎯 비교 기준", ["전사 평균대비", "목표치(+10%) 대비"], horizontal=True, key="map_compare")

        hierarchy_cols = ['부문', '총괄', '부서', '영업가족']
        agg_dict = {col: 'first' for col in hierarchy_cols if col != map_level}
        agg_dict.update({'대상건': 'sum', '전체스캔건': 'sum', 'M스캔건': 'sum'})
        
        map_agg = df_sel.groupby(map_level).agg(agg_dict).reset_index()
        map_agg["M스캔율_대상"] = safe_rate(map_agg["M스캔건"], map_agg["대상건"])
        
        map_baseline = avg_rate_target if "평균" in map_compare else round(avg_rate_target * 1.1, 1)
        map_agg["격차"] = map_agg["M스캔율_대상"] - map_baseline
        
        if map_min_target > 0: map_agg = map_agg[map_agg["대상건"] >= map_min_target].copy()
        map_agg = map_agg[map_agg["M스캔건"] > 0].sort_values("M스캔율_대상", ascending=False).reset_index(drop=True)
        if map_top_n: map_agg = map_agg.head(map_top_n).reset_index(drop=True)
        
        if map_agg.empty:
            st.info("M스캔 활용 데이터가 없습니다.")
        else:
            if map_type == "가로막대":
                fig_bar = px.bar(map_agg, y=map_level, x="M스캔율_대상", orientation="h", color="격차", text_auto=".1f%",
                                 color_continuous_scale=["#FF4444", "#888888", "#44CC44"], 
                                 title=f"전사 평균/목표 대비 조직별 M스캔율 분포",
                                 hover_data=["부문", "총괄", "부서", "영업가족", "대상건", "전체스캔건", "M스캔건"])
                fig_bar.update_layout(height=600, xaxis_title="M스캔율 (대상대비 %)", yaxis=dict(autorange="reversed"))
                fig_bar.add_vline(x=map_baseline, line_dash="dash", line_color="black", annotation_text=f"기준선 {map_baseline:.1f}%")
                st.plotly_chart(fig_bar, use_container_width=True)
            else:
                fig_pie = px.pie(map_agg, values="M스캔건", names=map_level, title="조직별 M스캔 건수 비중", hole=0.4)
                fig_pie.update_traces(textposition="inside", textinfo="percent+label")
                st.plotly_chart(fig_pie, use_container_width=True)

    # ==========================================
    # 탭 4: 가이드 & 프로세스
    # ==========================================
    with tab_guide:
        st.subheader("🔄 모바일가입확인서 발송 및 결재 프로세스")
        for step in PROCESS_FLOW:
            with st.expander(f"🔹 Step {step['step']}: {step['title']}"):
                st.markdown(step["desc"])
                
        st.subheader("❓ 자주 묻는 질문(FAQ)")
        for q, a in MOBILE_GUIDE["faq"]: st.markdown(f"**Q. {q}**\n\nA. {a}")

        st.divider()
        st.subheader("📝 책임판매 필수 서류 4종 (체결 전 100% 완비 원칙)")
        doc_df = pd.DataFrame(GUIDANCE_DOCS[1:], columns=GUIDANCE_DOCS[0]).set_index("No.")
        st.dataframe(doc_df, use_container_width=True, hide_index=True, height=280)
        
        do_cols = st.columns(2)
        for i, item in enumerate(MOBILE_GUIDE["do_list"][:4]):
            do_cols[i%2].markdown(item)

    # ==========================================
    # 탭 5: 매뉴얼 다운로드
    # ==========================================
    with tab_manual:
        st.subheader("📚 모바일동의 매뉴얼 다운로드")
        st.divider()
        found = False
        for mf in MANUAL_FILES:
            if os.path.exists(mf):
                found = True
                try:
                    with open(mf, "rb") as f:
                        st.download_button(label=f"📥 {mf}", data=f.read(), file_name=mf, mime="application/pdf", key=f"dl_{mf}", use_container_width=True)
                except Exception as e:
                    st.error(f"❌ {mf} 읽기 오류: {e}")
        if not found:
            st.warning("⚠️ 매뉴얼 파일이 현재 폴더에 없습니다. 실행 디렉토리에 PDF 파일을 복사해주세요.")

    # ==========================================
    # 다운로드 버튼
    # ==========================================
    st.divider()
    st.subheader("📥 리포트 내보내기")
    if st.button("현황 데이터 Excel 다운로드", use_container_width=True):
        wb = Workbook(); ws = wb.active; ws.title = "M스캔 현황"
        export_cols = ["순위", "월_표시"] if view_mode == "월별" else ["순위"]
        export_cols.extend(group_cols + ["대상건", "전체스캔건", "M스캔건", "M스캔율_대상", "M스캔율_완료", "기준치", "대비_격차"])
        export_cols = [c for c in export_cols if c in agg.columns]
        
        ws.append(export_cols)
        for _, row in agg.iterrows():
            ws.append([row[c] for c in export_cols])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        st.download_button("Excel 저장", buf, f"M스캔_현황_{datetime.now().strftime('%Y%m%d')}.xlsx", 
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ==========================================
# 7. 공문형식 리포트 생성 함수
# ==========================================
def generate_official_report(df_sel, df_targets, sel_months, title, dept, date_str, recipient):
    """공문형식 리포트 PDF 생성"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io
    
    try:
        pdfmetrics.registerFont(TTFont('NotoSansKR', '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'))
        font_name = 'NotoSansKR'
    except:
        try:
            pdfmetrics.registerFont(TTFont('NanumGothic', '/usr/share/fonts/truetype/nanum/NanumGothic.ttf'))
            font_name = 'NanumGothic'
        except:
            font_name = 'Helvetica'
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Korean', fontName=font_name, fontSize=10))
    styles.add(ParagraphStyle(name='Title', fontName=font_name, fontSize=16, bold=True, alignment=1))
    styles.add(ParagraphStyle(name='Subtitle', fontName=font_name, fontSize=12, bold=True))
    styles.add(ParagraphStyle(name='Small', fontName=font_name, fontSize=8))
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 10*mm))
    
    header_data = [
        ['문서번호:', f'{dept}-{datetime.now().strftime("%Y%m%d")}-001'],
        ['발급일자:', date_str],
        ['수신:', recipient],
        ['발신:', dept]
    ]
    header_table = Table(header_data, colWidths=[40*mm, 80*mm])
    header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 10*mm))
    
    elements.append(Paragraph("【목적】", styles['Subtitle']))
    elements.append(Paragraph("본 문서는 영업가족별 M스캔 목표관리 현황을 공유하고, 모바일동의(M스캔)의 중요성과 활용 방안을 안내하기 위해 작성되었습니다.", styles['Korean']))
    elements.append(Spacer(1, 5*mm))
    
    elements.append(Paragraph("【목표관리 현황】", styles['Subtitle']))
    
    if not df_targets.empty and "영업가족" in df_sel.columns:
        actual_stats = build_org_stats(df_sel, sel_months, ["영업가족"], "누적")
        merged_df = actual_stats.merge(df_targets, on="영업가족", how="left")
        merged_df["M스캔율_목표"] = merged_df["M스캔율_목표"].fillna(0)
        merged_df["M스캔율_달성율"] = ((merged_df["M스캔율_대상"] / merged_df["M스캔율_목표"].replace(0, float('nan'))) * 100).round(1).fillna(0)
        merged_df = merged_df.sort_values("M스캔율_달성율", ascending=False)
        
        table_data = [['영업가족', '대상건', 'M스캔건', 'M스캔율(실제)', 'M스캔율(목표)', '달성율', '특이사항']]
        for _, row in merged_df.head(15).iterrows():
            table_data.append([
                row['영업가족'],
                f"{int(row['대상건']):,}",
                f"{int(row['M스캔건']):,}",
                f"{row['M스캔율_대상']:.1f}%",
                f"{row['M스캔율_목표']:.1f}%",
                f"{row['M스캔율_달성율']:.1f}%",
                row.get('특이사항', '')[:20] + "..." if len(row.get('특이사항', '')) > 20 else row.get('특이사항', '')
            ])
        
        table = Table(table_data, colWidths=[35*mm, 18*mm, 18*mm, 22*mm, 22*mm, 18*mm, 27*mm])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("목표 데이터가 없습니다.", styles['Small']))
    
    elements.append(Spacer(1, 10*mm))
    
    elements.append(Paragraph("【모바일동의(M스캔)의 중요성】", styles['Subtitle']))
    for reason in MOBILE_GUIDE["reasons"]:
        elements.append(Paragraph(f"• {reason}", styles['Korean']))
    
    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph("【필수 서류 4종】", styles['Subtitle']))
    elements.append(Paragraph("계약 체결 전 아래 4종 서류 100% 완비가 필수입니다.", styles['Korean']))
    
    doc_table_data = [['No.', '서류명', '법적근거']]
    for i, doc in enumerate(GUIDANCE_DOCS[1:], 1):
        doc_table_data.append([str(i), doc[1], doc[2].replace('\n', ' ')])
    
    doc_table = Table(doc_table_data, colWidths=[10*mm, 50*mm, 60*mm])
    doc_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E75B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(doc_table)
    
    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph("【요청사항】", styles['Subtitle']))
    elements.append(Paragraph("1. 설정된 목표를 반드시 달성하여 주시기 바랍니다.", styles['Korean']))
    elements.append(Paragraph("2. 모바일동의(M스캔)를 적극 활용하여 업무 효율성과 법적 증빙력을 확보해 주시기 바랍니다.", styles['Korean']))
    elements.append(Paragraph("3. 2026년 5월부터는 서류 미비 시 내부 통제 미충족 조직으로 관리되오니 각별한 주의 바랍니다.", styles['Korean']))
    
    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph(f"{dept}", styles['Korean']))
    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph(f"문서책임자: _________________ (인)", styles['Small']))
    
    doc.build(elements)
    buf.seek(0)
    return buf

# ==========================================
# 8. Main
# ==========================================
def main():
    if not st.session_state.get("logged_in"):
        login_page()
    else:
        with st.sidebar:
            st.success("👋 접속 완료")
            if st.button("🚪 로그아웃", use_container_width=True):
                st.session_state.logged_in = False
                st.rerun()
            st.divider()
            st.caption("v14.2 | 하위필터 | 라벨수정 | 목표저장오류해결 | 그래프통합 | 용어변경")
        dashboard_page()

if __name__ == "__main__":
    main()
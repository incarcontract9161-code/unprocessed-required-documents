import streamlit as st
import pandas as pd
import os
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, Image as RLImage, PageBreak, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import io

# ==========================================
# 0. 페이지 설정 (최상단 필수)
# ==========================================
st.set_page_config(page_title="보험 서류 스캔 관리 대시보드", layout="wide", page_icon="📊")

# ==========================================
# 1. 전역 설정 & 인증
# ==========================================
# GitHub에 업로드할 엑셀 파일명
EXCEL_FILE = "insurance_data.xlsx"

# Render 배포 시 환경변수(APP_PASSWORD) 사용, 로컬 테스트 시 기본값 적용
APP_PASSWORD = os.environ.get("APP_PASSWORD", "incar961")

GUIDANCE_TEXT = (
    "【책임판매 필수서류 안내】\n"
    "개인정보동의서, 비교설명확인서, 고지의무확인서, 완전판매확인서(대상계약 限)는 "
    "금융소비자보호법 및 보험업 감독규정에 따라 신계약 체결 전 구비가 요구되는 필수 서류입니다. "
    "상기 서류는 소비자 보호 및 설명 의무 이행 여부를 확인하기 위한 내부 통제 관리 대상 서류로서, 실적 확정 입력 마감 시점까지 제출 완료를 원칙으로 하며 미비 시 내부 통제 리스크 관리 대상 계약으로 분류됩니다."
)
PRECAUTION_TEXT_COVER = (
    "【미처리 시 유의사항】\n"
    "실적 확정 입력 마감 시점까지 필수 서류가 제출되지 않은 계약과 조직에 대하여는 모집질서 및 분쟁  리스크 관리 대상으로 분류되어 관리됩니다.\n"
    "내부 통제 기준 충족 시까지,  내부 심사 및 결재 과정에서 승인 여부가 제한 될 수 있습니다. (리스크, 신규 운영자금 등 기타 지원 신청)"
)
PRECAUTION_TEXT_CONFIRM = "영업가족별 미처리 현황 및 유의사항에 대하여 인지하였으며,"
PRECAUTION_TEXT_SHEET = "본인은 신계약 필수 서류의 사전 구비 의무 및 미제출 시 내부 통제 관리 기준이 적용될 수 있음을 확인합니다."
SIGNATURE_CONFIRMATION_TEXT = "신계약 필수 서류의 사전 구비 의무 및 미제출 시 내부 통제 관리 기준이 적용 사항을 영업가족에게 안내하였음을 확인합니다."

# 필수 서류 상세 안내 표 데이터
REQUIRED_DOCS_TABLE = [
    ["No.", "서류명", "법적 관리 근거 및 관련 내부 통제 기준", "목적 및 주요 내용"],
    ["1", "개인정보동의서", 
     "개인정보보호법 15조 및\n대리점 표준 내부통제기준 27조",
     "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에\n따른 개인정보 처리로 신계약시 필수 징구"],
    ["2", "비교설명확인서",
     "보험업감독규정\n별표 5-6",
     "유사 상품 3개 이상 비교·설명 이행\n사실 고객 확인 서명"],
    ["3", "고지의무확인서",
     "금융소비자보호법 26조와\n동법시행령 24조",
     "판매자 중요사항 고지의무 이행 확인,\n권한·책임·보상 관련 핵심 사항 고지,\n소비자 오인 예방"],
    ["4", "완전판매확인서\n(대상: 종신, CI, CEO정기, 고액)",
     "금융소비자보호법 제17·19조 설명 적합성 적정성 관련 조항\n영업지원기준안",
     "약관,청약서 부본 제공, 중요 상품 이해 및\n자발적 가입 확인, 설명 의무 이행 증빙력 확보"]
]

# ==========================================
# 2. 데이터 로딩 (GitHub 엑셀 기반)
# ==========================================
@st.cache_data(ttl=300)  # 5분마다 자동 갱신
def load_data():
    """GitHub의 엑셀 파일을 읽어서 DataFrame 반환"""
    if not os.path.exists(EXCEL_FILE):
        st.error(f"⚠️ '{EXCEL_FILE}' 파일이 없습니다. GitHub에 엑셀 파일을 업로드해주세요.")
        return pd.DataFrame()
    
    try:
        df = pd.read_excel(EXCEL_FILE)
        
        if df.empty:
            return pd.DataFrame()
        
        # 날짜 및 필드 전처리
        df["보험시작일_dt"] = pd.to_datetime(df["보험시작일"], errors="coerce")
        df["월_피리어드"] = df["보험시작일_dt"].dt.to_period("M").astype(str)
        df["FA고지_c"] = df["FA고지"].fillna("").astype(str).str.strip()
        df["비교설명_c"] = df["비교설명"].fillna("").astype(str).str.strip()
        df["완전판매_c"] = df["완전판매"].fillna("").astype(str).str.strip()
        
        # 스캔 여부 판정 (스캔, M스캔, 보험사스캔 = 스캔 처리)
        def is_scanned(val):
            if pd.isna(val) or val == "":
                return False
            val_str = str(val).strip()
            return val_str in ["스캔", "M스캔", "보험사스캔"]
        
        # 미스캔 여부 판정
        def is_not_scanned(val):
            if pd.isna(val) or val == "":
                return True
            val_str = str(val).strip()
            return val_str == "미스캔"
        
        # 완판 대상 여부 판정 (스캔, M스캔, 미스캔만 대상 - 해당없음 제외)
        def is_cs_target(val):
            if pd.isna(val) or val == "":
                return False
            val_str = str(val).strip()
            return val_str in ["스캔", "M스캔", "미스캔"]
        
        # FA고지/비교설명: 스캔 여부
        df["FA고지_스캔"] = df["FA고지_c"].apply(is_scanned).astype(int)
        df["FA고지_미스캔"] = df["FA고지_c"].apply(is_not_scanned).astype(int)
        
        df["비교설명_스캔"] = df["비교설명_c"].apply(is_scanned).astype(int)
        df["비교설명_미스캔"] = df["비교설명_c"].apply(is_not_scanned).astype(int)
        
        # 완판: 대상 여부 및 스캔 여부
        df["완판_대상"] = df["완전판매_c"].apply(is_cs_target).astype(int)
        df["완판_스캔"] = df["완전판매_c"].apply(is_scanned).astype(int)
        df["완판_미스캔"] = df["완전판매_c"].apply(is_not_scanned).astype(int)
        
        # 집계 기준은 "증권번호 수"가 아니라 "증권번호별 필수 서류 수"입니다.
        # 개인정보동의서는 항상 필수 제출 대상으로 간주되어 별도 집계에서 제외합니다.
        df["FA_target"] = 1
        df["비교_target"] = 1
        df["완판_target"] = df["완판_대상"]

        df["FA_scan"] = df["FA고지_스캔"]
        df["비교_scan"] = df["비교설명_스캔"]
        df["완판_scan"] = df["완판_스캔"]

        df["FA_miss"] = (df["FA_target"] - df["FA_scan"]).clip(lower=0).astype(int)
        df["비교_miss"] = (df["비교_target"] - df["비교_scan"]).clip(lower=0).astype(int)
        df["완판_miss"] = (df["완판_target"] - df["완판_scan"]).clip(lower=0).astype(int)

        df["대상건"] = df[["FA_target", "비교_target", "완판_target"]].sum(axis=1).astype(int)
        df["스캔건"] = df[["FA_scan", "비교_scan", "완판_scan"]].sum(axis=1).astype(int)
        df["미스캔"] = df[["FA_miss", "비교_miss", "완판_miss"]].sum(axis=1).astype(int)

        return df
    except Exception as e:
        st.error(f"❌ 엑셀 파일 읽기 오류: {e}")
        return pd.DataFrame()

def get_file_update_time():
    """엑셀 파일의 마지막 수정 시간 반환"""
    if os.path.exists(EXCEL_FILE):
        timestamp = os.path.getmtime(EXCEL_FILE)
        return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
    return "알 수 없음"

# ==========================================
# 3. 집계 헬퍼
# ==========================================
# Numeric miss flags are precomputed in load_data to speed aggregation.

def calculate_scan_stats(df_group):
    """
    증번 기준 스캔율 계산
    - FA고지/비교설명: 증번당 2배수 (개인정보 제외)
    - 완판: 대상건만 집계 (스캔, M스캔, 미스캔만 대상)
    - 스캔 처리: 스캔, M스캔, 보험사스캔
    """
    cnt = len(df_group)  # 증번 수
    
    # FA고지/비교설명만 필수 (증번당 2개씩 = 증번수 * 2)
    FA_대상 = cnt * 2
    비교_대상 = cnt * 2
    
    FA_스캔 = df_group["FA고지_스캔"].sum() * 2  # 증번당 2개 기준
    비교_스캔 = df_group["비교설명_스캔"].sum() * 2
    
    # 완판: 대상건만 (해당없음 제외)
    완판_대상 = df_group["완판_대상"].sum()
    완판_스캔 = df_group["완판_스캔"].sum()
    완판_미스캔 = df_group["완판_미스캔"].sum()
    
    # 전체 대상 및 스캔
    전체_대상 = FA_대상 + 비교_대상 + 완판_대상
    전체_스캔 = FA_스캔 + 비교_스캔 + 완판_스캔
    전체_미스캔 = 전체_대상 - 전체_스캔
    
    # 스캔율
    스캔율 = round((전체_스캔 / 전체_대상 * 100), 1) if 전체_대상 > 0 else 0.0
    
    return {
        "증번수": cnt,
        "FA_대상": FA_대상,
        "FA_스캔": FA_스캔,
        "FA_미스캔": FA_대상 - FA_스캔,
        "비교_대상": 비교_대상,
        "비교_스캔": 비교_스캔,
        "비교_미스캔": 비교_대상 - 비교_스캔,
        "완판_대상": 완판_대상,
        "완판_스캔": 완판_스캔,
        "완판_미스캔": 완판_미스캔,
        "전체_대상": 전체_대상,
        "전체_스캔": 전체_스캔,
        "전체_미스캔": 전체_미스캔,
        "스캔율": 스캔율
    }


# ==========================================
# 4. 전체 계층 리포트
# ==========================================
@st.cache_data(ttl=300)
def build_hierarchy_report(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    rows = []
    for bm, df_bm in src.groupby("부문"):
        fa = int(df_bm["FA_miss"].sum()); bi = int(df_bm["비교_miss"].sum()); cs = int(df_bm["완판_miss"].sum())
        tot = fa + bi + cs; cnt = int(df_bm["대상건"].sum()); scan = int(df_bm["스캔건"].sum())
        rows.append({"구분":"부문계", "부문":bm, "총괄":"", "부서":"", "영업가족":"",
                     "FA":fa, "비교":bi, "완판":cs, "총미스캔":tot, "대상건":cnt, "스캔건":scan,
                     "미처리율":round(tot/cnt*100,1) if cnt else 0.0, "스캔율":round(scan/cnt*100,1) if cnt else 0.0})
        for tg, df_tg in df_bm.groupby("총괄"):
            fa2 = int(df_tg["FA_miss"].sum()); bi2 = int(df_tg["비교_miss"].sum()); cs2 = int(df_tg["완판_miss"].sum())
            tot2 = fa2 + bi2 + cs2; cnt2 = int(df_tg["대상건"].sum()); scan2 = int(df_tg["스캔건"].sum())
            rows.append({"구분":"총괄계", "부문":bm, "총괄":tg, "부서":"", "영업가족":"",
                         "FA":fa2, "비교":bi2, "완판":cs2, "총미스캔":tot2, "대상건":cnt2, "스캔건":scan2,
                         "미처리율":round(tot2/cnt2*100,1) if cnt2 else 0.0, "스캔율":round(scan2/cnt2*100,1) if cnt2 else 0.0})
            for ds, df_ds in df_tg.groupby("부서"):
                fa3 = int(df_ds["FA_miss"].sum()); bi3 = int(df_ds["비교_miss"].sum()); cs3 = int(df_ds["완판_miss"].sum())
                tot3 = fa3 + bi3 + cs3; cnt3 = int(df_ds["대상건"].sum()); scan3 = int(df_ds["스캔건"].sum())
                rows.append({"구분":"부서계", "부문":bm, "총괄":tg, "부서":ds, "영업가족":"",
                             "FA":fa3, "비교":bi3, "완판":cs3, "총미스캔":tot3, "대상건":cnt3, "스캔건":scan3,
                             "미처리율":round(tot3/cnt3*100,1) if cnt3 else 0.0, "스캔율":round(scan3/cnt3*100,1) if cnt3 else 0.0})
                for fg, df_fg in df_ds.groupby("영업가족"):
                    fa4 = int(df_fg["FA_miss"].sum()); bi4 = int(df_fg["비교_miss"].sum()); cs4 = int(df_fg["완판_miss"].sum())
                    t4 = fa4 + bi4 + cs4; c4 = int(df_fg["대상건"].sum()); s4 = int(df_fg["스캔건"].sum())
                    rows.append({"구분":"영업가족", "부문":bm, "총괄":tg, "부서":ds, "영업가족":fg,
                                 "FA":fa4, "비교":bi4, "완판":cs4, "총미스캔":t4, "대상건":c4, "스캔건":s4,
                                 "미처리율":round(t4/c4*100,1) if c4 else 0.0, "스캔율":round(s4/c4*100,1) if c4 else 0.0})
    return pd.DataFrame(rows)

@st.cache_data(ttl=300)
def build_monthly_hierarchy(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    rows = []
    for mon, dm in src.groupby("월_피리어드"):
        for bm, db in dm.groupby("부문"):
            fa_b = int(db["FA_miss"].sum()); bi_b = int(db["비교_miss"].sum()); cs_b = int(db["완판_miss"].sum())
            target_b = int(db["대상건"].sum()); scan_b = int(db["스캔건"].sum())
            rows.append({"월":mon, "구분":"부문계", "부문":bm, "총괄":"", "부서":"",
                         "FA":fa_b, "비교":bi_b, "완판":cs_b, "총미스캔":fa_b+bi_b+cs_b,
                         "대상건":target_b, "스캔건":scan_b, "미처리율":round((fa_b+bi_b+cs_b)/target_b*100,1) if target_b else 0.0, "스캔율":round(scan_b/target_b*100,1) if target_b else 0.0})
            for tg, dt in db.groupby("총괄"):
                fa_t = int(dt["FA_miss"].sum()); bi_t = int(dt["비교_miss"].sum()); cs_t = int(dt["완판_miss"].sum())
                target_t = int(dt["대상건"].sum()); scan_t = int(dt["스캔건"].sum())
                rows.append({"월":mon, "구분":"총괄계", "부문":bm, "총괄":tg, "부서":"",
                             "FA":fa_t, "비교":bi_t, "완판":cs_t, "총미스캔":fa_t+bi_t+cs_t,
                             "대상건":target_t, "스캔건":scan_t, "미처리율":round((fa_t+bi_t+cs_t)/target_t*100,1) if target_t else 0.0, "스캔율":round(scan_t/target_t*100,1) if target_t else 0.0})
                for ds, dd in dt.groupby("부서"):
                    fa = int(dd["FA_miss"].sum()); bi = int(dd["비교_miss"].sum()); cs = int(dd["완판_miss"].sum())
                    tot = fa + bi + cs; cnt = int(dd["대상건"].sum()); scan = int(dd["스캔건"].sum())
                    rows.append({"월":mon, "구분":"부서계", "부문":bm, "총괄":tg, "부서":ds,
                                 "FA":fa, "비교":bi, "완판":cs, "총미스캔":tot,
                                 "대상건":cnt, "스캔건":scan, "미처리율":round(tot/cnt*100,1) if cnt else 0.0, "스캔율":round(scan/cnt*100,1) if cnt else 0.0})
    return pd.DataFrame(rows)

@st.cache_data(ttl=300)
def build_monthly_hierarchy_pivot(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    rows = []
    for mon, dm in src.groupby("월_피리어드"):
        for bm, db in dm.groupby("부문"):
            fa_b = int(db["FA_miss"].sum()); bi_b = int(db["비교_miss"].sum()); cs_b = int(db["완판_miss"].sum())
            target_b = int(db["대상건"].sum()); scan_b = int(db["스캔건"].sum())
            rows.append({"월":mon, "구분":"부문계", "부문":bm, "총괄":"", "부서":"",
                         "FA":fa_b, "비교":bi_b, "완판":cs_b, "총미스캔":fa_b+bi_b+cs_b,
                         "대상건":target_b, "스캔건":scan_b, "미처리율":round((fa_b+bi_b+cs_b)/target_b*100,1) if target_b else 0.0, "스캔율":round(scan_b/target_b*100,1) if target_b else 0.0})
            for tg, dt in db.groupby("총괄"):
                fa_t = int(dt["FA_miss"].sum()); bi_t = int(dt["비교_miss"].sum()); cs_t = int(dt["완판_miss"].sum())
                target_t = int(dt["대상건"].sum()); scan_t = int(dt["스캔건"].sum())
                rows.append({"월":mon, "구분":"총괄계", "부문":bm, "총괄":tg, "부서":"",
                             "FA":fa_t, "비교":bi_t, "완판":cs_t, "총미스캔":fa_t+bi_t+cs_t,
                             "대상건":target_t, "스캔건":scan_t, "미처리율":round((fa_t+bi_t+cs_t)/target_t*100,1) if target_t else 0.0, "스캔율":round(scan_t/target_t*100,1) if target_t else 0.0})
                for ds, dd in dt.groupby("부서"):
                    fa = int(dd["FA_miss"].sum()); bi = int(dd["비교_miss"].sum()); cs = int(dd["완판_miss"].sum())
                    tot = fa + bi + cs; cnt = int(dd["대상건"].sum()); scan = int(dd["스캔건"].sum())
                    rows.append({"월":mon, "구분":"부서계", "부문":bm, "총괄":tg, "부서":ds,
                                 "FA":fa, "비교":bi, "완판":cs, "총미스캔":tot,
                                 "대상건":cnt, "스캔건":scan, "미처리율":round(tot/cnt*100,1) if cnt else 0.0, "스캔율":round(scan/cnt*100,1) if cnt else 0.0})
    pivot_src = pd.DataFrame(rows)
    if pivot_src.empty:
        return pivot_src
    metrics = ["FA","비교","완판","총미스캔","대상건","스캔건","미처리율","스캔율"]
    pivot_frames = []
    month_order = sorted(src["월_피리어드"].dropna().unique())
    for metric in metrics:
        temp = pivot_src.pivot_table(index=["구분","부문","총괄","부서"], columns="월", values=metric, aggfunc="first")
        temp.columns = [f"{month}_{metric}" for month in temp.columns]
        pivot_frames.append(temp)
    pivot = pd.concat(pivot_frames, axis=1).reset_index()
    ordered_columns = ["구분","부문","총괄","부서"]
    for month in month_order:
        for metric in metrics:
            ordered_columns.append(f"{month}_{metric}")
    pivot = pivot[[c for c in ordered_columns if c in pivot.columns]]
    return pivot.fillna(0)

# ==========================================
# 5. 관리대장 선정 대상
# ==========================================
@st.cache_data(ttl=300)
def get_ledger_targets(df, months):
    src = df[df["월_피리어드"].isin(months)].copy()
    if src.empty: return {}
    agg = src.groupby(["부문", "총괄", "부서", "영업가족"]).agg(
        FA=("FA_miss", "sum"), 비교=("비교_miss", "sum"),
        완판=("완판_miss", "sum"), 대상=("대상건", "sum"), 스캔=("스캔건", "sum")
    ).reset_index()
    agg["총미스캔"] = agg[["FA", "비교", "완판"]].sum(axis=1)
    agg = agg[agg["총미스캔"] > 0]
    return {dept: grp for dept, grp in agg.groupby("부서")}

# ==========================================
# 6. 한글 폰트 및 스타일
# ==========================================
@st.cache_resource
def register_korean_font():
    """한글 폰트 등록 - Streamlit Cloud 환경 최적화"""
    
    # Streamlit Cloud에서 packages.txt로 설치한 폰트 경로들
    font_candidates = [
        ("NotoSansKR", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        ("NotoSansKR", "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
        # Windows (로컬 테스트용)
        ("Malgun", r"C:\Windows\Fonts\malgun.ttf"),
        # 기타 Linux 경로
        ("DejaVu", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    
    for name, path in font_candidates:
        try:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont(name, path))
                return name
        except Exception:
            continue
    
    # 폰트를 찾을 수 없을 때 경고
    st.warning("⚠️ 한글 폰트를 찾을 수 없습니다. PDF/Excel에서 한글이 깨질 수 있습니다. GitHub에 packages.txt 파일이 있는지 확인하세요.")
    return "Helvetica"

HDR_CLR, ALT_CLR, SUB_CLR = "#4472C4", "#EEF3FB", "#D9E1F2"

def _pdf_styles(fn):
    S = getSampleStyleSheet()
    def ps(n, **kw): return ParagraphStyle(n, parent=S["Normal"], fontName=fn, **kw)
    return {
        "title":   ps("T",  fontSize=15, bold=True, alignment=1, spaceAfter=4),
        "sub":     ps("S",  fontSize=10, spaceAfter=3),
        "body":    ps("B",  fontSize=8,  spaceAfter=2),
        "notice":  ps("N",  fontSize=7.5, spaceAfter=3, textColor=colors.HexColor("#CC0000"), alignment=1),
        "date":    ps("D",  fontSize=8,  alignment=2, spaceAfter=4),
        "section": ps("SC", fontSize=9,  bold=True,  spaceAfter=2),
    }

def _tbl(data, cw, fn, header_rows=1, sub_rows=None, align="CENTER"):
    if not data or len(data) < 1: return Spacer(1,0)
    cw_scaled = [w * 1.4 for w in cw]
    align_map = {"LEFT":0, "CENTER":1, "RIGHT":2}
    align_value = align_map.get(align.upper(), 1)
    S = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "tbl_cell",
        parent=S["Normal"],
        fontName=fn,
        fontSize=8,
        leading=10,
        alignment=align_value,
        wordWrap="CJK"
    )
    wrapped_data = [
        [Paragraph(str(cell), cell_style) if not isinstance(cell, Paragraph) else cell for cell in row]
        for row in data
    ]
    t = Table(wrapped_data, colWidths=cw_scaled, repeatRows=header_rows)
    cmds = [
        ("FONTNAME", (0,0),(-1,-1), fn), ("FONTSIZE", (0,0),(-1,-1), 8),
        ("ALIGN", (0,0),(-1,-1), align.upper()), ("VALIGN", (0,0),(-1,-1), "MIDDLE"),
        ("WORDWRAP", (0,0),(-1,-1), "CJK"),
        ("GRID", (0,0),(-1,-1), 0.4, colors.grey),
        ("LEFTPADDING", (0,0),(-1,-1), 4), ("RIGHTPADDING", (0,0),(-1,-1), 4),
        ("TOPPADDING", (0,0),(-1,-1), 4), ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ("BACKGROUND", (0,0),(-1,header_rows-1), colors.HexColor("#DCE6F1")),
        ("TEXTCOLOR", (0,0),(-1,header_rows-1), colors.HexColor("#1F3864")),
    ]
    for i in range(header_rows, len(data)):
        if (i-header_rows)%2==1: cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F3F6FA")))
        if sub_rows and i in sub_rows: cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#E9EEF8")))
    t.setStyle(TableStyle(cmds)); return t

def _fig_to_image(fig, max_width=1000, height=360):
    try:
        img_buf = io.BytesIO()
        fig.write_image(img_buf, format="png", width=max_width, height=height, scale=2)
        img_buf.seek(0)
        img = RLImage(img_buf)
        desired_width = min(max_width, 820)
        img.drawWidth = desired_width
        img.drawHeight = height * (desired_width / max_width)
        return img, None
    except Exception as e:
        return None, str(e)


def _append_pdf_figure(E, fig, st_, max_width=1000, height=360):
    img, err = _fig_to_image(fig, max_width=max_width, height=height)
    if img is not None:
        E.append(img)
        E.append(Spacer(1,10))
        return
    E.append(Paragraph(f"차트 이미지를 생성하지 못했습니다: {err}", st_["notice"]))
    E.append(Spacer(1,6))


def _sig_table(labels, fn, cw=120):
    t = Table([labels,["____________________"]*len(labels),["(인)"]*len(labels)], colWidths=[cw*1.4]*len(labels))
    t.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),("FONTNAME",(0,0),(-1,-1),fn),
                           ("FONTSIZE",(0,0),(-1,-1),8.5),("TOPPADDING",(0,0),(-1,-1),5),
                           ("BOTTOMPADDING",(0,0),(-1,-1),5),("BOX",(0,0),(-1,-1),0.5,colors.grey),
                           ("INNERGRID",(0,0),(-1,-1),0.3,colors.lightgrey)]))
    return t

# ==========================================
# 7. 전체 계층 리포트 Excel
# ==========================================

def report_excel(df, months):
    wb = Workbook()
    ws = wb.active
    ws.title = "\uacc4\uce35\ub9ac\ud3ec\ud2b8"

    tfn = "Malgun Gothic"
    hf = Font(name=tfn, size=9, bold=True, color="FFFFFF")
    bf = Font(name=tfn, size=9)
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    h_fill = PatternFill("solid", fgColor="4472C4")
    alt_fill = PatternFill("solid", fgColor="EEF3FB")
    fills = {
        "\ubd80\ubb38\uacc4": PatternFill("solid", fgColor="1F3864"),
        "\ucd1d\uad04\uacc4": PatternFill("solid", fgColor="2E75B6"),
        "\ubd80\uc11c\uacc4": PatternFill("solid", fgColor="D9E1F2"),
        "data_alt": alt_fill,
    }
    fonts_wc = {
        "\ubd80\ubb38\uacc4": Font(name=tfn, size=9, bold=True, color="FFFFFF"),
        "\ucd1d\uad04\uacc4": Font(name=tfn, size=9, bold=True, color="FFFFFF"),
        "\ubd80\uc11c\uacc4": Font(name=tfn, size=9, bold=True),
    }

    today = datetime.now().strftime("%Y-%m-%d")
    period_str = ", ".join(months) if months else "\uc804\uccb4"
    report = build_hierarchy_report(df, months)
    monthly = build_monthly_hierarchy(df, months)
    pivot = build_monthly_hierarchy_pivot(df, months)
    if report.empty:
        return io.BytesIO()

    headers = [
        "\uad6c\ubd84", "\ubd80\ubb38", "\ucd1d\uad04", "\ubd80\uc11c", "\uc601\uc5c5\uac00\uc871",
        "FA\uace0\uc9c0", "\ube44\uad50\uc124\uba85", "\uc644\uc804\ud310\ub9e4", "\ucd1d\ubbf8\uc2a4\uce94",
        "\ub300\uc0c1\uac74", "\uc2a4\uce94\uac74", "\ubbf8\ucc98\ub9ac\uc728", "\uc2a4\uce94\uc728",
    ]
    widths = [12, 14, 14, 14, 16, 10, 10, 10, 10, 10, 10, 10, 10]

    ws.merge_cells("A1:M1")
    ws["A1"] = f"\uacc4\uce35 \ub9ac\ud3ec\ud2b8 | \uae30\uac04: {period_str} | \uc0dd\uc131\uc77c: {today}"
    ws["A1"].font = Font(name=tfn, size=12, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    for ci, (header, width) in enumerate(zip(headers, widths), 1):
        c = ws.cell(2, ci, header)
        c.font = hf
        c.fill = h_fill
        c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    row_idx = 3
    for _, row in report.iterrows():
        gbn = row["\uad6c\ubd84"]
        values = [
            gbn, row["\ubd80\ubb38"], row["\ucd1d\uad04"], row["\ubd80\uc11c"], row["\uc601\uc5c5\uac00\uc871"],
            int(row["FA"]), int(row["\ube44\uad50"]), int(row["\uc644\ud310"]), int(row["\ucd1d\ubbf8\uc2a4\uce94"]),
            int(row["\ub300\uc0c1\uac74"]), int(row.get("\uc2a4\uce94\uac74", 0)), float(row["\ubbf8\ucc98\ub9ac\uc728"]), float(row.get("\uc2a4\uce94\uc728", 0.0)),
        ]
        fill = fills.get(gbn, fills["data_alt"] if row_idx % 2 == 0 else None)
        font = fonts_wc.get(gbn, bf)
        for ci, value in enumerate(values, 1):
            c = ws.cell(row_idx, ci, value)
            c.font = font
            c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            if ci in (12, 13):
                c.number_format = '0.0"%"'
            elif ci >= 6:
                c.number_format = '#,##0'
            if fill:
                c.fill = fill
        row_idx += 1

    if not monthly.empty:
        ws2 = wb.create_sheet("\uc6d4\ubcc4\uacc4\uce35")
        ws2.merge_cells("A1:M1")
        ws2["A1"] = f"\uc6d4\ubcc4 \uacc4\uce35 \ub9ac\ud3ec\ud2b8 | \uae30\uac04: {period_str} | \uc0dd\uc131\uc77c: {today}"
        ws2["A1"].font = Font(name=tfn, size=12, bold=True)
        ws2["A1"].alignment = Alignment(horizontal="center")
        headers2 = ["\uc6d4", "\uad6c\ubd84", "\ubd80\ubb38", "\ucd1d\uad04", "\ubd80\uc11c", "FA\uace0\uc9c0", "\ube44\uad50\uc124\uba85", "\uc644\uc804\ud310\ub9e4", "\ucd1d\ubbf8\uc2a4\uce94", "\ub300\uc0c1\uac74", "\uc2a4\uce94\uac74", "\ubbf8\ucc98\ub9ac\uc728", "\uc2a4\uce94\uc728"]
        widths2 = [12, 12, 14, 14, 14, 10, 10, 10, 10, 10, 10, 10, 10]
        for ci, (header, width) in enumerate(zip(headers2, widths2), 1):
            c = ws2.cell(2, ci, header)
            c.font = hf
            c.fill = h_fill
            c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws2.column_dimensions[get_column_letter(ci)].width = width
        row_idx = 3
        for _, row in monthly.iterrows():
            gbn = row["\uad6c\ubd84"]
            values = [
                row["\uc6d4"], gbn, row["\ubd80\ubb38"], row["\ucd1d\uad04"], row["\ubd80\uc11c"],
                int(row["FA"]), int(row["\ube44\uad50"]), int(row["\uc644\ud310"]), int(row["\ucd1d\ubbf8\uc2a4\uce94"]),
                int(row["\ub300\uc0c1\uac74"]), int(row.get("\uc2a4\uce94\uac74", 0)), float(row["\ubbf8\ucc98\ub9ac\uc728"]), float(row.get("\uc2a4\uce94\uc728", 0.0)),
            ]
            fill = fills.get(gbn, fills["data_alt"] if row_idx % 2 == 0 else None)
            font = fonts_wc.get(gbn, bf)
            for ci, value in enumerate(values, 1):
                c = ws2.cell(row_idx, ci, value)
                c.font = font
                c.border = bdr
                c.alignment = Alignment(horizontal="center", vertical="center")
                if ci in (12, 13):
                    c.number_format = '0.0"%"'
                elif ci >= 6:
                    c.number_format = '#,##0'
                if fill:
                    c.fill = fill
            row_idx += 1

    if not pivot.empty:
        ws3 = wb.create_sheet("\uc6d4\ubcc4\ud53c\ubc97")
        ws3.merge_cells(f"A1:{get_column_letter(len(pivot.columns))}1")
        ws3["A1"] = f"\uc6d4\ubcc4 \ud53c\ubc97 \ub9ac\ud3ec\ud2b8 | \uae30\uac04: {period_str} | \uc0dd\uc131\uc77c: {today}"
        ws3["A1"].font = Font(name=tfn, size=12, bold=True)
        ws3["A1"].alignment = Alignment(horizontal="center")
        headers3 = pivot.columns.tolist()
        for ci, header in enumerate(headers3, 1):
            c = ws3.cell(2, ci, header)
            c.font = hf
            c.fill = h_fill
            c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws3.column_dimensions[get_column_letter(ci)].width = 15 if ci > 4 else 18
        for ri, (_, pr) in enumerate(pivot.iterrows(), 3):
            for ci, header in enumerate(headers3, 1):
                value = pr[header]
                c = ws3.cell(ri, ci, value)
                c.font = bf
                c.border = bdr
                c.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(value, (int, float)):
                    if "\uc728" in str(header):
                        c.number_format = '0.0"%"'
                    else:
                        c.number_format = '#,##0'
                if ri % 2 == 0:
                    c.fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def report_pdf(df, months):
    fn = register_korean_font()
    st_ = _pdf_styles(fn)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    today = datetime.now().strftime("%Y-%m-%d")
    period_str = ", ".join(months) if months else "\uc804\uccb4"
    elements = [
        Paragraph("\uacc4\uce35 \ub9ac\ud3ec\ud2b8", st_["title"]),
        Paragraph(f"\uae30\uac04: {period_str} | \uc0dd\uc131\uc77c: {today}", st_["date"]),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor(HDR_CLR)),
        Spacer(1, 6),
    ]

    report = build_hierarchy_report(df, months)
    if not report.empty:
        elements.append(Paragraph("\uc804\uccb4 \uacc4\uce35 \uc694\uc57d", st_["section"]))
        hdr = [["\uad6c\ubd84", "\ubd80\ubb38", "\ucd1d\uad04", "\ubd80\uc11c", "\uc601\uc5c5\uac00\uc871", "FA\uace0\uc9c0", "\ube44\uad50\uc124\uba85", "\uc644\uc804\ud310\ub9e4", "\ucd1d\ubbf8\uc2a4\uce94", "\ub300\uc0c1\uac74", "\uc2a4\uce94\uac74", "\ubbf8\ucc98\ub9ac\uc728", "\uc2a4\uce94\uc728"]]
        rows = []
        sub_rows = []
        for i, (_, r) in enumerate(report.iterrows(), 1):
            rows.append([
                r["\uad6c\ubd84"], r["\ubd80\ubb38"], r["\ucd1d\uad04"], r["\ubd80\uc11c"], r["\uc601\uc5c5\uac00\uc871"],
                f"{int(r['FA']):,}", f"{int(r['\ube44\uad50']):,}", f"{int(r['\uc644\ud310']):,}", f"{int(r['\ucd1d\ubbf8\uc2a4\uce94']):,}",
                f"{int(r['\ub300\uc0c1\uac74']):,}", f"{int(r.get('\uc2a4\uce94\uac74', 0)):,}",
                f"{float(r['\ubbf8\ucc98\ub9ac\uc728']):.1f}%", f"{float(r.get('\uc2a4\uce94\uc728', 0.0)):.1f}%",
            ])
            if r["\uad6c\ubd84"] in ("\ubd80\ubb38\uacc4", "\ucd1d\uad04\uacc4", "\ubd80\uc11c\uacc4"):
                sub_rows.append(i)
        elements.append(_tbl(hdr + rows, [34, 42, 42, 42, 54, 34, 34, 34, 36, 34, 34, 34, 34], fn, sub_rows=sub_rows))
        elements.append(Spacer(1, 8))

    monthly = build_monthly_hierarchy(df, months)
    if not monthly.empty:
        elements.append(PageBreak())
        elements.append(Paragraph("\uc6d4\ubcc4 \uacc4\uce35 \uc694\uc57d", st_["section"]))
        hdr = [["\uc6d4", "\uad6c\ubd84", "\ubd80\ubb38", "\ucd1d\uad04", "\ubd80\uc11c", "FA\uace0\uc9c0", "\ube44\uad50\uc124\uba85", "\uc644\uc804\ud310\ub9e4", "\ucd1d\ubbf8\uc2a4\uce94", "\ub300\uc0c1\uac74", "\uc2a4\uce94\uac74", "\ubbf8\ucc98\ub9ac\uc728", "\uc2a4\uce94\uc728"]]
        rows = []
        sub_rows = []
        for i, (_, r) in enumerate(monthly.iterrows(), 1):
            rows.append([
                r["\uc6d4"], r["\uad6c\ubd84"], r["\ubd80\ubb38"], r["\ucd1d\uad04"], r["\ubd80\uc11c"],
                f"{int(r['FA']):,}", f"{int(r['\ube44\uad50']):,}", f"{int(r['\uc644\ud310']):,}", f"{int(r['\ucd1d\ubbf8\uc2a4\uce94']):,}",
                f"{int(r['\ub300\uc0c1\uac74']):,}", f"{int(r.get('\uc2a4\uce94\uac74', 0)):,}",
                f"{float(r['\ubbf8\ucc98\ub9ac\uc728']):.1f}%", f"{float(r.get('\uc2a4\uce94\uc728', 0.0)):.1f}%",
            ])
            if r["\uad6c\ubd84"] in ("\ubd80\ubb38\uacc4", "\ucd1d\uad04\uacc4", "\ubd80\uc11c\uacc4"):
                sub_rows.append(i)
        elements.append(_tbl(hdr + rows, [34, 34, 42, 42, 42, 34, 34, 34, 36, 34, 34, 34, 34], fn, sub_rows=sub_rows))
        elements.append(Spacer(1, 8))

    pivot = build_monthly_hierarchy_pivot(df, months)
    if not pivot.empty:
        elements.append(PageBreak())
        elements.append(Paragraph("\uc6d4\ubcc4 \ud53c\ubc97 \uc694\uc57d", st_["section"]))
        headers = pivot.columns.tolist()
        rows = []
        for _, pr in pivot.iterrows():
            row = []
            for col, value in zip(headers, pr.tolist()):
                if isinstance(value, (int, float)) and not pd.isna(value):
                    if "\uc728" in str(col):
                        row.append(f"{float(value):.1f}%")
                    else:
                        row.append(f"{int(value):,}")
                else:
                    row.append(str(value))
            rows.append(row)
        fixed = [24, 24, 28, 32]
        month_cols = max(1, len(headers) - 4)
        remaining = max(12, int((542 - sum(fixed)) / month_cols))
        widths = fixed + [remaining] * month_cols
        elements.append(_tbl([headers] + rows, widths, fn))

    doc.build(elements)
    buf.seek(0)
    return buf


def report_fullpage_pdf(df, months, agg_group, map_level, dash_doc_types=None, dash_chart_mode="group", dash_top_n=15, map_type="treemap"):
    fn = register_korean_font()
    st_ = _pdf_styles(fn)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)

    month_col = "\uc6d4_\ud53c\ub9ac\uc5b4\ub4dc"
    bi_miss_col = "\ube44\uad50_miss"
    cs_miss_col = "\uc644\ud310_miss"
    miss_col = "\ubbf8\uc2a4\uce94"
    target_col = "\ub300\uc0c1\uac74"
    scan_col = "\uc2a4\uce94\uac74"
    org_label = "\uc870\uc9c1"
    total_label = "\ucd1d \ubbf8\uc2a4\uce94"
    fa_label = "FA\uace0\uc9c0"
    bi_label = "\ube44\uad50\uc124\uba85"
    cs_label = "\uc644\uc804\ud310\ub9e4"

    today = datetime.now().strftime("%Y-%m-%d")
    period_str = ", ".join(months) if months else "\uc804\uccb4"
    elements = [
        Paragraph("\uc804\uccb4 \ud398\uc774\uc9c0 \ub9ac\ud3ec\ud2b8", st_["title"]),
        Paragraph(f"\uae30\uac04: {period_str} | \uc0dd\uc131\uc77c: {today}", st_["date"]),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor(HDR_CLR)),
        Spacer(1, 8),
    ]

    df_sel = df[df[month_col].isin(months)].copy() if months else df.copy()
    fa_t = int(df_sel["FA_miss"].sum())
    bi_t = int(df_sel[bi_miss_col].sum())
    cs_t = int(df_sel[cs_miss_col].sum())
    miss_total = int(df_sel[miss_col].sum())
    target_total = int(df_sel[target_col].sum())
    scan_total = int(df_sel[scan_col].sum())
    miss_rate = round(miss_total / target_total * 100, 1) if target_total else 0.0
    scan_rate = round(scan_total / target_total * 100, 1) if target_total else 0.0

    elements.append(Paragraph("\ud575\uc2ec KPI", st_["section"]))
    summary = [
        ["\uc870\ud68c \ub370\uc774\ud130", f"{len(df_sel):,}"],
        ["\ucd1d \ub300\uc0c1\uac74", f"{target_total:,}"],
        ["\ucd1d \uc2a4\uce94\uac74", f"{scan_total:,}"],
        ["\ucd1d \ubbf8\uc2a4\uce94", f"{miss_total:,}"],
        ["\ubbf8\ucc98\ub9ac\uc728 / \uc2a4\uce94\uc728", f"{miss_rate:.1f}% / {scan_rate:.1f}%"],
        ["FA\uace0\uc9c0 / \ube44\uad50\uc124\uba85 / \uc644\uc804\ud310\ub9e4", f"{fa_t:,} / {bi_t:,} / {cs_t:,}"],
    ]
    elements.append(_tbl(summary, [110, 180], fn, header_rows=0, align="LEFT"))
    elements.append(Spacer(1, 8))

    dash_doc_types = dash_doc_types or [total_label]
    dash_chart_mode = "group" if dash_chart_mode in ("group", "\uadf8\ub8f9\ud615") else "stack"
    map_type_norm = "pie" if map_type in ("pie", "\ud30c\uc774 \ucc28\ud2b8") else "treemap"

    agg = df_sel.groupby(agg_group).agg(
        fa_miss_sum=("FA_miss", "sum"),
        bi_miss_sum=(bi_miss_col, "sum"),
        cs_miss_sum=(cs_miss_col, "sum"),
        target_sum=(target_col, "sum"),
        scan_sum=(scan_col, "sum"),
    ).reset_index()
    agg["total_miss"] = agg[["fa_miss_sum", "bi_miss_sum", "cs_miss_sum"]].sum(axis=1)
    agg["miss_rate"] = ((agg["total_miss"] / agg["target_sum"].replace(0, pd.NA)) * 100).round(1).fillna(0.0)
    agg["scan_rate"] = ((agg["scan_sum"] / agg["target_sum"].replace(0, pd.NA)) * 100).round(1).fillna(0.0)
    agg = agg.rename(columns={agg_group: org_label}).sort_values("total_miss", ascending=False).head(dash_top_n)

    if not agg.empty:
        elements.append(Paragraph(f"\ub300\uc2dc\ubcf4\ub4dc \uc694\uc57d ({agg_group})", st_["section"]))
        hdr = [[org_label, "\ub300\uc0c1\uac74", "\uc2a4\uce94\uac74", total_label, "\ubbf8\ucc98\ub9ac\uc728", "\uc2a4\uce94\uc728", fa_label, bi_label, cs_label]]
        rows = [[
            r[org_label], f"{int(r['target_sum']):,}", f"{int(r['scan_sum']):,}", f"{int(r['total_miss']):,}",
            f"{float(r['miss_rate']):.1f}%", f"{float(r['scan_rate']):.1f}%", f"{int(r['fa_miss_sum']):,}", f"{int(r['bi_miss_sum']):,}", f"{int(r['cs_miss_sum']):,}"
        ] for _, r in agg.iterrows()]
        elements.append(_tbl(hdr + rows, [84, 48, 48, 56, 46, 46, 42, 42, 42], fn))
        elements.append(Spacer(1, 6))

        try:
            metric_map = {
                total_label: "total_miss",
                fa_label: "fa_miss_sum",
                bi_label: "bi_miss_sum",
                cs_label: "cs_miss_sum",
            }
            if len(dash_doc_types) == 1:
                metric_key = metric_map.get(dash_doc_types[0], "total_miss")
                # 내림차순 정렬 확보 (이미 sort_values로 정렬됨)
                fig_dash = px.bar(
                    agg,
                    x=org_label,
                    y=metric_key,
                    title=f"{dash_doc_types[0]} TOP {dash_top_n}",
                    text=metric_key,
                    color=metric_key,
                    color_continuous_scale="Reds" if metric_key == "total_miss" else "Blues",
                    labels={org_label: "조직", metric_key: "건수"}
                )
                fig_dash.update_traces(
                    texttemplate="%{text:,.0f}", 
                    textposition="outside",
                    hovertemplate="<b>%{x}</b><br>%{y:,.0f}건<extra></extra>"
                )
                fig_dash.update_layout(
                    title=f"{dash_doc_types[0]} TOP {dash_top_n} (내림차순)",
                    xaxis_tickangle=-45,
                    xaxis_title="조직",
                    yaxis_title="건수",
                    height=400,
                    margin=dict(l=60, r=60, t=80, b=120),
                    hovermode="x unified",
                    xaxis=dict(autorange=True),
                    showlegend=False
                )
            else:
                metric_cols = [metric_map.get(doc_type, "total_miss") for doc_type in dash_doc_types]
                chart_df = agg[[org_label] + metric_cols].copy()
                chart_df.columns = [org_label] + dash_doc_types
                chart_df = chart_df.melt(org_label, var_name="\uc11c\ub958", value_name="\uac74\uc218")
                fig_dash = px.bar(
                    chart_df,
                    x=org_label,
                    y="\uac74\uc218",
                    color="\uc11c\ub958",
                    barmode=dash_chart_mode,
                    title=f"\uc11c\ub958\ubcc4 \ubbf8\uc2a4\uce94 TOP {dash_top_n}",
                    text="\uac74\uc218",
                )
                fig_dash.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
            
            _append_pdf_figure(elements, fig_dash, st_, max_width=1000, height=340)
        except Exception as err:
            elements.append(Paragraph(f"\ub300\uc2dc\ubcf4\ub4dc \ucc28\ud2b8 \uc0dd\uc131 \uc624\ub958: {err}", st_["notice"]))
            elements.append(Spacer(1, 6))

        try:
            trend_fig = go.Figure()
            trend_fig.add_trace(
                go.Scatter(
                    x=agg[org_label],
                    y=agg["miss_rate"],
                    mode="lines+markers+text",
                    text=[f"{v:.1f}%" for v in agg["miss_rate"]],
                    textposition="top center",
                    line=dict(shape="spline", color="#C0392B", width=3),
                    marker=dict(size=7),
                    name="\ubbf8\ucc98\ub9ac\uc728",
                )
            )
            trend_fig.add_trace(
                go.Scatter(
                    x=agg[org_label],
                    y=agg["scan_rate"],
                    mode="lines+markers+text",
                    text=[f"{v:.1f}%" for v in agg["scan_rate"]],
                    textposition="bottom center",
                    line=dict(shape="spline", color="#1F618D", width=3),
                    marker=dict(size=7),
                    name="\uc2a4\uce94\uc728",
                )
            )
            trend_fig.update_layout(title=f"\ubbf8\ucc98\ub9ac\uc728 / \uc2a4\uce94\uc728 TOP {dash_top_n}", xaxis_tickangle=-35, height=320, margin=dict(l=20, r=20, t=60, b=20))
            _append_pdf_figure(elements, trend_fig, st_, max_width=1000, height=320)
        except Exception as err:
            elements.append(Paragraph(f"\ucd94\uc774 \ucc28\ud2b8 \uc0dd\uc131 \uc624\ub958: {err}", st_["notice"]))
            elements.append(Spacer(1, 6))

    map_agg = df_sel.groupby(map_level).agg(
        miss_sum=(miss_col, "sum"),
        target_sum=(target_col, "sum"),
        scan_sum=(scan_col, "sum"),
    ).reset_index()
    map_agg["miss_rate"] = ((map_agg["miss_sum"] / map_agg["target_sum"].replace(0, pd.NA)) * 100).round(1).fillna(0.0)
    map_agg["scan_rate"] = ((map_agg["scan_sum"] / map_agg["target_sum"].replace(0, pd.NA)) * 100).round(1).fillna(0.0)
    map_agg = map_agg.sort_values("miss_sum", ascending=False).head(dash_top_n)

    if not map_agg.empty:
        hdr = [[org_label, "\ub300\uc0c1\uac74", "\uc2a4\uce94\uac74", total_label, "\ubbf8\ucc98\ub9ac\uc728", "\uc2a4\uce94\uc728"]]
        rows = [[
            r[map_level], f"{int(r['target_sum']):,}", f"{int(r['scan_sum']):,}", f"{int(r['miss_sum']):,}", f"{float(r['miss_rate']):.1f}%", f"{float(r['scan_rate']):.1f}%"
        ] for _, r in map_agg.iterrows()]
        elements.append(Paragraph(f"{map_level}\ubcc4 \uc9d1\uacc4 \uc694\uc57d", st_["section"]))
        elements.append(_tbl(hdr + rows, [100, 55, 55, 58, 55, 55], fn))
        elements.append(Spacer(1, 6))

        try:
            if map_type_norm == "pie":
                fig_map = px.pie(
                    map_agg,
                    values="miss_sum",
                    names=map_level,
                    title=f"{map_level}\ubcc4 {total_label} \ube44\uc911",
                    hole=0.4,
                    color_discrete_sequence=px.colors.qualitative.Set3,
                )
                fig_map.update_traces(textposition="inside", textinfo="percent+label")
            else:
                fig_map = px.treemap(
                    map_agg,
                    path=[map_level],
                    values="miss_sum",
                    color="miss_rate",
                    title=f"{map_level}\ubcc4 {total_label} \ud2b8\ub9ac\ub9f5",
                    color_continuous_scale="RdYlGn_r",
                )
            fig_map.update_layout(
                margin=dict(l=60, r=60, t=80, b=60), 
                width=1000, 
                height=380,
                showlegend=True if map_type_norm == "treemap" else False
            )
            _append_pdf_figure(elements, fig_map, st_, max_width=1000, height=340)
        except Exception as err:
            elements.append(Paragraph(f"\uc9d1\uacc4 \ucc28\ud2b8 \uc0dd\uc131 \uc624\ub958: {err}", st_["notice"]))
            elements.append(Spacer(1, 6))

    pivot = build_monthly_hierarchy_pivot(df, months)
    if not pivot.empty:
        elements.append(PageBreak())
        elements.append(Paragraph("\uc6d4\ubcc4 \ud53c\ubc97 \uc694\uc57d", st_["section"]))
        headers = pivot.columns.tolist()
        rows = []
        for _, pr in pivot.iterrows():
            row = []
            for col, value in zip(headers, pr.tolist()):
                if isinstance(value, (int, float)) and not pd.isna(value):
                    if "\uc728" in str(col):
                        row.append(f"{float(value):.1f}%")
                    else:
                        row.append(f"{int(value):,}")
                else:
                    row.append(str(value))
            rows.append(row)
        fixed = [30, 30, 35, 45]
        month_cols = max(1, len(headers) - 4)
        remaining = max(12, int((542 - sum(fixed)) / month_cols))
        widths = fixed + [remaining] * month_cols
        elements.append(_tbl([headers] + rows, widths, fn))

    doc.build(elements)
    buf.seek(0)
    return buf


def ledger_pdf(families_by_dept, period_text, df_src):
    fn, st_, buf = register_korean_font(), _pdf_styles(register_korean_font()), io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=12*mm,leftMargin=12*mm, topMargin=15*mm,bottomMargin=15*mm)
    today = datetime.now().strftime("%Y년 %m월 %d일"); E = []
    center_date_style = ParagraphStyle("CenterDate", parent=st_["date"], alignment=1)
    title_left = ParagraphStyle("TitleLeft", parent=st_["title"], alignment=0)
    indent_style = ParagraphStyle("IndentSub", parent=st_["sub"], leftIndent=8, alignment=0, spaceAfter=2)
    date_left = ParagraphStyle("DateLeft", parent=st_["date"], alignment=0)
    date_indent = ParagraphStyle("DateIndent", parent=st_["date"], leftIndent=8, alignment=0)
    section_left = ParagraphStyle("SectionLeft", parent=st_["section"], alignment=0)
    notice_left = ParagraphStyle("NoticeLeft", parent=st_["notice"], leftIndent=8, alignment=0)
    for dept_name, grp_df in families_by_dept.items():
        sec, tg = grp_df.iloc[0]["부문"], grp_df.iloc[0]["총괄"]
        E += [Paragraph("신계약 필수서류 미처리 확인서", title_left), HRFlowable(width="100%",thickness=1.5,color=colors.HexColor(HDR_CLR)), Spacer(1,4),
              Paragraph(f"부서: {sec}  > {tg}  >  <b>{dept_name}</b>", indent_style), Paragraph(f"적용기간: {period_text}", date_indent), Spacer(1,6)]
        dept_src = df_src[df_src["부서"]==dept_name]
        E += [Paragraph("【필수 서류 상세 안내】", st_["section"]),
              _tbl(REQUIRED_DOCS_TABLE, [12, 60, 90, 198], fn, header_rows=1, align="LEFT"), Spacer(1,8),
              Paragraph(GUIDANCE_TEXT, notice_left), Spacer(1,8)]
        if not dept_src.empty:
            E.append(Paragraph("▶ 영업가족별 · 월별 · 양식별 미처리 현황", section_left))
            fam_mon = dept_src.groupby(["영업가족","월_피리어드"]).agg(FA=("FA_miss","sum"),비교=("비교_miss","sum"),완판=("완판_miss","sum")).reset_index()
            fam_mon["계"] = fam_mon[["FA","비교","완판"]].sum(axis=1); fam_mon = fam_mon[fam_mon["계"] > 0]
            if not fam_mon.empty:
                td=[["영업가족","월","FA고지","비교설명","완전판매","계"]]
                for _, r in fam_mon.iterrows(): td.append([
                r["영업가족"],
                r["월_피리어드"],
                f"{int(r.FA):,}",
                f"{int(r.비교):,}",
                f"{int(r.완판):,}",
                f"{int(r['계']):,}"
            ])
                E.append(_tbl(td,[130,50,45,45,45,45],fn, align="LEFT"))
                E.append(Spacer(1,4))
        E += [Paragraph(PRECAUTION_TEXT_COVER, notice_left), Spacer(1,4),
              Paragraph(PRECAUTION_TEXT_CONFIRM, notice_left), Spacer(1,4),
              Paragraph(SIGNATURE_CONFIRMATION_TEXT, notice_left), Spacer(1,8),
              Paragraph("작성일: _______________", center_date_style), Spacer(1,4),
              _sig_table(["부문장 확인","총괄 확인","부서장 확인"],fn,120), PageBreak()]
        for _, fam in grp_df.drop_duplicates("영업가족").iterrows():
            fam_name = fam["영업가족"]
            E += [Paragraph("신계약 필수서류 미처리 확인서", title_left), HRFlowable(width="100%",thickness=1.5,color=colors.HexColor(HDR_CLR)), Spacer(1,4),
                  Paragraph(f"소속: {sec}  > {tg}  > {dept_name}  >  <b>{fam_name}</b>", indent_style), Paragraph(f"적용기간: {period_text}", date_indent), Spacer(1,6)]
            fam_src = df_src[(df_src["영업가족"]==fam_name) & df_src["소속"].notna()]
            sosok = fam_src.groupby(["소속","월_피리어드"]).agg(FA=("FA_miss","sum"),비교=("비교_miss","sum"),완판=("완판_miss","sum")).reset_index()
            sosok["계"] = sosok[["FA","비교","완판"]].sum(axis=1); sosok = sosok[sosok["계"] > 0]
            E.append(Paragraph("▶ 소속별 · 월별 · 양식별 미처리 건수", section_left))
            if not sosok.empty:
                td2=[["소속","월","FA고지","비교설명","완전판매","계"]]
                for _,r in sosok.iterrows(): td2.append([
                    r["소속"],
                    r["월_피리어드"],
                    f"{int(r.FA):,}",
                    f"{int(r.비교):,}",
                    f"{int(r.완판):,}",
                    f"{int(r['계']):,}"
                ])
                E.append(_tbl(td2,[130,50,45,45,45,45],fn, align="LEFT"))
                E.append(Spacer(1,4))
            else:
                E.append(Paragraph("(해당 데이터 없음)", st_["body"]))
            E.append(Spacer(1,6))
            sum_d=[["FA고지","비교설명","완전판매","총계"],[f"{int(fam["FA"]):,}",f"{int(fam["비교"]):,}",f"{int(fam["완판"]):,}",f"{int(fam["총미스캔"]):,}"]]
            E += [Paragraph("▶ 양식별 미처리 요약", section_left), _tbl(sum_d,[90,90,90,90],fn, align="LEFT"), Spacer(1,8),
                  Paragraph("【필수 서류 상세 안내】", st_["section"]),
                  _tbl(REQUIRED_DOCS_TABLE, [12, 60, 90, 198], fn, header_rows=1, align="LEFT"), Spacer(1,8),
                  Paragraph(GUIDANCE_TEXT, notice_left), Spacer(1,8),
                  Paragraph(PRECAUTION_TEXT_COVER, notice_left), Spacer(1,4),
                  Paragraph(PRECAUTION_TEXT_SHEET, notice_left), Spacer(1,8),
                  Paragraph("작성일: _______________", center_date_style)]
            sig2=Table([[f"영업가족대표 서명: ____________________ (인)"]], colWidths=[120*1.4*3])
            sig2.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"LEFT"),("FONTNAME",(0,0),(-1,-1),fn),("FONTSIZE",(0,0),(-1,-1),9.5),
                                      ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),("BOX",(0,0),(-1,-1),0.5,colors.grey)]))
            E += [sig2, PageBreak()]
    doc.build(E); buf.seek(0); return buf

# ==========================================
# 10. 관리대장 Excel
# ==========================================
def ledger_excel(families_by_dept, period_text, df_src):
    wb = Workbook(); ws0 = wb.active; ws0.title="목차"
    tfn = "맑은 고딕"
    hf = Font(name=tfn,size=9,bold=True,color="FFFFFF")
    bf = Font(name=tfn,size=9)
    nf = Font(name=tfn,size=8,italic=True,color="CC0000")
    sig_f = Font(name=tfn,size=9,bold=True)
    bdr, h_fill, alt_fill = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin")), PatternFill("solid",fgColor="4472C4"), PatternFill("solid",fgColor="EEF3FB")
    today = datetime.now().strftime("%Y년 %m월 %d일")
    ws0.merge_cells("A1:F1"); ws0["A1"]=f"관리대장 목차  ·  {period_text}  ·  발급: {today}"
    ws0["A1"].font=Font(name=tfn,size=13,bold=True); ws0["A1"].alignment=Alignment(horizontal="center")
    for ci,h in enumerate(["부서","영업가족","FA고지","비교설명","완전판매","총미스캔"],1):
        c=ws0.cell(3,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center")
    ir=4
    for dept,grp in families_by_dept.items():
        for _,fam in grp.drop_duplicates("영업가족").iterrows():
            for ci,v in enumerate([dept,fam["영업가족"],int(fam["FA"]),int(fam["비교"]),int(fam["완판"]),int(fam["총미스캔"])],1):
                c=ws0.cell(ir,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                if isinstance(v,(int,float)): c.number_format = "#,##0"
                if ir%2==0: c.fill=alt_fill
            ir+=1
    for ci,w in enumerate([22,25,13,13,13,14],1): ws0.column_dimensions[get_column_letter(ci)].width=w
    for dept_name,grp_df in families_by_dept.items():
        sec, tg = grp_df.iloc[0]["부문"], grp_df.iloc[0]["총괄"]
        sname=f"표지_{dept_name[:10]}".replace("/","_"); ws_c=wb.create_sheet(title=sname)
        ws_c.merge_cells("A1:G1"); ws_c["A1"]=f"[{dept_name}]  신계약 필수서류 미처리 확인서"; ws_c["A1"].font=Font(name=tfn,size=14,bold=True)
        ws_c["A2"]=f"{sec}  > {tg}  > {dept_name}"; ws_c["A2"].font=Font(name=tfn,size=10)
        ws_c["A3"]=f"    적용기간: {period_text}"; ws_c["A3"].font=bf; ws_c["A3"].alignment=Alignment(horizontal="left")
        r=5
        ws_c.cell(r,1,"【필수 서류 상세 안내】").font=Font(name=tfn,size=10,bold=True); r+=1
        for ci, header in enumerate(["No.", "서류명", "법적 근거", "목적 및 주요 내용"], 1):
            c = ws_c.cell(r, ci, header)
            c.font = hf; c.fill = h_fill; c.border = bdr; c.alignment = Alignment(horizontal="center", vertical="center")
        ws_c.column_dimensions[get_column_letter(1)].width = 6
        ws_c.column_dimensions[get_column_letter(2)].width = 20
        ws_c.column_dimensions[get_column_letter(3)].width = 25
        ws_c.column_dimensions[get_column_letter(4)].width = 45
        r += 1
        docs_data = [
            ["1", "개인정보동의서", "개인정보보호법 15조 및\n대리점 표준 내부통제기준 27조", 
             "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에 따른 개인정보 처리로 신계약시 필수 징구"],
            ["2", "비교설명확인서", "보험업감독규정\n별표 5-6", 
             "유사 상품 3개 이상 비교·설명 이행 사실 고객 확인 서명"],
            ["3", "고지의무확인서", "금융소비자보호법 26조와\n동법시행령 24조", 
             "판매자 권한·책임·보상 관련 핵심 사항 고지, 소비자 소인 예방"],
            ["4", "완전판매확인서\n(대상: 종신, CI, CEO경기, 고액)", "금융소비자보호법 제17·19조\n영업지원기준안", 
             "약관,청약서 부본 제공, 중요 상품 이해 및 자발적 가입 확인, 설명 의무 이행 증빙력 확보"]
        ]
        for row_data in docs_data:
            for ci, val in enumerate(row_data, 1):
                c = ws_c.cell(r, ci, val)
                c.font = bf; c.border = bdr; c.alignment = Alignment(horizontal="left" if ci > 1 else "center", vertical="top", wrapText=True)
            ws_c.row_dimensions[r].height = 35
            r += 1
        r += 1
        ws_c.cell(r,1,GUIDANCE_TEXT).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=45; r+=2
        ws_c.cell(r,1,"▶ 영업가족별 · 월별 · 양식별 미처리 현황").font=Font(name=tfn,size=10,bold=True); r+=1
        dept_src=df_src[df_src["부서"]==dept_name]
        if not dept_src.empty:
            fam_mon=dept_src.groupby(["영업가족","월_피리어드"]).agg(FA=("FA_miss","sum"),비교=("비교_miss","sum"),완판=("완판_miss","sum")).reset_index()
            fam_mon["계"]=fam_mon[["FA","비교","완판"]].sum(axis=1); fam_mon=fam_mon[fam_mon["계"] > 0]
            hdrs, cws = ["영업가족","월","FA고지","비교설명","완전판매","계"], [25,20,13,13,13,13]
            for ci,(h,w) in enumerate(zip(hdrs,cws),1):
                c=ws_c.cell(r,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center"); ws_c.column_dimensions[get_column_letter(ci)].width=w
            for i,(_,rv) in enumerate(fam_mon.iterrows()):
                row_v=[rv["영업가족"],rv["월_피리어드"],int(rv.FA),int(rv.비교),int(rv.완판),int(rv["계"])]
                af=alt_fill if i%2==1 else None
                for ci,v in enumerate(row_v,1):
                    c=ws_c.cell(r+1+i,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                    if isinstance(v,(int,float)): c.number_format = "#,##0"
                    if af: c.fill=af
            r+=len(fam_mon)+2
        ws_c.cell(r,1,PRECAUTION_TEXT_COVER).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=35; r+=2
        ws_c.cell(r,1,PRECAUTION_TEXT_CONFIRM).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=35; r+=2

        ws_c.cell(r,1,"작성일: _______________").font=bf; r+=2
        for i,sig in enumerate(["부문장 확인","총괄 확인","부서장 확인"]):
            ws_c.cell(r,i*2+1,sig).font=sig_f
            ws_c.cell(r,i*2+2,"________________ (인)").font=Font(name=tfn,color="888888")

        for _,fam in grp_df.drop_duplicates("영업가족").iterrows():
            fam_name=fam["영업가족"]; fn_safe=fam_name[:14].replace("/","_").replace("  ","")
            ws_f=wb.create_sheet(title=fn_safe); ws_f.merge_cells("A1:G1"); ws_f["A1"]=f"[{fam_name}]  신계약 필수서류 미처리 확인서"; ws_f["A1"].font=Font(name=tfn,size=13,bold=True)
            ws_f["A2"]=f"{sec}  > {tg}  > {dept_name}  > {fam_name}"; ws_f["A2"].font=Font(name=tfn,size=9,italic=True)
            ws_f["A3"]=f"    적용기간: {period_text}"; ws_f["A3"].font=bf; ws_f["A3"].alignment=Alignment(horizontal="left")
            r_f=5; ws_f.cell(r_f,1,"▶ 소속별 · 월별 · 양식별 미처리 건수").font=Font(name=tfn,size=10,bold=True); r_f+=1
            fam_src=df_src[(df_src["영업가족"]==fam_name) & df_src["소속"].notna()]
            sosok=fam_src.groupby(["소속","월_피리어드"]).agg(FA=("FA_miss","sum"),비교=("비교_miss","sum"),완판=("완판_miss","sum")).reset_index()
            sosok["계"]=sosok[["FA","비교","완판"]].sum(axis=1); sosok=sosok[sosok["계"] > 0]
            sh, sc = ["소속","월","FA고지","비교설명","완전판매","계"], [25,20,13,13,13,13]
            for ci,(h,w) in enumerate(zip(sh,sc),1): c=ws_f.cell(r_f,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center"); ws_f.column_dimensions[get_column_letter(ci)].width=w
            if not sosok.empty:
                for i,(_,sr) in enumerate(sosok.iterrows()):
                    rv2=[sr["소속"],sr["월_피리어드"],int(sr.FA),int(sr.비교),int(sr.완판),int(sr["계"])]; af=alt_fill if i%2==1 else None
                    for ci,v in enumerate(rv2,1):
                        c=ws_f.cell(r_f+1+i,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                        if isinstance(v,(int,float)): c.number_format = "#,##0"
                        if af: c.fill=af
                r_f += len(sosok) + 2
            else:
                r_f += 1
            ws_f.cell(r_f,1,"▶ 양식별 요약").font=Font(name=tfn,size=10,bold=True); r_f+=1
            for ci,h in enumerate(["FA고지","비교설명","완전판매","총계"],1): c=ws_f.cell(r_f,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center")
            for ci,v in enumerate([int(fam["FA"]),int(fam["비교"]),int(fam["완판"]),int(fam["총미스캔"])],1):
                c=ws_f.cell(r_f+1,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                if isinstance(v,(int,float)): c.number_format = "#,##0"
            r_f+=3
            # 미처리 건수 표 하단 유의사항
            ws_f.cell(r_f,1,PRECAUTION_TEXT_COVER).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=30; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_SHEET).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=30; r_f+=2
            
            # 필수 서류 상세 안내 표 추가
            ws_f.cell(r_f,1,"【필수 서류 상세 안내】").font=Font(name=tfn,size=10,bold=True); r_f+=1
            # 표 헤더
            for ci, header in enumerate(["No.", "서류명", "법적 근거", "목적 및 주요 내용"], 1):
                c = ws_f.cell(r_f, ci, header)
                c.font = hf; c.fill = h_fill; c.border = bdr; c.alignment = Alignment(horizontal="center", vertical="center")
            ws_f.column_dimensions[get_column_letter(1)].width = 6
            ws_f.column_dimensions[get_column_letter(2)].width = 20
            ws_f.column_dimensions[get_column_letter(3)].width = 25
            ws_f.column_dimensions[get_column_letter(4)].width = 45
            r_f += 1
            
            # 표 데이터
            docs_data = [
                ["1", "개인정보동의서", "개인정보보호법 15조 및\n대리점 표준 내부통제기준 27조", 
                 "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에 따른 개인정보 처리로 신계약시 필수 징구"],
                ["2", "비교설명확인서", "보험업감독규정\n별표 5-6", 
                 "유사 상품 3개 이상 비교·설명 이행 사실 고객 확인 서명"],
                ["3", "고지의무확인서", "금융소비자보호법 26조와\n동법시행령 24조", 
                 "판매자 권한·책임·보상 관련 핵심 사항 고지, 소비자 소인 예방"],
                ["4", "완전판매확인서\n(대상: 종신, CI, CEO경기, 고액)", "금융소비자보호법 제17·19조\n영업지원기준안", 
                 "약관,청약서 부본 제공, 중요 상품 이해 및 자발적 가입 확인, 설명 의무 이행 증빙력 확보"]
            ]
            for row_data in docs_data:
                for ci, val in enumerate(row_data, 1):
                    c = ws_f.cell(r_f, ci, val)
                    c.font = bf; c.border = bdr; c.alignment = Alignment(horizontal="left" if ci > 1 else "center", vertical="top", wrapText=True)
                ws_f.row_dimensions[r_f].height = 35
                r_f += 1
            r_f += 1
            ws_f.cell(r_f,1,GUIDANCE_TEXT).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=45; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_COVER).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=35; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_SHEET).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=35; r_f+=2

            ws_f.cell(r_f,1,"작성일: _______________").font=bf; r_f+=1
            ws_f.cell(r_f,1,"영업가족대표 서명: ________________ (인)").font=sig_f
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ==========================================
# 11. UI – 단일 비밀번호 로그인
# ==========================================
def login_page():
    st.title("🔐 시스템 접속")
    st.markdown("**단일 비밀번호**로 접속합니다.")
    pwd = st.text_input("접속 비밀번호를 입력하세요", type="password")
    if st.button("접속하기", use_container_width=True, type="primary"):
        if pwd == APP_PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")

# ==========================================
# 12. 통합 대시보드
# ==========================================
def dashboard_page():
    st.title("\uc11c\ub958 \ucc98\ub9ac \ud604\ud669 \ub300\uc2dc\ubcf4\ub4dc")

    df = load_data()
    if df.empty:
        st.warning("\ub370\uc774\ud130\uac00 \uc5c6\uc2b5\ub2c8\ub2e4. insurance_data.xlsx \ud30c\uc77c\uc744 \ud655\uc778\ud574 \uc8fc\uc138\uc694.")
        return

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        st.success(f"\ucd1d {len(df):,}\uac74\uc758 \ub370\uc774\ud130 \ub85c\ub4dc \uc644\ub8cc")
    with col2:
        st.info(f"\uae30\uc900: {get_file_update_time()}")
    with col3:
        if st.button("\uc0c8\ub85c\uace0\uce68"):
            st.cache_data.clear()
            st.rerun()

    month_col = "\uc6d4_\ud53c\ub9ac\uc5b4\ub4dc"
    bi_miss_col = "\ube44\uad50_miss"
    cs_miss_col = "\uc644\ud310_miss"
    miss_col = "\ubbf8\uc2a4\uce94"
    target_col = "\ub300\uc0c1\uac74"
    scan_col = "\uc2a4\uce94\uac74"
    dept_col = "\ubd80\ubb38"
    tg_col = "\ucd1d\uad04"
    ds_col = "\ubd80\uc11c"
    fam_col = "\uc601\uc5c5\uac00\uc871"

    all_months = sorted(df[month_col].dropna().unique())
    st.subheader("\ubd84\uc11d \uae30\uac04 \uc120\ud0dd")
    sel_months = st.multiselect("\uc6d4 \uc120\ud0dd", all_months, default=[all_months[-1]] if all_months else [])
    if not sel_months:
        st.warning("\ucd5c\uc18c 1\uac1c \uc774\uc0c1\uc758 \uc6d4\uc744 \uc120\ud0dd\ud574 \uc8fc\uc138\uc694.")
        return

    period_text = f"{sel_months[0]} ~ {sel_months[-1]}" if len(sel_months) > 1 else sel_months[0]
    df_sel = df[df[month_col].isin(sel_months)].copy()
    if df_sel.empty:
        st.info("\uc120\ud0dd\ud55c \uae30\uac04\uc5d0 \ub370\uc774\ud130\uac00 \uc5c6\uc2b5\ub2c8\ub2e4.")
        return

    fa_t = int(df_sel["FA_miss"].sum())
    bi_t = int(df_sel[bi_miss_col].sum())
    cs_t = int(df_sel[cs_miss_col].sum())
    miss_total = int(df_sel[miss_col].sum())
    target_total = int(df_sel[target_col].sum())
    scan_total = int(df_sel[scan_col].sum())
    miss_rate = round(miss_total / target_total * 100, 1) if target_total else 0.0
    scan_rate = round(scan_total / target_total * 100, 1) if target_total else 0.0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("\ucd1d \uacc4\uc57d\uac74\uc218", f"{len(df_sel):,}\uac74")
    m2.metric("\ucd1d \ub300\uc0c1\uc11c\ub958\uc218", f"{target_total:,}\uac74")
    m3.metric("\ubbf8\ucc98\ub9ac\uc728", f"{miss_rate:.1f}%")
    m4.metric("\uc2a4\uce94\uc728", f"{scan_rate:.1f}%")
    st.caption(f"FA / \ube44\uad50 / \uc644\ud310 \ubbf8\uc2a4\uce94: {fa_t:,} / {bi_t:,} / {cs_t:,}")
    st.divider()

    tab_dash, tab_map, tab_report, tab_ledger = st.tabs(["\ud604\ud669 \ub300\uc2dc\ubcf4\ub4dc", "\ubbf8\ucc98\ub9ac\ub9f5", "\uacc4\uce35 \ub9ac\ud3ec\ud2b8", "\uad00\ub9ac\ub300\uc7a5 \ucd9c\ub825"])

    with tab_dash:
        cs1, cs2 = st.columns([2, 1])
        with cs1:
            search_text = st.text_input("\uc870\uc9c1 \uac80\uc0c9", placeholder="\uc870\uc9c1\uba85 \uc785\ub825")
        with cs2:
            agg_group = st.selectbox("\uc9d1\uacc4 \uae30\uc900", [dept_col, tg_col, ds_col, fam_col], key="agg_group")

        agg = df_sel.groupby(agg_group).agg(
            fa_miss_sum=("FA_miss", "sum"),
            bi_miss_sum=(bi_miss_col, "sum"),
            cs_miss_sum=(cs_miss_col, "sum"),
            target_sum=(target_col, "sum"),
            scan_sum=(scan_col, "sum"),
        ).reset_index()
        agg["total_miss"] = agg[["fa_miss_sum", "bi_miss_sum", "cs_miss_sum"]].sum(axis=1)
        agg["miss_rate"] = ((agg["total_miss"] / agg["target_sum"].replace(0, pd.NA)) * 100).round(1).fillna(0.0)
        agg["scan_rate"] = ((agg["scan_sum"] / agg["target_sum"].replace(0, pd.NA)) * 100).round(1).fillna(0.0)
        agg = agg.rename(columns={
            agg_group: "\uc870\uc9c1",
            "fa_miss_sum": "FA\uace0\uc9c0_\ubbf8\uc2a4\uce94",
            "bi_miss_sum": "\ube44\uad50\uc124\uba85_\ubbf8\uc2a4\uce94",
            "cs_miss_sum": "\uc644\uc804\ud310\ub9e4_\ubbf8\uc2a4\uce94",
            "target_sum": "\ub300\uc0c1\uac74",
            "scan_sum": "\uc2a4\uce94\uac74",
            "total_miss": "\ucd1d_\ubbf8\uc2a4\uce94",
            "miss_rate": "\ubbf8\ucc98\ub9ac\uc728",
            "scan_rate": "\uc2a4\uce94\uc728",
        })
        if search_text:
            agg = agg[agg["\uc870\uc9c1"].astype(str).str.contains(search_text, case=False, na=False)]
        agg = agg.sort_values("\ucd1d_\ubbf8\uc2a4\uce94", ascending=False).reset_index(drop=True)
        agg.insert(0, "\uc21c\uc704", range(1, len(agg) + 1))

        if agg.empty:
            st.info("\uc870\uac74\uc5d0 \ub9de\ub294 \ub370\uc774\ud130\uac00 \uc5c6\uc2b5\ub2c8\ub2e4.")
        else:
            st.dataframe(
                agg[["\uc21c\uc704", "\uc870\uc9c1", "\ub300\uc0c1\uac74", "\uc2a4\uce94\uac74", "\ucd1d_\ubbf8\uc2a4\uce94", "\ubbf8\ucc98\ub9ac\uc728", "\uc2a4\uce94\uc728", "FA\uace0\uc9c0_\ubbf8\uc2a4\uce94", "\ube44\uad50\uc124\uba85_\ubbf8\uc2a4\uce94", "\uc644\uc804\ud310\ub9e4_\ubbf8\uc2a4\uce94"]].style.format({
                    "\uc21c\uc704": "{:,}",
                    "\ub300\uc0c1\uac74": "{:,}",
                    "\uc2a4\uce94\uac74": "{:,}",
                    "\ucd1d_\ubbf8\uc2a4\uce94": "{:,}",
                    "\ubbf8\ucc98\ub9ac\uc728": "{:.1f}%",
                    "\uc2a4\uce94\uc728": "{:.1f}%",
                    "FA\uace0\uc9c0_\ubbf8\uc2a4\uce94": "{:,}",
                    "\ube44\uad50\uc124\uba85_\ubbf8\uc2a4\uce94": "{:,}",
                    "\uc644\uc804\ud310\ub9e4_\ubbf8\uc2a4\uce94": "{:,}",
                }),
                use_container_width=True,
                hide_index=True,
            )

            top_n = st.slider("\ucc28\ud2b8 \ud45c\uc2dc \uac1c\uc218", 5, 30, 20, key="dash_top_n")
            top = agg.head(top_n)
            c1, c2 = st.columns(2)
            with c1:
                doc_types = st.multiselect("\ud45c\uc2dc \uc11c\ub958", ["FA\uace0\uc9c0", "\ube44\uad50\uc124\uba85", "\uc644\uc804\ud310\ub9e4", "\ucd1d \ubbf8\uc2a4\uce94"], default=["\ucd1d \ubbf8\uc2a4\uce94"], key="dash_doc_types")
                if doc_types:
                    max_v = top["\ucd1d_\ubbf8\uc2a4\uce94"].max()
                    yr = [0, max_v * 1.2] if max_v > 0 else [0, 10]
                    if len(doc_types) == 1 and doc_types[0] == "\ucd1d \ubbf8\uc2a4\uce94":
                        fig = go.Figure()
                        fig.add_trace(go.Bar(
                            x=top["\uc870\uc9c1"],
                            y=top["\ucd1d_\ubbf8\uc2a4\uce94"],
                            text=top["\ucd1d_\ubbf8\uc2a4\uce94"],
                            textposition="outside",
                            texttemplate="%{text:,.0f}",
                            hovertemplate="%{x}<br>\uac74\uc218: %{y:,.0f}<extra></extra>",
                            marker_color=top["\ucd1d_\ubbf8\uc2a4\uce94"],
                            marker_colorscale="Reds",
                        ))
                    elif len(doc_types) == 1:
                        col_map = {
                            "FA\uace0\uc9c0": "FA\uace0\uc9c0_\ubbf8\uc2a4\uce94",
                            "\ube44\uad50\uc124\uba85": "\ube44\uad50\uc124\uba85_\ubbf8\uc2a4\uce94",
                            "\uc644\uc804\ud310\ub9e4": "\uc644\uc804\ud310\ub9e4_\ubbf8\uc2a4\uce94",
                        }
                        fig = px.bar(top, x="\uc870\uc9c1", y=col_map[doc_types[0]], title=f"{doc_types[0]} \ubbf8\uc2a4\uce94 TOP {top_n}", text=col_map[doc_types[0]], color=col_map[doc_types[0]], color_continuous_scale="Blues")
                        fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside", hovertemplate="%{x}<br>\uac74\uc218: %{y:,.0f}<extra></extra>")
                    else:
                        chart_mode = st.radio("\ucc28\ud2b8 \uc720\ud615", ["\uadf8\ub8f9\ud615", "\ub204\uc801\ud615"], horizontal=True, key="dash_chart_mode")
                        col_map = {
                            "FA\uace0\uc9c0": "FA\uace0\uc9c0_\ubbf8\uc2a4\uce94",
                            "\ube44\uad50\uc124\uba85": "\ube44\uad50\uc124\uba85_\ubbf8\uc2a4\uce94",
                            "\uc644\uc804\ud310\ub9e4": "\uc644\uc804\ud310\ub9e4_\ubbf8\uc2a4\uce94",
                            "\ucd1d \ubbf8\uc2a4\uce94": "\ucd1d_\ubbf8\uc2a4\uce94",
                        }
                        p = top[["\uc870\uc9c1"] + [col_map[d] for d in doc_types]].copy()
                        p.columns = ["\uc870\uc9c1"] + doc_types
                        p = p.melt("\uc870\uc9c1", var_name="\uc885\ub958", value_name="\uac74\uc218")
                        fig = px.bar(p, x="\uc870\uc9c1", y="\uac74\uc218", color="\uc885\ub958", text="\uac74\uc218", barmode="group" if chart_mode == "\uadf8\ub8f9\ud615" else "stack")
                        fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside", hovertemplate="%{x}<br>%{fullData.name}: %{y:,.0f}<extra></extra>")
                    fig.update_layout(xaxis_tickangle=-45, yaxis=dict(range=yr, tickformat=","), height=420)
                    st.plotly_chart(fig, use_container_width=True)
            with c2:
                max_v = top["\ucd1d_\ubbf8\uc2a4\uce94"].max()
                yr = [0, max_v * 1.2] if max_v > 0 else [0, 10]
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(
                    x=top["\uc870\uc9c1"],
                    y=top["\ucd1d_\ubbf8\uc2a4\uce94"],
                    mode="lines+markers+text",
                    text=[f"{int(v):,}" for v in top["\ucd1d_\ubbf8\uc2a4\uce94"]],
                    textposition="top center",
                    hovertemplate="%{x}<br>\uac74\uc218: %{y:,.0f}<extra></extra>",
                    line=dict(shape="spline", color="#CC0000"),
                    marker=dict(size=6),
                ))
                fig2.update_layout(title=f"\ubbf8\ucc98\ub9ac \uac74\uc218 \ucd94\uc774 TOP {top_n}", xaxis_tickangle=-45, yaxis=dict(range=yr, tickformat=","), height=420)
                st.plotly_chart(fig2, use_container_width=True)

    with tab_map:
        st.subheader("\ubbf8\ucc98\ub9ac \ubd84\ud3ec \uc2dc\uac01\ud654")
        mc1, mc2 = st.columns([1, 2])
        with mc1:
            map_level = st.selectbox("\uc9d1\uacc4 \ub2e8\uc704", [dept_col, tg_col, ds_col, fam_col], key="map_level")
        with mc2:
            map_type = st.radio("\ucc28\ud2b8 \uc720\ud615", ["\ud30c\uc774 \ucc28\ud2b8", "\ud2b8\ub9ac\ub9f5"], horizontal=True, key="map_type")

        map_agg = df_sel.groupby(map_level).agg(
            miss_sum=(miss_col, "sum"),
            target_sum=(target_col, "sum"),
            scan_sum=(scan_col, "sum"),
        ).reset_index()
        map_agg["miss_rate"] = ((map_agg["miss_sum"] / map_agg["target_sum"].replace(0, pd.NA)) * 100).round(1).fillna(0.0)
        map_agg["scan_rate"] = ((map_agg["scan_sum"] / map_agg["target_sum"].replace(0, pd.NA)) * 100).round(1).fillna(0.0)
        map_agg = map_agg.rename(columns={map_level: "\uc870\uc9c1"})
        map_agg = map_agg[map_agg["miss_sum"] > 0].sort_values("miss_sum", ascending=False)

        if map_agg.empty:
            st.info("\ubbf8\ucc98\ub9ac \uac74\uc218\uac00 \uc788\ub294 \ub370\uc774\ud130\uac00 \uc5c6\uc2b5\ub2c8\ub2e4.")
        else:
            if map_type == "\ud30c\uc774 \ucc28\ud2b8":
                fig_pie = px.pie(map_agg, values="miss_sum", names="\uc870\uc9c1", title=f"{map_level}\ubcc4 \ubbf8\uc2a4\uce94 \uac74\uc218 \ube44\uc911", hole=0.4, color_discrete_sequence=px.colors.qualitative.Set3)
                fig_pie.update_traces(textposition="inside", textinfo="percent+label", hovertemplate="%{label}<br>건수: %{value:,.0f}<br>비중: %{percent}<extra></extra>")
                fig_pie.update_layout(height=500)
                st.plotly_chart(fig_pie, use_container_width=True)
            else:
                fig_tree = px.treemap(map_agg, path=["\uc870\uc9c1"], values="miss_sum", color="miss_rate", title=f"{map_level}\ubcc4 \ubbf8\ucc98\ub9ac \ubd84\ud3ec", color_continuous_scale="RdYlGn_r")
                fig_tree.update_layout(height=500)
                st.plotly_chart(fig_tree, use_container_width=True)

            map_display = map_agg.rename(columns={
                "miss_sum": "\ubbf8\uc2a4\uce94",
                "target_sum": "\ub300\uc0c1\uac74",
                "scan_sum": "\uc2a4\uce94\uac74",
                "miss_rate": "\ubbf8\ucc98\ub9ac\uc728",
                "scan_rate": "\uc2a4\uce94\uc728",
            })
            st.dataframe(
                map_display.style.format({
                    "\ubbf8\uc2a4\uce94": "{:,}",
                    "\ub300\uc0c1\uac74": "{:,}",
                    "\uc2a4\uce94\uac74": "{:,}",
                    "\ubbf8\ucc98\ub9ac\uc728": "{:.1f}%",
                    "\uc2a4\uce94\uc728": "{:.1f}%",
                }),
                use_container_width=True,
                hide_index=True,
            )

    with tab_report:
        st.subheader("\uc804\uccb4 \ub370\uc774\ud130 \uae30\ubc18 \uacc4\uce35\ubcc4 \ubbf8\ucc98\ub9ac \ud604\ud669")
        report_df = build_hierarchy_report(df, sel_months)
        if report_df.empty:
            st.info("\ub370\uc774\ud130\uac00 \uc5c6\uc2b5\ub2c8\ub2e4.")
        else:
            def style_row(row):
                if row["\uad6c\ubd84"] == "\ubd80\ubb38\uacc4":
                    return ["background-color:#1F3864;color:white;font-weight:bold"] * len(row)
                if row["\uad6c\ubd84"] == "\ucd1d\uad04\uacc4":
                    return ["background-color:#2E75B6;color:white;font-weight:bold"] * len(row)
                if row["\uad6c\ubd84"] == "\ubd80\uc11c\uacc4":
                    return ["background-color:#D9E1F2;font-weight:bold"] * len(row)
                return [""] * len(row)

            st.dataframe(
                report_df.style.apply(style_row, axis=1).format({
                    "FA": "{:,}",
                    "\ube44\uad50": "{:,}",
                    "\uc644\ud310": "{:,}",
                    "\ucd1d\ubbf8\uc2a4\uce94": "{:,}",
                    "\ub300\uc0c1\uac74": "{:,}",
                    "\uc2a4\uce94\uac74": "{:,}",
                    "\ubbf8\ucc98\ub9ac\uc728": "{:.1f}%",
                    "\uc2a4\uce94\uc728": "{:.1f}%",
                }),
                use_container_width=True,
                hide_index=True,
                height=500,
            )

            st.divider()
            pivot_df = build_monthly_hierarchy_pivot(df, sel_months)
            if not pivot_df.empty:
                pivot_display = pivot_df.copy()
                for col in pivot_display.columns:
                    if col not in ["\uad6c\ubd84", dept_col, tg_col, ds_col] and not col.endswith("_\ubbf8\ucc98\ub9ac\uc728") and not col.endswith("_\uc2a4\uce94\uc728"):
                        pivot_display[col] = pivot_display[col].apply(lambda x: int(x) if pd.notna(x) else "")
                st.dataframe(
                    pivot_display.style.format({
                        col: "{:,}" for col in pivot_display.columns
                        if col not in ["\uad6c\ubd84", dept_col, tg_col, ds_col] and not col.endswith("_\ubbf8\ucc98\ub9ac\uc728") and not col.endswith("_\uc2a4\uce94\uc728")
                    }).format({
                        col: "{:.1f}%" for col in pivot_display.columns if col.endswith("_\ubbf8\ucc98\ub9ac\uc728") or col.endswith("_\uc2a4\uce94\uc728")
                    }),
                    use_container_width=True,
                    hide_index=True,
                    height=420,
                )

            cr1, cr2 = st.columns(2)
            with cr1:
                if st.button("\uacc4\uce35 \ub9ac\ud3ec\ud2b8 Excel", use_container_width=True):
                    with st.spinner("\uc0dd\uc131 \uc911..."):
                        buf = report_excel(df, sel_months)
                    st.download_button("Excel \ub2e4\uc6b4\ub85c\ub4dc", buf, f"\uacc4\uce35\ub9ac\ud3ec\ud2b8_{period_text.replace(' ', '_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_rpt_xl")
            with cr2:
                if st.button("\uacc4\uce35 \ub9ac\ud3ec\ud2b8 PDF", use_container_width=True):
                    with st.spinner("\uc0dd\uc131 \uc911..."):
                        buf2 = report_pdf(df, sel_months)
                    st.download_button("PDF \ub2e4\uc6b4\ub85c\ub4dc", buf2, f"\uacc4\uce35\ub9ac\ud3ec\ud2b8_{period_text.replace(' ', '_')}.pdf", "application/pdf", key="dl_rpt_pdf")

            st.markdown("---")
            if st.button("\uc804\uccb4 \ud398\uc774\uc9c0 PDF", use_container_width=True, key="dl_fullpage_pdf"):
                with st.spinner("\uc804\uccb4 \ud398\uc774\uc9c0 PDF \uc0dd\uc131 \uc911..."):
                    agg_group_state = st.session_state.get("agg_group", dept_col)
                    map_level_state = st.session_state.get("map_level", dept_col)
                    dash_doc_types = st.session_state.get("dash_doc_types", ["\ucd1d \ubbf8\uc2a4\uce94"])
                    dash_chart_mode = st.session_state.get("dash_chart_mode", "\uadf8\ub8f9\ud615")
                    dash_top_n = st.session_state.get("dash_top_n", 10)
                    map_type_state = st.session_state.get("map_type", "\ud2b8\ub9ac\ub9f5")
                    buf_full = report_fullpage_pdf(df, sel_months, agg_group_state, map_level_state, dash_doc_types, dash_chart_mode, dash_top_n, map_type_state)
                    st.download_button("\uc804\uccb4 \ud398\uc774\uc9c0 PDF \ub2e4\uc6b4\ub85c\ub4dc", buf_full, f"\uc804\uccb4\ud398\uc774\uc9c0\ub9ac\ud3ec\ud2b8_{period_text.replace(' ', '_')}.pdf", "application/pdf", key="dl_fullpage_pdf_btn")

    with tab_ledger:
        st.subheader("\uad00\ub9ac\ub300\uc7a5 \uc120\uc815 \ubc0f \ucd9c\ub825")
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            sel_bm = st.selectbox("\ubd80\ubb38", ["\uc804\uccb4"] + sorted(df_sel[dept_col].dropna().unique().tolist()), key="lg_bm")
        df_l1 = df_sel if sel_bm == "\uc804\uccb4" else df_sel[df_sel[dept_col] == sel_bm]
        with cf2:
            sel_tg = st.selectbox("\ucd1d\uad04", ["\uc804\uccb4"] + sorted(df_l1[tg_col].dropna().unique().tolist()), key="lg_tg")
        df_l2 = df_l1 if sel_tg == "\uc804\uccb4" else df_l1[df_l1[tg_col] == sel_tg]
        with cf3:
            sel_ds = st.selectbox("\ubd80\uc11c", ["\uc804\uccb4"] + sorted(df_l2[ds_col].dropna().unique().tolist()), key="lg_ds")
        df_l3 = df_l2 if sel_ds == "\uc804\uccb4" else df_l2[df_l2[ds_col] == sel_ds]

        targets = get_ledger_targets(df_l3, sel_months)
        if not targets:
            st.success("\ubbf8\uc2a4\uce94 \ubc1c\uc0dd \ub300\uc0c1\uc774 \uc5c6\uc2b5\ub2c8\ub2e4.")
        else:
            prev = [{
                dept_col: r[dept_col],
                tg_col: r[tg_col],
                ds_col: dept,
                fam_col: r[fam_col],
                "FA": int(r["FA"]),
                "\ube44\uad50": int(r["\ube44\uad50"]),
                "\uc644\ud310": int(r["\uc644\ud310"]),
                "\ucd1d\ubbf8\uc2a4\uce94": int(r["\ucd1d\ubbf8\uc2a4\uce94"]),
                target_col: int(r["\ub300\uc0c1"]),
                scan_col: int(r["\uc2a4\uce94"]),
            } for dept, grp in targets.items() for _, r in grp.iterrows()]
            prev_df = pd.DataFrame(prev)
            st.dataframe(
                prev_df.style.format({
                    "FA": "{:,}",
                    "\ube44\uad50": "{:,}",
                    "\uc644\ud310": "{:,}",
                    "\ucd1d\ubbf8\uc2a4\uce94": "{:,}",
                    target_col: "{:,}",
                    scan_col: "{:,}",
                }),
                use_container_width=True,
                hide_index=True,
            )

            all_depts = sorted(targets.keys())
            sel_depts = st.multiselect("\ucd9c\ub825 \ubd80\uc11c", all_depts, default=all_depts, key="lg_sel_dept")
            if not sel_depts:
                st.warning("\ucd9c\ub825\ud560 \ubd80\uc11c\ub97c 1\uac1c \uc774\uc0c1 \uc120\ud0dd\ud574 \uc8fc\uc138\uc694.")
            else:
                out_targets = {d: targets[d] for d in sel_depts if d in targets}
                cd1, cd2 = st.columns(2)
                with cd1:
                    if st.button("\uad00\ub9ac\ub300\uc7a5 PDF", use_container_width=True, key="gen_pdf"):
                        with st.spinner("\uc0dd\uc131 \uc911..."):
                            pb = ledger_pdf(out_targets, period_text, df_l3)
                        st.download_button("PDF \ub2e4\uc6b4\ub85c\ub4dc", pb, f"\uad00\ub9ac\ub300\uc7a5_{period_text.replace(' ', '_')}.pdf", "application/pdf", key="dl_ldg_pdf")
                with cd2:
                    if st.button("\uad00\ub9ac\ub300\uc7a5 Excel", use_container_width=True, key="gen_xl"):
                        with st.spinner("\uc0dd\uc131 \uc911..."):
                            xb = ledger_excel(out_targets, period_text, df_l3)
                        st.download_button("Excel \ub2e4\uc6b4\ub85c\ub4dc", xb, f"\uad00\ub9ac\ub300\uc7a5_{period_text.replace(' ', '_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_ldg_xl")

# ==========================================
# 13. main
# ==========================================
def main():
    if not st.session_state.get("logged_in"):
        login_page()
    else:
        with st.sidebar:
            st.success("👋 접속 완료")
            if st.button("🚪 로그아웃", use_container_width=True):
                st.session_state.logged_in = False
                st.rerun()
            st.divider()
            st.caption("v5.0 | GitHub엑셀기반·실시간배포 | © 2026")
        dashboard_page()

if __name__ == "__main__":
    main()

import streamlit as st
import pandas as pd
import sqlite3
import os
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, PageBreak, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import io

# ==========================================
# 0. 페이지 설정 (최상단 필수)
# ==========================================
st.set_page_config(page_title="보험 서류 스캔 관리 대시보드", layout="wide", page_icon="📊")

# ==========================================
# 1. 전역 설정
# ==========================================
DB_PATH = "data/insurance_db.sqlite"
if not os.path.exists("data"):
    os.makedirs("data")

CREDENTIALS = {
    "admin": {"password": os.environ.get("ADMIN_PW", "admin961"), "role": "admin"},
    "incar": {"password": os.environ.get("USER_PW", "incar123"), "role": "user"},
}

GUIDANCE_TEXT = (
    "【책임판매 필수서류 안내】\n"
    "개인정보동의서, 비교설명확인서, 고지의무확인서, 완전판매확인서(대상계약 限)는 "
    "금융소비자보호법 및 보험업 감독규정에 따라 신계약 체결 전 반드시 완비되어야 하며, "
    "미비 시 리스크 계약으로 간주됩니다."
)
PRECAUTION_TEXT_COVER = (
    "【미처리 시 유의사항】\n"
    "대상 건은 모집질서 위반 및 특정 리스크에 준하여 관리됩니다.\n"
    "신계약 리스크 점검 강화, 회사 지원금 및 특인 제한, 내부 징계 등 불이익이 발생할 수 있으므로 "
    "확인 및 안내되었음을 확인합니다."
)
PRECAUTION_TEXT_SHEET = "본인은 위 내용을 안내받았음을 확인합니다."

# ==========================================
# 2. DB 관리
# ==========================================
@st.cache_resource
def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            보종 TEXT, 보험사 TEXT, 상품군 TEXT, 상품명 TEXT, 증권번호 TEXT,
            계약자 TEXT, 보험료 REAL, 보험시작일 TEXT, 보험종료일 TEXT,
            부문 TEXT, 총괄 TEXT, 소속 TEXT, 영업가족 TEXT, 부서 TEXT,
            담당자 TEXT, 담당자사번 TEXT, 개인정보 TEXT,
            FA고지 TEXT, 비교설명 TEXT, 완전판매 TEXT, 업로드일시 TEXT
        )
    """)
    for col in ["보험시작일", "부문", "총괄", "소속", "영업가족", "부서", "담당자", "증권번호"]:
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_{col} ON contracts({col})")
    conn.commit()

# ==========================================
# 3. 데이터 로딩
# ==========================================
def load_data():
    if "df" not in st.session_state:
        conn = get_db()
        with st.spinner("🔄 데이터 로딩 중..."):
            df = pd.read_sql("SELECT * FROM contracts", conn)
        if df.empty:
            st.session_state.df = pd.DataFrame()
            return st.session_state.df
        
        df["보험시작일_dt"] = pd.to_datetime(df["보험시작일"], errors="coerce")
        df["월_피리어드"] = df["보험시작일_dt"].dt.to_period("M").astype(str)
        df["FA고지_c"] = df["FA고지"].fillna("").astype(str).str.strip()
        df["비교설명_c"] = df["비교설명"].fillna("").astype(str).str.strip()
        df["완전판매_c"] = df["완전판매"].fillna("").astype(str).str.strip()
        st.session_state.df = df
    return st.session_state.df

def upload_data(df, replace=True):
    required = ["FA고지", "비교설명", "완전판매", "보험시작일", "소속", "총괄"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"필수 컬럼 누락: {missing}"); return False
    df["업로드일시"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df.to_sql("contracts", get_db(), if_exists="replace" if replace else "append", index=False)
    st.session_state.pop("df", None)
    return True

def delete_month(month):
    conn = get_db()
    conn.execute("DELETE FROM contracts WHERE substr(보험시작일,1,7)=?", (month,))
    conn.commit()
    st.session_state.pop("df", None)
    st.success(f"✅ {month} 데이터 삭제 완료")

# ==========================================
# 4. 집계 헬퍼
# ==========================================
def _miss(s):
    return (s.fillna("").astype(str).str.strip() == "미스캔").sum()

def _miss_cs(s):
    s2 = s.fillna("").astype(str).str.strip()
    return ((s2 != "해당없음") & (s2 == "미스캔")).sum()

def fmt_rate(val, denom):
    if denom == 0: return "0.0%"
    return f"{val/denom*100:.1f}%"

# ==========================================
# 5. 전체 계층 리포트 (피벗 형태)
# ==========================================
def build_hierarchy_report(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    
    rows = []
    for bm, df_bm in src.groupby("부문"):
        fa=_miss(df_bm["FA고지_c"]); bi=_miss(df_bm["비교설명_c"]); cs=_miss_cs(df_bm["완전판매_c"])
        tot=fa+bi+cs; cnt=len(df_bm)
        rows.append({"구분":"부문계", "부문":bm, "총괄":"", "부서":"", "영업가족":"",
                     "FA":fa, "비교":bi, "완판":cs, "총미스캔":tot, "대상건":cnt,
                     "미처리율":round(tot/cnt*100,1) if cnt else 0.0})
        
        for tg, df_tg in df_bm.groupby("총괄"):
            fa2=_miss(df_tg["FA고지_c"]); bi2=_miss(df_tg["비교설명_c"]); cs2=_miss_cs(df_tg["완전판매_c"])
            tot2=fa2+bi2+cs2; cnt2=len(df_tg)
            rows.append({"구분":"총괄계", "부문":bm, "총괄":tg, "부서":"", "영업가족":"",
                         "FA":fa2, "비교":bi2, "완판":cs2, "총미스캔":tot2, "대상건":cnt2,
                         "미처리율":round(tot2/cnt2*100,1) if cnt2 else 0.0})
            
            for ds, df_ds in df_tg.groupby("부서"):
                fa3=_miss(df_ds["FA고지_c"]); bi3=_miss(df_ds["비교설명_c"]); cs3=_miss_cs(df_ds["완전판매_c"])
                tot3=fa3+bi3+cs3; cnt3=len(df_ds)
                rows.append({"구분":"부서계", "부문":bm, "총괄":tg, "부서":ds, "영업가족":"",
                             "FA":fa3, "비교":bi3, "완판":cs3, "총미스캔":tot3, "대상건":cnt3,
                             "미처리율":round(tot3/cnt3*100,1) if cnt3 else 0.0})
                
                for fg, df_fg in df_ds.groupby("영업가족"):
                    fa4=_miss(df_fg["FA고지_c"]); bi4=_miss(df_fg["비교설명_c"]); cs4=_miss_cs(df_fg["완전판매_c"])
                    t4=fa4+bi4+cs4; c4=len(df_fg)
                    rows.append({"구분":"영업가족", "부문":bm, "총괄":tg, "부서":ds, "영업가족":fg,
                                 "FA":fa4, "비교":bi4, "완판":cs4, "총미스캔":t4, "대상건":c4,
                                 "미처리율":round(t4/c4*100,1) if c4 else 0.0})
    return pd.DataFrame(rows)

def build_monthly_hierarchy(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    
    rows = []
    for mon, dm in src.groupby("월_피리어드"):
        for bm, db in dm.groupby("부문"):
            fa_b=_miss(db["FA고지_c"]); bi_b=_miss(db["비교설명_c"]); cs_b=_miss_cs(db["완전판매_c"])
            rows.append({"월":mon, "구분":"부문계", "부문":bm, "총괄":"", "부서":"",
                         "FA":fa_b, "비교":bi_b, "완판":cs_b, "총미스캔":fa_b+bi_b+cs_b,
                         "대상건":len(db), "미처리율":round((fa_b+bi_b+cs_b)/len(db)*100,1) if len(db) else 0.0})
        for tg, dt in db.groupby("총괄"):
            fa_t=_miss(dt["FA고지_c"]); bi_t=_miss(dt["비교설명_c"]); cs_t=_miss_cs(dt["완전판매_c"])
            rows.append({"월":mon, "구분":"총괄계", "부문":bm, "총괄":tg, "부서":"",
                         "FA":fa_t, "비교":bi_t, "완판":cs_t, "총미스캔":fa_t+bi_t+cs_t,
                         "대상건":len(dt), "미처리율":round((fa_t+bi_t+cs_t)/len(dt)*100,1) if len(dt) else 0.0})
            for ds, dd in dt.groupby("부서"):
                fa=_miss(dd["FA고지_c"]); bi=_miss(dd["비교설명_c"]); cs=_miss_cs(dd["완전판매_c"])
                tot=fa+bi+cs; cnt=len(dd)
                rows.append({"월":mon, "구분":"부서계", "부문":bm, "총괄":tg, "부서":ds,
                             "FA":fa, "비교":bi, "완판":cs, "총미스캔":tot,
                             "대상건":cnt, "미처리율":round(tot/cnt*100,1) if cnt else 0.0})
    return pd.DataFrame(rows)

# ==========================================
# 6. 관리대장 선정 대상
# ==========================================
def get_ledger_targets(df, months):
    src = df[df["월_피리어드"].isin(months)].copy()
    if src.empty: return {}
    agg = src.groupby(["부문", "총괄", "부서", "영업가족"]).agg(
        FA=("FA고지_c", _miss), 비교=("비교설명_c", _miss),
        완판=("완전판매_c", _miss_cs), 대상=("증권번호", "count")
    ).reset_index()
    agg["총미스캔"] = agg[["FA", "비교", "완판"]].sum(axis=1)
    agg = agg[agg["총미스캔"] > 0]
    return {dept: grp for dept, grp in agg.groupby("부서")}

# ==========================================
# 7. 한글 폰트 및 스타일
# ==========================================
@st.cache_resource
def register_korean_font():
    for name, path in [
        ("Malgun", r"C:\Windows\Fonts\malgun.ttf"),
        ("NotoSansKR", "/usr/share/fonts/truetype/noto/NotoSansKR-Regular.otf"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
    ]:
        try:
            pdfmetrics.registerFont(TTFont(name, path)); return name
        except Exception:
            continue
    return "Helvetica"

HDR_CLR  = "#4472C4"
ALT_CLR  = "#EEF3FB"
SUB_CLR  = "#D9E1F2"

def _pdf_styles(fn):
    S = getSampleStyleSheet()
    def ps(n, **kw): return ParagraphStyle(n, parent=S["Normal"], fontName=fn, **kw)
    return {
        "title":   ps("T",  fontSize=15, bold=True, alignment=1, spaceAfter=4),
        "sub":     ps("S",  fontSize=10, spaceAfter=3),
        "body":    ps("B",  fontSize=8,  spaceAfter=2),
        "notice":  ps("N",  fontSize=7.5, spaceAfter=3, textColor=colors.HexColor("#CC0000"), alignment=1),
        "date":    ps("D",  fontSize=8,  alignment=2, spaceAfter=4),
        "section": ps("SC", fontSize=9,  bold=True,  spaceAfter=2),
    }

def _tbl(data, cw, fn, header_rows=1, sub_rows=None):
    if not data or len(data) < 1: return Spacer(1,0)
    # ✅ 가로 너비 40% 자동 스케일링
    cw_scaled = [w * 1.4 for w in cw]
    t = Table(data, colWidths=cw_scaled, repeatRows=header_rows)
    cmds = [
        ("FONTNAME",     (0,0),(-1,-1), fn),
        ("FONTSIZE",     (0,0),(-1,-1), 7.5),
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("GRID",         (0,0),(-1,-1), 0.4, colors.grey),
        ("LEFTPADDING",  (0,0),(-1,-1), 4),
        ("RIGHTPADDING", (0,0),(-1,-1), 4),
        ("TOPPADDING",   (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ("BACKGROUND",   (0,0),(-1,header_rows-1), colors.HexColor(HDR_CLR)),
        ("TEXTCOLOR",    (0,0),(-1,header_rows-1), colors.whitesmoke),
        ("FONTSIZE",     (0,0),(-1,header_rows-1), 8),
    ]
    for i in range(header_rows, len(data)):
        if (i-header_rows)%2==1:
            cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor(ALT_CLR)))
        if sub_rows and i in sub_rows:
            cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor(SUB_CLR)))
    t.setStyle(TableStyle(cmds)); return t

def _sig_table(labels, fn, cw=140):
    t = Table([labels,["____________________"]*len(labels),["(인)"]*len(labels)],
              colWidths=[cw*1.4]*len(labels))
    t.setStyle(TableStyle([
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("FONTNAME",     (0,0),(-1,-1),fn),
        ("FONTSIZE",     (0,0),(-1,-1),8.5),
        ("TOPPADDING",   (0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("BOX",          (0,0),(-1,-1),0.5,colors.grey),
        ("INNERGRID",    (0,0),(-1,-1),0.3,colors.lightgrey),
    ]))
    return t

# ==========================================
# 8. 전체 계층 리포트 Excel (피벗/병합)
# ==========================================
def report_excel(df, months):
    wb  = Workbook(); ws = wb.active; ws.title="계층별_미처리현황"
    tfn = "맑은 고딕"
    hf  = Font(name=tfn,size=9,bold=True,color="FFFFFF")
    bf  = Font(name=tfn,size=9)
    bdr = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    fills = {
        "부문계":  PatternFill("solid",fgColor="1F3864"),
        "총괄계":  PatternFill("solid",fgColor="2E75B6"),
        "부서계":  PatternFill("solid",fgColor="D9E1F2"),
        "영업가족_alt": PatternFill("solid",fgColor="EEF3FB"),
    }
    fonts_wc = {
        "부문계": Font(name=tfn,size=9,bold=True,color="FFFFFF"),
        "총괄계": Font(name=tfn,size=9,bold=True,color="FFFFFF"),
        "부서계": Font(name=tfn,size=9,bold=True),
    }
    h_fill = PatternFill("solid",fgColor="4472C4")
    today  = datetime.now().strftime("%Y년 %m월 %d일")
    period_str = ", ".join(months) if months else "전체"
    
    ws.merge_cells("A1:K1")
    ws["A1"] = f"서류 미처리 현황 계층별 집계  ·  기간: {period_str}  ·  발급: {today}"
    ws["A1"].font = Font(name=tfn,size=12,bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    headers = ["구분","부문","총괄","부서","영업가족","FA고지","비교설명","완전판매","총미스캔","대상건","미처리율"]
    cws     = [14,20,20,20,24,12,12,12,14,12,16] # ✅ 40% 확대
    for ci,(h,w) in enumerate(zip(headers,cws),1):
        c=ws.cell(2,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr
        c.alignment=Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width=w

    report = build_hierarchy_report(df, months)
    if report.empty: return io.BytesIO()

    ri = 3
    for _, row in report.iterrows():
        gbn = row["구분"]
        rate_str = f"{row['미처리율']:.1f}%"
        vals = [gbn, row["부문"], row["총괄"], row["부서"], row["영업가족"],
                row["FA"], row["비교"], row["완판"], row["총미스캔"], row["대상건"], rate_str]
        fill = fills.get(gbn, fills["영업가족_alt"] if ri%2==0 else None)
        fnt  = fonts_wc.get(gbn, bf)
        for ci,v in enumerate(vals,1):
            c=ws.cell(ri,ci,v); c.font=fnt; c.border=bdr
            c.alignment=Alignment(horizontal="center",vertical="center")
            if fill: c.fill=fill
        ri += 1

    # 월별 시트
    ws2 = wb.create_sheet("월별_계층집계")
    monthly = build_monthly_hierarchy(df, months)
    if not monthly.empty:
        ws2.merge_cells("A1:K1")
        ws2["A1"] = f"월별 계층 미처리 집계  ·  기간: {period_str}  ·  발급: {today}"
        ws2["A1"].font = Font(name=tfn,size=12,bold=True)
        ws2["A1"].alignment=Alignment(horizontal="center")
        ws2.row_dimensions[1].height=22
        mhdr=["월","구분","부문","총괄","부서","FA고지","비교설명","완전판매","총미스캔","대상건","미처리율"]
        mcws=[18,14,20,20,24,12,12,12,14,12,16]
        for ci,(h,w) in enumerate(zip(mhdr,mcws),1):
            c=ws2.cell(2,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr
            c.alignment=Alignment(horizontal="center"); ws2.column_dimensions[get_column_letter(ci)].width=w
        for ri2,(_, r) in enumerate(monthly.iterrows(),3):
            gbn=r["구분"]
            rate_str = f"{r['미처리율']:.1f}%"
            vals2=[r["월"],gbn,r["부문"],r["총괄"],r["부서"],
                   r["FA"],r["비교"],r["완판"],r["총미스캔"],r["대상건"], rate_str]
            fill2=fills.get(gbn, fills["영업가족_alt"] if ri2%2==0 else None)
            fnt2=fonts_wc.get(gbn,bf)
            for ci,v in enumerate(vals2,1):
                c=ws2.cell(ri2,ci,v); c.font=fnt2; c.border=bdr
                c.alignment=Alignment(horizontal="center")
                if fill2: c.fill=fill2

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ==========================================
# 9. 전체 계층 리포트 PDF
# ==========================================
def report_pdf(df, months):
    fn  = register_korean_font()
    st_ = _pdf_styles(fn)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=10*mm,leftMargin=10*mm,
                            topMargin=12*mm,bottomMargin=12*mm)
    today      = datetime.now().strftime("%Y년 %m월 %d일")
    period_str = ", ".join(months) if months else "전체"
    E = []
    E.append(Paragraph("서류 미처리 현황 계층별 집계", st_["title"]))
    E.append(Paragraph(f"기간: {period_str}  |  발급일자: {today}", st_["date"]))
    E.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor(HDR_CLR)))
    E.append(Spacer(1,6))
    
    report = build_hierarchy_report(df, months)
    if not report.empty:
        E.append(Paragraph("▶ 부문 / 총괄 / 부서 / 영업가족 계층 집계", st_["section"]))
        hdr=[["구분","부문","총괄","부서","영업가족","FA","비교","완판","총미스캔","대상건","미처리율"]]
        drows=[]; sub_idx=[]
        for i,(_,r) in enumerate(report.iterrows()):
            rate_str = f"{r['미처리율']:.1f}%"
            drows.append([r["구분"],r["부문"],r["총괄"],r["부서"],r["영업가족"],
                          r["FA"],r["비교"],r["완판"],r["총미스캔"],r["대상건"],rate_str])
            if r["구분"] in ("부문계","총괄계","부서계"): sub_idx.append(i+1)
        cw=[22,36,36,36,42,16,16,16,22,18,22]
        E.append(_tbl(hdr+drows,cw,fn,sub_rows=sub_idx))
        E.append(Spacer(1,8))

    monthly = build_monthly_hierarchy(df, months)
    if not monthly.empty:
        E.append(PageBreak())
        E.append(Paragraph("▶ 월별 계층별 미처리 집계", st_["section"]))
        mhdr=[["월","구분","부문","총괄","부서","FA","비교","완판","총미스캔","대상건","미처리율"]]
        mrows=[]
        for _,r in monthly.iterrows():
            rate_str = f"{r['미처리율']:.1f}%"
            mrows.append([r["월"],r["구분"],r["부문"],r["총괄"],r["부서"],
                          r["FA"],r["비교"],r["완판"],r["총미스캔"],r["대상건"],rate_str])
        msub=[i+1 for i,(_,r) in enumerate(monthly.iterrows()) if r["구분"] in ("부문계","총괄계","부서계")]
        cw2=[28,22,36,36,40,16,16,16,22,18,22]
        E.append(_tbl(mhdr+mrows,cw2,fn,sub_rows=msub))

    doc.build(E); buf.seek(0); return buf

# ==========================================
# 10. 관리대장 PDF
# ==========================================
def ledger_pdf(families_by_dept, period_text, df_src):
    fn  = register_korean_font()
    st_ = _pdf_styles(fn)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=12*mm,leftMargin=12*mm,
                            topMargin=15*mm,bottomMargin=15*mm)
    today = datetime.now().strftime("%Y년 %m월 %d일")
    E = []
    for dept_name, grp_df in families_by_dept.items():
        sec = grp_df.iloc[0]["부문"]; tg = grp_df.iloc[0]["총괄"]

        E.append(Paragraph("미처리 서류 확인서", st_["title"]))
        E.append(HRFlowable(width="100%",thickness=1.5,color=colors.HexColor(HDR_CLR)))
        E.append(Spacer(1,4))
        E.append(Paragraph(f"부서: {sec}  > {tg}  >  <b>{dept_name}</b>", st_["sub"]))
        E.append(Paragraph(f"적용기간: {period_text}  발급일자: {today}", st_["date"]))
        E.append(Spacer(1,6))

        dept_src = df_src[df_src["부서"]==dept_name]
        if not dept_src.empty:
            E.append(Paragraph("▶ 영업가족별 · 월별 · 양식별 미처리 현황", st_["section"]))
            fam_mon = dept_src.groupby(["영업가족","월_피리어드"]).agg(
                FA=("FA고지_c",_miss),비교=("비교설명_c",_miss),완판=("완전판매_c",_miss_cs)
            ).reset_index()
            fam_mon["계"] = fam_mon[["FA","비교","완판"]].sum(axis=1)
            fam_mon = fam_mon[fam_mon["계"] > 0]
            if not fam_mon.empty:
                td=[["영업가족","월","FA고지","비교설명","완전판매","계"]]
                for _,r in fam_mon.iterrows():
                    td.append([r["영업가족"],r["월_피리어드"],int(r.FA),int(r.비교),int(r.완판),int(r["계"])])
                E.append(_tbl(td,[80,36,32,32,32,28],fn))
            E.append(Spacer(1,8))

        E.append(Paragraph(GUIDANCE_TEXT, st_["notice"]))
        E.append(Paragraph(PRECAUTION_TEXT_COVER, st_["notice"]))
        E.append(Spacer(1,8))
        E.append(Paragraph(f"작성일: {today}", st_["date"]))
        E.append(_sig_table(["부문장 확인","총괄 확인","부서장 확인"],fn,140))
        E.append(PageBreak())

        for _, fam in grp_df.drop_duplicates("영업가족").iterrows():
            fam_name = fam["영업가족"]
            E.append(Paragraph("서류 미처리 확인서", st_["title"]))
            E.append(HRFlowable(width="100%",thickness=1.5,color=colors.HexColor(HDR_CLR)))
            E.append(Spacer(1,4))
            E.append(Paragraph(f"소속: {sec}  > {tg}  > {dept_name}  >  <b>{fam_name}</b>",st_["sub"]))
            E.append(Paragraph(f"적용기간: {period_text}  발급일자: {today}", st_["date"]))
            E.append(Spacer(1,6))

            fam_src = df_src[(df_src["영업가족"]==fam_name) & df_src["소속"].notna()]
            sosok = fam_src.groupby(["소속","월_피리어드"]).agg(
                FA=("FA고지_c",_miss),비교=("비교설명_c",_miss),완판=("완전판매_c",_miss_cs)
            ).reset_index()
            sosok["계"] = sosok[["FA","비교","완판"]].sum(axis=1)
            sosok = sosok[sosok["계"] > 0]

            E.append(Paragraph("▶ 소속별 · 월별 · 양식별 미처리 건수", st_["section"]))
            if not sosok.empty:
                td2=[["소속","월","FA고지","비교설명","완전판매","계"]]
                for _,r in sosok.iterrows():
                    td2.append([r["소속"],r["월_피리어드"],int(r.FA),int(r.비교),int(r.완판),int(r["계"])])
                E.append(_tbl(td2,[80,36,32,32,32,28],fn))
            else:
                E.append(Paragraph("(해당 데이터 없음)", st_["body"]))
            E.append(Spacer(1,6))

            E.append(Paragraph("▶ 양식별 미처리 요약", st_["section"]))
            sum_d=[["FA고지","비교설명","완전판매","총계"],
                   [str(int(fam["FA"])),str(int(fam["비교"])),str(int(fam["완판"])),str(int(fam["총미스캔"]))]]
            E.append(_tbl(sum_d,[60,60,60,60],fn))
            E.append(Spacer(1,8))

            E.append(Paragraph(GUIDANCE_TEXT, st_["notice"]))
            E.append(Spacer(1,4))
            E.append(Paragraph("본인은 위 미처리 건수에 대하여 인지하였으며, 관련 리스크 관리에 최선을 다할 것을 확인합니다.", st_["body"]))
            E.append(Paragraph(PRECAUTION_TEXT_SHEET, st_["notice"]))
            E.append(Spacer(1,10))
            E.append(Paragraph(f"작성일: {today}", st_["date"]))
            sig2=Table([[f"영업가족대표 서명: ____________________ (인)"]],colWidths=[260])
            sig2.setStyle(TableStyle([
                ("ALIGN",(0,0),(-1,-1),"LEFT"),("FONTNAME",(0,0),(-1,-1),fn),
                ("FONTSIZE",(0,0),(-1,-1),9.5),("TOPPADDING",(0,0),(-1,-1),8),
                ("BOTTOMPADDING",(0,0),(-1,-1),8),("BOX",(0,0),(-1,-1),0.5,colors.grey),
            ]))
            E.append(sig2)
            E.append(PageBreak())

    doc.build(E); buf.seek(0); return buf

# ==========================================
# 11. 관리대장 Excel
# ==========================================
def ledger_excel(families_by_dept, period_text, df_src):
    wb  = Workbook(); ws0=wb.active; ws0.title="목차"
    tfn = "맑은 고딕"
    hf  = Font(name=tfn,size=9,bold=True,color="FFFFFF")
    bf  = Font(name=tfn,size=9)
    nf  = Font(name=tfn,size=8,italic=True,color="CC0000")
    sig_f=Font(name=tfn,size=9,bold=True)
    bdr = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    h_fill  =PatternFill("solid",fgColor="4472C4")
    alt_fill=PatternFill("solid",fgColor="EEF3FB")
    today   =datetime.now().strftime("%Y년 %m월 %d일")
    
    ws0.merge_cells("A1:F1")
    ws0["A1"]=f"관리대장 목차  ·  {period_text}  ·  발급: {today}"
    ws0["A1"].font=Font(name=tfn,size=13,bold=True)
    ws0["A1"].alignment=Alignment(horizontal="center")
    for ci,h in enumerate(["부서","영업가족","FA고지","비교설명","완전판매","총미스캔"],1):
        c=ws0.cell(3,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr
        c.alignment=Alignment(horizontal="center")
    ir=4
    for dept,grp in families_by_dept.items():
        for _,fam in grp.drop_duplicates("영업가족").iterrows():
            for ci,v in enumerate([dept,fam["영업가족"],int(fam["FA"]),int(fam["비교"]),int(fam["완판"]),int(fam["총미스캔"])],1):
                c=ws0.cell(ir,ci,v); c.font=bf; c.border=bdr
                c.alignment=Alignment(horizontal="center")
                if ir%2==0: c.fill=alt_fill
            ir+=1
    for ci,w in enumerate([22,25,13,13,13,14],1): # ✅ 40% 확대
        ws0.column_dimensions[get_column_letter(ci)].width=w

    for dept_name,grp_df in families_by_dept.items():
        sec=grp_df.iloc[0]["부문"]; tg=grp_df.iloc[0]["총괄"]
        sname=f"표지_{dept_name[:10]}".replace("/","_")
        ws_c=wb.create_sheet(title=sname)
        ws_c.merge_cells("A1:G1")
        ws_c["A1"]=f"[{dept_name}]  미처리 서류 확인서"
        ws_c["A1"].font=Font(name=tfn,size=14,bold=True)
        ws_c["A2"]=f"{sec}  > {tg}  > {dept_name}"
        ws_c["A2"].font=Font(name=tfn,size=10)
        ws_c["A3"]=f"적용기간: {period_text}  |  발급일자: {today}"
        ws_c["A3"].font=bf

        r=5
        ws_c.cell(r,1,"▶ 영업가족별 · 월별 · 양식별 미처리 현황").font=Font(name=tfn,size=10,bold=True); r+=1
        dept_src=df_src[df_src["부서"]==dept_name]
        if not dept_src.empty:
            fam_mon=dept_src.groupby(["영업가족","월_피리어드"]).agg(
                FA=("FA고지_c",_miss),비교=("비교설명_c",_miss),완판=("완전판매_c",_miss_cs)
            ).reset_index()
            fam_mon["계"]=fam_mon[["FA","비교","완판"]].sum(axis=1)
            fam_mon=fam_mon[fam_mon["계"] > 0]
            hdrs=["영업가족","월","FA고지","비교설명","완전판매","계"]; cws=[25,20,13,13,13,13] # ✅ 40% 확대
            for ci,(h,w) in enumerate(zip(hdrs,cws),1):
                c=ws_c.cell(r,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr
                c.alignment=Alignment(horizontal="center")
                ws_c.column_dimensions[get_column_letter(ci)].width=w
            for i,(_,rv) in enumerate(fam_mon.iterrows()):
                row_v=[rv["영업가족"],rv["월_피리어드"],int(rv.FA),int(rv.비교),int(rv.완판),int(rv["계"])]
                af=alt_fill if i%2==1 else None
                for ci,v in enumerate(row_v,1):
                    c=ws_c.cell(r+1+i,ci,v); c.font=bf; c.border=bdr
                    c.alignment=Alignment(horizontal="center")
                    if af: c.fill=af
            r+=len(fam_mon)+2

        ws_c.cell(r,1,GUIDANCE_TEXT).font=nf
        ws_c.cell(r,1).alignment=Alignment(wrapText=True)
        ws_c.row_dimensions[r].height=45; r+=2
        ws_c.cell(r,1,PRECAUTION_TEXT_COVER).font=nf
        ws_c.cell(r,1).alignment=Alignment(wrapText=True)
        ws_c.row_dimensions[r].height=45; r+=3
        ws_c.cell(r,1,f"작성일: {today}").font=bf; r+=2
        for i,sig in enumerate(["부문장 확인","총괄 확인","부서장 확인"]):
            ws_c.cell(r,i*2+1,sig).font=sig_f
            ws_c.cell(r,i*2+2,"________________ (인)").font=Font(name=tfn,color="888888")

        for _,fam in grp_df.drop_duplicates("영업가족").iterrows():
            fam_name=fam["영업가족"]
            fn_safe=fam_name[:14].replace("/","_").replace(" ","")
            ws_f=wb.create_sheet(title=fn_safe)
            ws_f.merge_cells("A1:G1")
            ws_f["A1"]=f"[{fam_name}]  서류 미처리 확인서"
            ws_f["A1"].font=Font(name=tfn,size=13,bold=True)
            ws_f["A2"]=f"{sec}  > {tg}  > {dept_name}  > {fam_name}"
            ws_f["A2"].font=Font(name=tfn,size=9,italic=True)
            ws_f["A3"]=f"적용기간: {period_text}  |  발급일자: {today}"; ws_f["A3"].font=bf

            r_f=5
            ws_f.cell(r_f,1,"▶ 소속별 · 월별 · 양식별 미처리 건수").font=Font(name=tfn,size=10,bold=True); r_f+=1
            fam_src=df_src[(df_src["영업가족"]==fam_name) & df_src["소속"].notna()]
            sosok=fam_src.groupby(["소속","월_피리어드"]).agg(
                FA=("FA고지_c",_miss),비교=("비교설명_c",_miss),완판=("완전판매_c",_miss_cs)
            ).reset_index()
            sosok["계"]=sosok[["FA","비교","완판"]].sum(axis=1)
            sosok=sosok[sosok["계"] > 0]
            sh=["소속","월","FA고지","비교설명","완전판매","계"]; sc=[25,20,13,13,13,13] # ✅ 40% 확대
            for ci,(h,w) in enumerate(zip(sh,sc),1):
                c=ws_f.cell(r_f,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr
                c.alignment=Alignment(horizontal="center")
                ws_f.column_dimensions[get_column_letter(ci)].width=w
            if not sosok.empty:
                for i,(_,sr) in enumerate(sosok.iterrows()):
                    rv2=[sr["소속"],sr["월_피리어드"],int(sr.FA),int(sr.비교),int(sr.완판),int(sr["계"])]
                    af=alt_fill if i%2==1 else None
                    for ci,v in enumerate(rv2,1):
                        c=ws_f.cell(r_f+1+i,ci,v); c.font=bf; c.border=bdr
                        c.alignment=Alignment(horizontal="center")
                        if af: c.fill=af
                r_f+=len(sosok)+2

            ws_f.cell(r_f,1,"▶ 양식별 요약").font=Font(name=tfn,size=10,bold=True); r_f+=1
            for ci,h in enumerate(["FA고지","비교설명","완전판매","총계"],1):
                c=ws_f.cell(r_f,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr
                c.alignment=Alignment(horizontal="center")
            for ci,v in enumerate([int(fam["FA"]),int(fam["비교"]),int(fam["완판"]),int(fam["총미스캔"])],1):
                c=ws_f.cell(r_f+1,ci,v); c.font=bf; c.border=bdr
                c.alignment=Alignment(horizontal="center")
            r_f+=3

            ws_f.cell(r_f,1,GUIDANCE_TEXT).font=nf
            ws_f.cell(r_f,1).alignment=Alignment(wrapText=True)
            ws_f.row_dimensions[r_f].height=40; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_SHEET).font=nf; r_f+=2
            ws_f.cell(r_f,1,f"작성일: {today}").font=bf; r_f+=1
            ws_f.cell(r_f,1,"영업가족대표 서명: ________________ (인)").font=sig_f

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ==========================================
# 12. UI – 로그인
# ==========================================
def login_page():
    _,col,_ = st.columns([1,1.4,1])
    with col:
        st.markdown("## 🔐 로그인")
        with st.container(border=True):
            user = st.text_input("아이디")
            pwd  = st.text_input("비밀번호", type="password")
            if st.button("로그인", use_container_width=True, type="primary"):
                if user in CREDENTIALS and CREDENTIALS[user]["password"] == pwd:
                    st.session_state.logged_in = True
                    st.session_state.role = CREDENTIALS[user]["role"]
                    st.rerun()
                else:
                    st.error("아이디 또는 비밀번호가 올바르지 않습니다.")

# ==========================================
# 13. UI – 관리자
# ==========================================
def admin_page():
    st.title("👨‍💼 관리자 대시보드")
    t1,t2 = st.tabs(["📤 데이터 업로드", "🗑️ 월별 삭제"])
    with t1:
        f = st.file_uploader("엑셀 업로드 (.xlsx/.xls)", type=["xlsx","xls"])
        if f:
            df_up = pd.read_excel(f)
            st.dataframe(df_up.head(5), use_container_width=True, hide_index=True)
            mode = st.radio("저장 방식", ["전체 대체","추가"], horizontal=True)
            if st.button("💾 저장", type="primary"):
                if upload_data(df_up, replace=(mode=="전체 대체")):
                    st.success("✅ 저장 완료!")
    with t2:
        try:
            months = pd.read_sql("SELECT DISTINCT substr(보험시작일,1,7) as m FROM contracts ORDER BY m DESC", get_db())["m"].tolist()
            if months:
                sel = st.selectbox("삭제할 월", months)
                ok = st.checkbox(f"⚠️ {sel} 삭제에 동의합니다.")
                if st.button("삭제", type="secondary", disabled=not ok): delete_month(sel)
            else: st.info("데이터 없음")
        except: st.info("데이터 없음")

# ==========================================
# 14. UI – 사용자
# ==========================================
def user_page():
    st.title("📊 서류 처리 현황 대시보드")
    df = load_data()
    if df.empty:
        st.warning("📭 관리자가 데이터를 업로드해주세요."); return
    
    all_months = sorted(df["월_피리어드"].dropna().unique())
    st.subheader("📅 분석 기간 선택")
    sel_months = st.multiselect("월 선택 (복수 가능)", all_months,
                               default=[all_months[-1]] if all_months else [])
    if not sel_months:
        st.warning("⚠️ 최소 1개 이상의 월을 선택해주세요."); return

    period_text = f"{sel_months[0]} ~ {sel_months[-1]}" if len(sel_months) > 1 else sel_months[0]
    df_sel = df[df["월_피리어드"].isin(sel_months)].copy()

    try:
        upd = pd.read_sql("SELECT MAX(업로드일시) as u FROM contracts", get_db()).iloc[0,0]
        st.info(f"📅 데이터 기준일: `{upd}` | 분석 기간: `{period_text}` | 총 {len(df_sel):,}건")
    except: pass

    fa_t = _miss(df_sel["FA고지_c"]); bi_t = _miss(df_sel["비교설명_c"]); cs_t = _miss_cs(df_sel["완전판매_c"])
    tot = fa_t + bi_t + cs_t; rate = round(tot/len(df_sel)*100,1) if len(df_sel) > 0 else 0.0
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("📄 총 계약건수", f"{len(df_sel):,}건")
    m2.metric("⚠️ 총 미처리건수", f"{tot:,}건")
    m3.metric("📉 미처리율", f"{rate:.1f}%")
    m4.metric("FA / 비교 / 완판", f"{fa_t} / {bi_t} / {cs_t}")
    st.divider()

    tab_dash, tab_report, tab_ledger = st.tabs([
        "📈 현황 대시보드", "📊 전체 계층 리포트", "📋 관리대장 선정 & 출력"
    ])

    # ── TAB 1 : 현황 대시보드 ─────────────────────────
    with tab_dash:
        cs1, cs2 = st.columns([2, 1])
        with cs1: search_text = st.text_input("🔍 조직 검색", placeholder="조직명 입력...")
        with cs2: agg_group = st.selectbox("집계 기준 (랭킹 단위)", ["부문","총괄","부서","영업가족"])

        agg = df_sel.groupby(agg_group).agg(
            FA고지_미스캔=("FA고지_c", _miss),
            비교설명_미스캔=("비교설명_c", _miss),
            완전판매_미스캔=("완전판매_c", _miss_cs),
            대상건=("증권번호", "count")
        ).reset_index()
        agg["총_미스캔"] = agg[["FA고지_미스캔","비교설명_미스캔","완전판매_미스캔"]].sum(axis=1)
        agg["미처리율"] = (agg["총_미스캔"] / agg["대상건"] * 100).round(1)
        agg = agg.rename(columns={agg_group: "조직"})
        if search_text:
            agg = agg[agg["조직"].astype(str).str.contains(search_text, case=False, na=False)]
        agg = agg.sort_values("총_미스캔", ascending=False).reset_index(drop=True)
        agg.insert(0, "순위", range(1, len(agg)+1))

        if agg.empty:
            st.info("조건에 맞는 데이터가 없습니다.")
        else:
            disp = ["순위","조직","총_미스캔","미처리율","FA고지_미스캔","비교설명_미스캔","완전판매_미스캔"]
            st.dataframe(agg[disp], use_container_width=True, hide_index=True)

            top_n = st.slider("차트 표시 개수", 5, 30, 30)
            top = agg.head(top_n)
            c1, c2 = st.columns(2)
            with c1:
                doc_types = st.multiselect("표시 서류", ["FA고지","비교설명","완전판매","총 미스캔"], default=["총 미스캔"])
                if doc_types:
                    max_v = top["총_미스캔"].max(); yr = [0, max_v*1.2] if max_v > 0 else [0, 10]
                    if len(doc_types)==1 and doc_types[0]=="총 미스캔":
                        fig = go.Figure(); fig.add_trace(go.Bar(
                            x=top["조직"], y=top["총_미스캔"], text=top["총_미스캔"],
                            textposition="outside", marker_color=top["총_미스캔"], marker_colorscale="Reds"))
                        fig.update_layout(title=f"미처리 건수 TOP {top_n} ({agg_group} 기준)",
                            xaxis_tickangle=-45, yaxis=dict(range=yr), height=420)
                        st.plotly_chart(fig, use_container_width=True)
                    elif len(doc_types)==1:
                        cm = {"FA고지":"FA고지_미스캔","비교설명":"비교설명_미스캔","완전판매":"완전판매_미스캔"}
                        y = cm[doc_types[0]]
                        fig = px.bar(top, x="조직", y=y, title=f"{doc_types[0]} 미스캔 TOP {top_n}",
                                     text=y, color=y, color_continuous_scale="Blues")
                        fig.update_layout(xaxis_tickangle=-45, height=420)
                        st.plotly_chart(fig, use_container_width=True)
                    else:
                        ct = st.radio("차트 유형", ["그룹형","누적형"], horizontal=True)
                        cm2 = {"FA고지":"FA고지_미스캔","비교설명":"비교설명_미스캔",
                               "완전판매":"완전판매_미스캔","총 미스캔":"총_미스캔"}
                        mc = [cm2.get(d, d) for d in doc_types]
                        p = top[["조직"]+mc].copy(); p.columns = ["조직"]+doc_types
                        p = p.melt("조직", var_name="종류", value_name="건수")
                        fig = px.bar(p, x="조직", y="건수", color="종류",
                                     barmode="group" if ct=="그룹형" else "stack",
                                     color_discrete_map={"FA고지":"#FF6B6B","비교설명":"#4ECDC4",
                                                         "완전판매":"#45B7D1","총 미스캔":"#9B59B6"})
                        fig.update_layout(xaxis_tickangle=-45, height=420)
                        st.plotly_chart(fig, use_container_width=True)
            with c2:
                max_v = top["총_미스캔"].max(); yr = [0, max_v*1.2] if max_v > 0 else [0, 10]
                fig2 = go.Figure(); fig2.add_trace(go.Scatter(
                    x=top["조직"], y=top["총_미스캔"], mode="lines+markers",
                    line=dict(shape="spline", color="#CC0000"), marker=dict(size=6)))
                fig2.update_layout(title=f"미처리 건수 추이 TOP {top_n}",
                    xaxis_tickangle=-45, yaxis=dict(range=yr), height=420)
                st.plotly_chart(fig2, use_container_width=True)

    # ── TAB 2 : 전체 계층 리포트 ─────────────────────────
    with tab_report:
        st.subheader("📊 전체 데이터 기반 계층별 미처리 현황")
        st.info("백데이터 전체 기준 · 부문계 → 총괄계 → 부서계 → 영업가족 계층 집계")

        report_df = build_hierarchy_report(df, sel_months)
        if report_df.empty:
            st.info("데이터가 없습니다.")
        else:
            def style_row(row):
                if row["구분"]=="부문계":   return ["background-color:#1F3864;color:white;font-weight:bold"]*len(row)
                elif row["구분"]=="총괄계": return ["background-color:#2E75B6;color:white;font-weight:bold"]*len(row)
                elif row["구분"]=="부서계": return ["background-color:#D9E1F2;font-weight:bold"]*len(row)
                return [""]*len(row)

            # ✅ 미처리율 포맷팅 적용
            disp_df = report_df.copy()
            disp_df["미처리율"] = disp_df["미처리율"].apply(lambda x: f"{x:.1f}%")
            st.dataframe(disp_df.style.apply(style_row, axis=1), use_container_width=True, hide_index=True, height=500)

            st.markdown("#### 📅 월별 계층 집계")
            monthly_df = build_monthly_hierarchy(df, sel_months)
            if not monthly_df.empty:
                m_disp = monthly_df.copy()
                m_disp["미처리율"] = m_disp["미처리율"].apply(lambda x: f"{x:.1f}%")
                st.dataframe(m_disp, use_container_width=True, hide_index=True, height=350)

            st.divider()
            cr1, cr2 = st.columns(2)
            with cr1:
                if st.button("📥 전체 계층 리포트 Excel 생성 (피벗/병합)", use_container_width=True):
                    with st.spinner("Excel 생성 중..."):
                        buf = report_excel(df, sel_months)
                    st.download_button("⬇️ Excel 다운로드", buf, f"계층리포트_{period_text.replace(' ','_')}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_rpt_xl")
            with cr2:
                if st.button("📥 전체 계층 리포트 PDF 생성", use_container_width=True):
                    with st.spinner("PDF 생성 중..."):
                        buf2 = report_pdf(df, sel_months)
                    st.download_button("⬇️ PDF 다운로드", buf2, f"계층리포트_{period_text.replace(' ','_')}.pdf",
                        "application/pdf", key="dl_rpt_pdf")

    # ── TAB 3 : 관리대장 선정 & 출력 ───────────────────
    with tab_ledger:
        st.subheader("📋 관리대장 선정 및 출력")
        st.markdown("#### 🔽 대상 필터")
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            sel_bm = st.selectbox("부문", ["전체"]+sorted(df_sel["부문"].dropna().unique().tolist()), key="lg_bm")
        df_l1 = df_sel if sel_bm=="전체" else df_sel[df_sel["부문"]==sel_bm]
        with cf2:
            sel_tg = st.selectbox("총괄", ["전체"]+sorted(df_l1["총괄"].dropna().unique().tolist()), key="lg_tg")
        df_l2 = df_l1 if sel_tg=="전체" else df_l1[df_l1["총괄"]==sel_tg]
        with cf3:
            sel_ds = st.selectbox("부서", ["전체"]+sorted(df_l2["부서"].dropna().unique().tolist()), key="lg_ds")
        df_l3 = df_l2 if sel_ds=="전체" else df_l2[df_l2["부서"]==sel_ds]

        targets = get_ledger_targets(df_l3, sel_months)
        if not targets:
            st.success("✅ 미스캔 발생 대상이 없습니다.")
        else:
            prev = []
            for dept, grp in targets.items():
                for _, r in grp.iterrows():
                    prev.append({"부문":r["부문"],"총괄":r["총괄"],"부서":dept,
                                 "영업가족":r["영업가족"],"FA":int(r["FA"]),
                                 "비교":int(r["비교"]),"완판":int(r["완판"]),"총미스캔":int(r["총미스캔"])})
            prev_df = pd.DataFrame(prev)
            st.markdown(f"#### 📌 선정 대상 — 총 **{len(prev_df)}** 개 영업가족")
            st.dataframe(prev_df, use_container_width=True, hide_index=True)

            st.markdown("#### ✅ 출력할 부서 선택")
            all_depts = sorted(targets.keys())
            sel_depts = st.multiselect("출력 부서 (미선택 시 전체)", all_depts, default=all_depts, key="lg_sel_dept")
            if not sel_depts:
                st.warning("⚠️ 출력할 부서를 1개 이상 선택하세요.")
            else:
                out_targets = {d: targets[d] for d in sel_depts if d in targets}
                sel_prev = [r for r in prev if r["부서"] in sel_depts]
                st.info(f"📄 출력 대상: **{len(sel_depts)}개 부서** · **{len(sel_prev)}개 영업가족**  |   "
                         "생성 구조: `[부서 표지] → [영업가족 낱장(소속·월·양식별 집계+서명)]`")

                cd1, cd2 = st.columns(2)
                with cd1:
                    if st.button("📥 관리대장 PDF 생성", use_container_width=True, key="gen_pdf"):
                        with st.spinner("PDF 생성 중..."):
                            pb = ledger_pdf(out_targets, period_text, df_l3)
                        st.download_button("⬇️ PDF 다운로드", pb, f"관리대장_{period_text.replace(' ','_')}.pdf",
                            "application/pdf", key="dl_ldg_pdf")
                with cd2:
                    if st.button("📥 관리대장 Excel 생성", use_container_width=True, key="gen_xl"):
                        with st.spinner("Excel 생성 중..."):
                            xb = ledger_excel(out_targets, period_text, df_l3)
                        st.download_button("⬇️ Excel 다운로드", xb, f"관리대장_{period_text.replace(' ','_')}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_ldg_xl")

# ==========================================
# 15. main
# ==========================================
def main():
    init_db()
    if not st.session_state.get("logged_in"):
        login_page()
    else:
        with st.sidebar:
            st.success(f"👋 {st.session_state.role.upper()} 로그인 중")
            if st.button("🚪 로그아웃", use_container_width=True):
                st.session_state.logged_in = False
                st.session_state.role = None
                st.rerun()
            st.divider()
            st.caption("v3.2 | 표너비40%확대·미처리율소수점1자리·문법완전수정 | © 2026")
        if st.session_state.role == "admin": admin_page()
        else: user_page()

if __name__ == "__main__":
    main()
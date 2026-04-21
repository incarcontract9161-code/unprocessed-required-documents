import streamlit as st
import pandas as pd
import os
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, PageBreak, HRFlowable, Image)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import io

# ==========================================
# 0. 페이지 설정 (최상단 필수)
# ==========================================
st.set_page_config(page_title="보험 서류 스캔 관리 대시보드", layout="wide", page_icon="📊")

# ==========================================
# 1. 전역 설정 & 인증
# ==========================================
EXCEL_FILE = "insurance_data.xlsx"
APP_PASSWORD = os.environ.get("APP_PASSWORD", "incar961")

GUIDANCE_TEXT = (
    "【책임판매 필수서류 안내】\n "
    "개인정보동의서, 비교설명확인서, 고지의무확인서, 완전판매확인서(대상계약 限)는  "
    "금융소비자보호법 및 보험업 감독규정에 따라 신계약 체결 전 반드시 완비되어야 하며,  "
    "미비 시 리스크 계약으로 간주됩니다. "
)
PRECAUTION_TEXT_COVER = (
    "【미처리 시 유의사항】\n "
    "미결 조직은 모집질서 위반 및 특정 리스크에 준하여 관리됩니다.\n "
    "신계약 리스크 점검 강화, 회사 지원금 및 특인 제한 등 불이익이 발생할 수 있습니다. "
)
PRECAUTION_TEXT_CONFIRM =  "영업가족별 미처리 현황 및 유의사항에 대하여 인지하였으며, "
PRECAUTION_TEXT_SHEET =  "영업가족별 미처리현황 및 유의사항에 대하여 안내받았음을 확인합니다. "
SIGNATURE_CONFIRMATION_TEXT =  "영업가족에게 안내하였음을 확인합니다. "

REQUIRED_DOCS_TABLE = [
    [ "No. ",  "서류명 ",  "법적 근거 ",  "목적 및 주요 내용 "],
    [ "1 ",  "개인정보동의서 ", "개인정보보호법 15조 및\n대리점 표준 내부통제기준 27조 ", "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에\n따른 개인정보 처리로 신계약시 필수 징구 "],
    [ "2 ",  "비교설명확인서 ", "보험업감독규정\n별표 5-6 ", "유사 상품 3개 이상 비교·설명 이행\n사실 고객 확인 서명 "],
    [ "3 ",  "고지의무확인서 ", "금융소비자보호법 26조와\n동법시행령 24조 ", "판매자 권한·책임·보상 관련 핵심 사항 고지,\n소비자 오인 예방 "],
    [ "4 ",  "완전판매확인서\n(대상: 종신, CI, CEO경기, 고액) ", "금융소비자보호법 제17·19조\n영업지원기준안 ", "약관,청약서 부본 제공, 중요 상품 이해 및\n자발적 가입 확인, 설명 의무 이행 증빙력 확보 "]
]

# ==========================================
# 2. 데이터 로딩 (GitHub 엑셀 기반)
# ==========================================
@st.cache_data(ttl=300)
def load_data():
    if not os.path.exists(EXCEL_FILE):
        st.error(f"⚠️ '{EXCEL_FILE}' 파일이 없습니다. GitHub에 엑셀 파일을 업로드해주세요.")
        return pd.DataFrame()
    try:
        df = pd.read_excel(EXCEL_FILE)
        if df.empty: return pd.DataFrame()
        
        df["보험시작일_dt"] = pd.to_datetime(df["보험시작일"], errors="coerce")
        df["월_피리어드"] = df["보험시작일_dt"].dt.to_period("M").astype(str)

        df["FA고지_c"] = df["FA고지"].fillna(" ").astype(str).str.strip()
        df["비교설명_c"] = df["비교설명"].fillna(" ").astype(str).str.strip()
        df["완전판매_c"] = df["완전판매"].fillna(" ").astype(str).str.strip()

        def is_scanned(val): return str(val).strip() in ["스캔", "M스캔", "보험사스캔"] if pd.notna(val) and str(val).strip() != " " else False
        def is_not_scanned(val): return str(val).strip() == "미스캔" if pd.notna(val) and str(val).strip() != " " else True
        def is_cs_target(val): return str(val).strip() in ["스캔", "M스캔", "미스캔"] if pd.notna(val) and str(val).strip() != " " else False

        df["FA_스캔"] = df["FA고지_c"].apply(is_scanned).astype(int)
        df["비교_스캔"] = df["비교설명_c"].apply(is_scanned).astype(int)
        df["완전판매_대상"] = df["완전판매_c"].apply(is_cs_target).astype(int)
        df["완전판매_스캔"] = df["완전판매_c"].apply(is_scanned).astype(int)

        df["FA_miss"] = (df["FA고지_c"] == "미스캔").astype(int)
        df["비교_miss"] = (df["비교설명_c"] == "미스캔").astype(int)
        df["완판_miss"] = (df["완전판매_c"] == "미스캔").astype(int)
        df["미스캔"] = df[["FA_miss", "비교_miss", "완판_miss"]].sum(axis=1)
        
        return df
    except Exception as e:
        st.error(f"❌ 엑셀 파일 읽기 오류: {e}")
        return pd.DataFrame()

def get_file_update_time():
    if os.path.exists(EXCEL_FILE):
        return datetime.fromtimestamp(os.path.getmtime(EXCEL_FILE)).strftime("%Y-%m-%d %H:%M:%S")
    return "알 수 없음"

# ==========================================
# 3. 집계 헬퍼
# ==========================================
def calculate_scan_stats(df_group):
    cnt = len(df_group)  # 대상건 = 증권개수
    
    # FA고지, 비교설명: 증번당 필수 대상
    FA_대상 = cnt
    비교_대상 = cnt
    # 완판: 해당사항없음 제외 대상
    완판_대상 = int(df_group["완전판매_대상"].sum())
    
    # 대상스캔건 = FA대상 + 비교대상 + 완판대상
    대상스캔건 = FA_대상 + 비교_대상 + 완판_대상
    
    # 스캔건수 (스캔, M스캔, 보험사스캔)
    FA_스캔 = int(df_group["FA_스캔"].sum())
    비교_스캔 = int(df_group["비교_스캔"].sum())
    완판_스캔 = int(df_group["완전판매_스캔"].sum())
    
    전체스캔 = FA_스캔 + 비교_스캔 + 완판_스캔
    전체미스캔 = 대상스캔건 - 전체스캔
    스캔율 = round((전체스캔 / 대상스캔건 * 100), 1) if 대상스캔건 > 0 else 0.0
    
    return {
        "증번수": cnt,
        "대상건": cnt,
        "대상스캔건": 대상스캔건,
        "전체스캔": 전체스캔,
        "전체미스캔": 전체미스캔,
        "스캔율": 스캔율,
        "FA_대상": FA_대상, "FA_스캔": FA_스캔, "FA_miss": FA_대상 - FA_스캔,
        "비교_대상": 비교_대상, "비교_스캔": 비교_스캔, "비교_miss": 비교_대상 - 비교_스캔,
        "완판_대상": 완판_대상, "완판_스캔": 완판_스캔, "완판_miss": 완판_대상 - 완판_스캔,
    }

def build_group_scan_stats(df, group_col):
    rows = []
    for org, df_group in df.groupby(group_col):
        s = calculate_scan_stats(df_group)
        rows.append({
            "조직": org, "대상건": s["대상건"], "대상스캔건": s["대상스캔건"], "전체스캔": s["전체스캔"],
            "총미스캔": s["전체미스캔"], "스캔율": s["스캔율"], "미처리율": round(100 - s["스캔율"], 1),
            "FA_miss": s["FA_miss"], "비교_miss": s["비교_miss"], "완판_miss": s["완판_miss"],
        })
    return pd.DataFrame(rows)

# ==========================================
# 4. 전체 계층 리포트 (KeyError 해결: 반환 컬럼 명시적 통일)
# ==========================================
@st.cache_data(ttl=300)
def build_hierarchy_report(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    
    has_sosok = "소속" in src.columns
    levels = ["부문", "총괄", "부서"] + (["소속"] if has_sosok else [])
    rows = []

    def recursive_agg(df_grp, level_idx, path_vals):
        if level_idx >= len(levels):  # 영업가족 상세
            for val, grp in df_grp.groupby("영업가족", dropna=False):
                s = calculate_scan_stats(grp)
                row = {c: path_vals[i] if i < len(path_vals) else "" for i, c in enumerate(levels)}
                row["영업가족"] = val
                row.update({"구분": "영업가족", "증번수": s["증번수"], "대상건": s["대상건"], 
                            "대상스캔건": s["대상스캔건"], "전체스캔": s["전체스캔"], 
                            "전체미스캔": s["전체미스캔"], "스캔율": s["스캔율"]})
                rows.append(row)
        else:
            col = levels[level_idx]
            for val, grp in df_grp.groupby(col, dropna=False):
                s = calculate_scan_stats(grp)
                row = {c: path_vals[i] if i < len(path_vals) else (val if i == level_idx else "") for i, c in enumerate(levels)}
                row["영업가족"] = ""
                if col == "부문": row["구분"] = "부문계"
                elif col == "총괄": row["구분"] = "총괄계"
                elif col == "부서": row["구분"] = "부서계"
                elif col == "소속": row["구분"] = "소속계"
                
                row.update({"증번수": s["증번수"], "대상건": s["대상건"], 
                            "대상스캔건": s["대상스캔건"], "전체스캔": s["전체스캔"], 
                            "전체미스캔": s["전체미스캔"], "스캔율": s["스캔율"]})
                rows.append(row)
                recursive_agg(grp, level_idx + 1, path_vals + [val])

    recursive_agg(src, 0, [])
    report = pd.DataFrame(rows)
    
    sort_cols = levels + ["영업가족"]
    report = report.sort_values(sort_cols, ascending=True).reset_index(drop=True)
    
    # 총계 행 추가
    ts = calculate_scan_stats(src)
    t_row = {col: "" for col in report.columns}
    t_row["구분"] = "총계"
    t_row.update({"증번수": ts["증번수"], "대상건": ts["대상건"], "대상스캔건": ts["대상스캔건"], 
                  "전체스캔": ts["전체스캔"], "전체미스캔": ts["전체미스캔"], "스캔율": ts["스캔율"]})
    report = pd.concat([report, pd.DataFrame([t_row])], ignore_index=True)
    return report

@st.cache_data(ttl=300)
def build_monthly_hierarchy(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    
    has_sosok = "소속" in src.columns
    idx = ["부문", "총괄", "부서"] + (["소속"] if has_sosok else [])
    
    agg = src.groupby(["월_피리어드"] + idx).agg(
        FA_miss=("FA_miss", "sum"), 비교_miss=("비교_miss", "sum"), 완판_miss=("완판_miss", "sum"),
        대상건=("증권번호", "count")
    ).reset_index()
    agg["총미스캔"] = agg[["FA_miss", "비교_miss", "완판_miss"]].sum(axis=1)
    agg["미처리율"] = round(agg["총미스캔"] / agg["대상건"] * 100, 1)
    agg["구분"] = "부서계"
    
    pivot_cols = ["총미스캔"]
    pivoted = agg.pivot_table(index=idx + ["구분"], columns="월_피리어드", values=pivot_cols, fill_value=0)
    pivoted.columns = [f"{m}_총미스캔" for m in pivoted.columns.get_level_values(1)]
    pivoted = pivoted.reset_index()
    
    sorted_months = sorted([c.split("_")[0] for c in pivoted.columns if "_총미스캔" in c])
    final_cols = idx + ["구분"] + [f"{m}_총미스캔" for m in sorted_months]
    return pivoted[final_cols]

# ==========================================
# 5. 관리대장 선정 대상
# ==========================================
@st.cache_data(ttl=300)
def get_ledger_targets(df, months):
    src = df[df["월_피리어드"].isin(months)].copy()
    if src.empty: return {}
    agg = src.groupby(["부문", "총괄", "부서", "영업가족"]).agg(
        FA=("FA_miss", "sum"), 비교=("비교_miss", "sum"),
        완판=("완판_miss", "sum"), 대상=("증권번호", "count")
    ).reset_index()
    agg["총미스캔"] = agg[["FA", "비교", "완판"]].sum(axis=1)
    agg = agg[agg["총미스캔"] > 0]
    return {dept: grp for dept, grp in agg.groupby("부서")}

# ==========================================
# 6. 한글 폰트 및 스타일
# ==========================================
@st.cache_resource
def register_korean_font():
    font_candidates = [
        ("NotoSansKR", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        ("NotoSansKR", "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
        ("Malgun", r"C:\Windows\Fonts\malgun.ttf"),
        ("DejaVu", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for name, path in font_candidates:
        try:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont(name, path))
                return name
        except Exception: continue
    st.warning("⚠️ 한글 폰트를 찾을 수 없습니다. PDF/Excel에서 한글이 깨질 수 있습니다.")
    return "Helvetica"

HDR_CLR = "#4472C4"
def _pdf_styles(fn):
    S = getSampleStyleSheet()
    def ps(n, **kw): return ParagraphStyle(n, parent=S["Normal"], fontName=fn, **kw)
    return {
        "title":   ps("T", fontSize=15, bold=True, alignment=1, spaceAfter=4),
        "sub":     ps("S", fontSize=10, spaceAfter=3),
        "body":    ps("B", fontSize=8,  spaceAfter=2),
        "notice":  ps("N", fontSize=7.5, spaceAfter=3, textColor=colors.HexColor("#CC0000"), alignment=1),
        "date":    ps("D", fontSize=8,  alignment=2, spaceAfter=4),
        "section": ps("SC", fontSize=9, bold=True, spaceAfter=2),
    }

def _tbl(data, cw, fn, header_rows=1, sub_rows=None, align="CENTER"):
    if not data or len(data) < 1: return Spacer(1,0)
    cw_scaled = [w * 1.7 for w in cw]
    align_map = {"LEFT":0, "CENTER":1, "RIGHT":2}
    align_value = align_map.get(align.upper(), 1)
    S = getSampleStyleSheet()
    cell_style = ParagraphStyle("tbl_cell", parent=S["Normal"], fontName=fn, fontSize=8, leading=10, alignment=align_value, wordWrap="CJK")
    wrapped_data = [[Paragraph(str(cell), cell_style) if not isinstance(cell, Paragraph) else cell for cell in row] for row in data]
    t = Table(wrapped_data, colWidths=cw_scaled, repeatRows=header_rows)
    cmds = [
        ("FONTNAME", (0,0),(-1,-1), fn), ("FONTSIZE", (0,0),(-1,-1), 8),
        ("ALIGN", (0,0),(-1,-1), align.upper()), ("VALIGN", (0,0),(-1,-1), "MIDDLE"),
        ("WORDWRAP", (0,0),(-1,-1), "CJK"), ("GRID", (0,0),(-1,-1), 0.4, colors.grey),
        ("LEFTPADDING", (0,0),(-1,-1), 2), ("RIGHTPADDING", (0,0),(-1,-1), 2),
        ("TOPPADDING", (0,0),(-1,-1), 2), ("BOTTOMPADDING",(0,0),(-1,-1), 2),
        ("BACKGROUND", (0,0),(-1,header_rows-1), colors.HexColor("#DCE6F1")),
        ("TEXTCOLOR", (0,0),(-1,header_rows-1), colors.HexColor("#1F3864")),
    ]
    for i in range(header_rows, len(data)):
        if (i-header_rows)%2==1: cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F3F6FA")))
        if sub_rows and i in sub_rows: cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#E9EEF8")))
    t.setStyle(TableStyle(cmds)); return t

def _sig_table(labels, fn, cw=120):
    t = Table([labels,["____________________"]*len(labels),["(인)"]*len(labels)], colWidths=[cw*1.4]*len(labels))
    t.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1), "CENTER"),("FONTNAME",(0,0),(-1,-1),fn),
        ("FONTSIZE",(0,0),(-1,-1),8.5),("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),("BOX",(0,0),(-1,-1),0.5,colors.grey),
        ("INNERGRID",(0,0),(-1,-1),0.3,colors.lightgrey)]))
    return t

def plotly_to_image(fig, width_mm=160, height_mm=90):
    try:
        img_bytes = pio.to_image(fig, format='png', width=1200, height=600, scale=2)
        return Image(io.BytesIO(img_bytes), width=width_mm*mm, height=height_mm*mm)
    except Exception:
        return Paragraph("<i>차트 렌더링 환경 미준비 (kaleido 필요)</i>", _pdf_styles(register_korean_font())["notice"])

# ==========================================
# 7. 전체 계층 리포트 Excel
# ==========================================
def report_excel(df, months):
    wb = Workbook(); ws = wb.active; ws.title="계층별_미처리현황"
    tfn, hf, bf = "맑은 고딕", Font(name="맑은 고딕",size=9,bold=True,color="FFFFFF"), Font(name="맑은 고딕",size=9)
    bdr = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
    fills = {"부문계":PatternFill("solid",fgColor="1F3864"), "총괄계":PatternFill("solid",fgColor="2E75B6"),
             "부서계":PatternFill("solid",fgColor="D9E1F2"), "소속계":PatternFill("solid",fgColor="E2EFDA"), "영업가족_alt":PatternFill("solid",fgColor="EEF3FB")}
    fonts_wc = {"부문계":Font(name=tfn,size=9,bold=True,color="FFFFFF"), "총괄계":Font(name=tfn,size=9,bold=True,color="FFFFFF"),
                "부서계":Font(name=tfn,size=9,bold=True), "소속계":Font(name=tfn,size=9,bold=True)}
    h_fill = PatternFill("solid",fgColor="4472C4")
    today, period_str = datetime.now().strftime("%Y년 %m월 %d일"), ", ".join(months) if months else "전체"
    
    has_sosok = "소속" in df.columns
    headers = ["구분", "부문", "총괄", "부서"] + (["소속"] if has_sosok else []) + ["영업가족", "증번수", "대상건", "대상스캔건", "전체스캔", "전체미스캔", "스캔율"]
    cws = [14,20,20,20] + ([18] if has_sosok else []) + [24,12,12,14,14,12,16]
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"] = f"서류 미처리 현황 계층별 집계  ·  기간: {period_str}  ·  발급: {today}"
    ws["A1"].font = Font(name=tfn,size=12,bold=True); ws["A1"].alignment = Alignment(horizontal="center"); ws.row_dimensions[1].height = 22
    
    for ci,(h,w) in enumerate(zip(headers,cws),1):
        c=ws.cell(2,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width=w

    report = build_hierarchy_report(df, months)
    if report.empty: return io.BytesIO()
    ri = 3
    for _, row in report.iterrows(): 
        gbn = row["구분"]; rate_str = f"{row['스캔율']:.1f}%"
        vals = [gbn, row["부문"], row["총괄"], row["부서"]] + ([row.get("소속","")] if has_sosok else []) + [row["영업가족"], row["증번수"], row["대상건"], row["대상스캔건"], row["전체스캔"], row["전체미스캔"], rate_str]
        fill = fills.get(gbn, fills["영업가족_alt"] if ri%2==0 else None)
        fnt  = fonts_wc.get(gbn, bf)
        for ci,v in enumerate(vals,1):
            c=ws.cell(ri,ci,v); c.font=fnt; c.border=bdr; c.alignment=Alignment(horizontal="center",vertical="center")
            if isinstance(v,(int,float)): c.number_format = "#,##0"
            if fill: c.fill=fill
        ri += 1

    monthly = build_monthly_hierarchy(df, months)
    if not monthly.empty:
        ws2 = wb.create_sheet("월별_계층집계(피벗)")
        ws2.merge_cells(f"A1:{get_column_letter(len(monthly.columns))}1")
        ws2["A1"] = f"월별 계층 미처리 집계(가로피벗)  ·  기간: {period_str}"
        ws2["A1"].font = Font(name=tfn,size=12,bold=True); ws2["A1"].alignment=Alignment(horizontal="center")
        for ci,h in enumerate(monthly.columns,1):
            c=ws2.cell(2,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center")
            ws2.column_dimensions[get_column_letter(ci)].width = 16
        for ri2, r in enumerate(monthly.itertuples(index=False), 3):
            for ci,v in enumerate(r,1):
                c=ws2.cell(ri2,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                if isinstance(v,(int,float)): c.number_format = "#,##0"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ==========================================
# 8. 전체 계층 리포트 PDF
# ==========================================
def report_pdf(df, months):
    fn, st_, buf = register_korean_font(), _pdf_styles(register_korean_font()), io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=12*mm,leftMargin=12*mm, topMargin=12*mm,bottomMargin=12*mm)
    today, period_str = datetime.now().strftime("%Y년 %m월 %d일"), ", ".join(months) if months else "전체"
    E = [Paragraph("서류 미처리 현황 계층별 집계 ", st_["title"]), Paragraph(f"기간: {period_str}  |  발급일자: {today} ", st_["date"]), HRFlowable(width="100%",thickness=1,color=colors.HexColor(HDR_CLR)), Spacer(1,6)]
    
    report = build_hierarchy_report(df, months)
    if not report.empty:
        has_sosok = "소속" in report.columns
        hdr = [["구분", "부문", "총괄", "부서"] + (["소속"] if has_sosok else []) + ["영업가족", "증번수", "대상건", "대상스캔건", "전체스캔", "전체미스캔", "스캔율"]]
        drows, sub_idx = [], []
        for i,(_,r) in enumerate(report.iterrows()):
            drows.append([r["구분"], r["부문"], r["총괄"], r["부서"]] + ([r.get("소속","")] if has_sosok else []) + [r["영업가족"],
                        f"{int(r['증번수']):,}", f"{int(r['대상건']):,}", f"{int(r['대상스캔건']):,}", f"{int(r['전체스캔']):,}", f"{int(r['전체미스캔']):,}", f"{r['스캔율']:.1f}%"])
            if r["구분"] in ("부문계", "총괄계", "부서계", "소속계"): sub_idx.append(i+1)
        col_w = [24,32,32,32] + ([22] if has_sosok else []) + [45,22,24,24,24,24,22]
        E.append(_tbl(hdr+drows, col_w, fn, sub_rows=sub_idx)); E.append(Spacer(1,8))
        
    monthly = build_monthly_hierarchy(df, months)
    if not monthly.empty:
        E.append(PageBreak()); E.append(Paragraph("▶ 월별 계층별 미처리 집계 (가로피벗)", st_["section"]))
        mrows = [list(r) for r in monthly.itertuples(index=False)]
        msb = [i+1 for i in range(len(monthly))]
        E.append(_tbl([list(monthly.columns)]+mrows, [18]*len(monthly.columns), fn, sub_rows=msb))
    doc.build(E); buf.seek(0); return buf

# ==========================================
# 8-2. 현재 세팅 기준 전체 페이지 PDF (차트 이미지 삽입 복구)
# ==========================================
def report_fullpage_pdf(df, months, agg_group, map_level, dash_doc_types=None, dash_chart_mode="그룹형", dash_top_n=15, map_type="🔲 트리맵"):
    fn, st_, buf = register_korean_font(), _pdf_styles(register_korean_font()), io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    today, period_str = datetime.now().strftime("%Y년 %m월 %d일"), ", ".join(months) if months else "전체"
    E = [
        Paragraph("전체 페이지 요약 리포트 ", st_["title"]),
        Paragraph(f"기간: {period_str}  |  발급일자: {today} ", st_["date"]),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor(HDR_CLR)), Spacer(1, 6),
    ]
    df_sel = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    stats_all = calculate_scan_stats(df_sel)
    miss_rate = round(100 - stats_all["스캔율"], 1)
    
    E.append(Paragraph("▶ 주요 KPI ", st_["section"]))
    E.append(_tbl([
        ["총 계약건수", f"{stats_all['증번수']:,}"], ["스캔대상건수", f"{stats_all['대상스캔건']:,}"],
        ["전체스캔", f"{stats_all['전체스캔']:,}"], ["전체미스캔", f"{stats_all['전체미스캔']:,}"],
        ["스캔율 / 미처리율", f"{stats_all['스캔율']:.1f}% / {miss_rate:.1f}%"],
    ], [90, 120], fn, header_rows=0, align="LEFT"))
    E.append(Spacer(1, 8))

    agg = build_group_scan_stats(df_sel, agg_group).sort_values("총미스캔", ascending=False).head(dash_top_n)
    if not agg.empty:
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(x=agg["조직"], y=agg["총미스캔"], marker_color="#FF6B6B", text=agg["총미스캔"], textposition="outside"))
        fig1.update_layout(title=f"미처리 건수 TOP {dash_top_n} ({agg_group})", margin=dict(l=10, r=10, t=30, b=5), font=dict(size=10))
        E.append(plotly_to_image(fig1))
        
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=agg["조직"], y=agg["총미스캔"], mode="lines+markers", line=dict(color="#2E75B6", width=3), marker=dict(size=6)))
        fig2.update_layout(title="미처리 추이 라인차트", margin=dict(l=10, r=10, t=30, b=5))
        E.append(Spacer(1, 5))
        E.append(plotly_to_image(fig2))
        E.append(Spacer(1, 8))

    E.append(Paragraph("▶ 계층별 상세 데이터 ", st_["section"]))
    report = build_hierarchy_report(df, months)
    if not report.empty:
        has_sosok = "소속" in report.columns
        hdr = [["구분", "부문", "총괄", "부서"] + (["소속"] if has_sosok else []) + ["영업가족", "증번수", "대상건", "대상스캔건", "전체스캔", "전체미스캔", "스캔율"]]
        drows = [[r["구분"], r["부문"], r["총괄"], r["부서"]] + ([r.get("소속","")] if has_sosok else []) + [r["영업가족"],
                    f"{int(r['증번수']):,}", f"{int(r['대상건']):,}", f"{int(r['대상스캔건']):,}", f"{int(r['전체스캔']):,}", f"{int(r['전체미스캔']):,}", f"{r['스캔율']:.1f}%"] for _, r in report.iterrows()]
        col_w = [24,32,32,32] + ([22] if has_sosok else []) + [45,22,24,24,24,24,22]
        E.append(_tbl(hdr + drows, col_w, fn))
    doc.build(E); buf.seek(0); return buf

# ==========================================
# 9. 관리대장 PDF & 10. Excel (기존 유지)
# ==========================================
def ledger_pdf(families_by_dept, period_text, df_src):
    fn, st_, buf = register_korean_font(), _pdf_styles(register_korean_font()), io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=12*mm,leftMargin=12*mm, topMargin=15*mm,bottomMargin=15*mm)
    today = datetime.now().strftime("%Y년 %m월 %d일"); E = []
    center_date_style = ParagraphStyle("CenterDate", parent=st_["date"], alignment=1)
    title_left = ParagraphStyle("TitleLeft", parent=st_["title"], alignment=0)
    indent_style = ParagraphStyle("IndentSub", parent=st_["sub"], leftIndent=8, alignment=0, spaceAfter=2)
    section_left = ParagraphStyle("SectionLeft", parent=st_["section"], alignment=0)
    notice_left = ParagraphStyle("NoticeLeft", parent=st_["notice"], leftIndent=8, alignment=0)
    date_indent = ParagraphStyle("DateIndent", parent=st_["date"], leftIndent=8, alignment=0)
    
    for dept_name, grp_df in families_by_dept.items():
        sec, tg = grp_df.iloc[0]["부문"], grp_df.iloc[0]["총괄"]
        E += [Paragraph("신계약 필수서류 미처리 확인서 ", title_left), HRFlowable(width="100%",thickness=1.5,color=colors.HexColor(HDR_CLR)), Spacer(1,4),
              Paragraph(f"부서: {sec}   > {tg}   >   {dept_name} ", indent_style), Paragraph(f"적용기간: {period_text} ", date_indent), Spacer(1,6)]
        dept_src = df_src[df_src["부서"]==dept_name]
        E += [Paragraph("【필수 서류 상세 안내】 ", st_["section"]),
              _tbl(REQUIRED_DOCS_TABLE, [12, 60, 90, 198], fn, header_rows=1, align="LEFT"), Spacer(1,8),
              Paragraph(GUIDANCE_TEXT, notice_left), Spacer(1,8)]
        if not dept_src.empty:
            E.append(Paragraph("▶ 영업가족별 · 월별 · 양식별 미처리 현황 ", section_left))
            fam_mon = dept_src.groupby(["영업가족", "월_피리어드"]).agg(FA=("FA_miss", "sum"),비교=("비교_miss", "sum"),완판=("완판_miss", "sum")).reset_index()
            fam_mon["계"] = fam_mon[["FA", "비교", "완판"]].sum(axis=1); fam_mon = fam_mon[fam_mon["계"] > 0]
            if not fam_mon.empty:
                td=[["영업가족", "월", "FA고지", "비교설명", "완전판매", "계"]]
                for _, r in fam_mon.iterrows(): td.append([r["영업가족"], r["월_피리어드"], f"{int(r.FA):,}", f"{int(r.비교):,}", f"{int(r.완판):,}", f"{int(r['계']):,}"])
                E.append(_tbl(td,[130,50,45,45,45,45],fn, align="LEFT")); E.append(Spacer(1,4))
        E += [Paragraph(PRECAUTION_TEXT_COVER, notice_left), Spacer(1,4), Paragraph(PRECAUTION_TEXT_CONFIRM, notice_left), Spacer(1,4),
              Paragraph(SIGNATURE_CONFIRMATION_TEXT, notice_left), Spacer(1,8),
              Paragraph("작성일: _______________ ", center_date_style), Spacer(1,4), _sig_table(["부문장 확인", "총괄 확인", "부서장 확인"],fn,120), PageBreak()]
        for _, fam in grp_df.drop_duplicates("영업가족").iterrows():
            fam_name = fam["영업가족"]
            E += [Paragraph("신계약 필수서류 미처리 확인서 ", title_left), HRFlowable(width="100%",thickness=1.5,color=colors.HexColor(HDR_CLR)), Spacer(1,4),
                  Paragraph(f"소속: {sec}   > {tg}   > {dept_name}   >   {fam_name} ", indent_style), Paragraph(f"적용기간: {period_text} ", date_indent), Spacer(1,6)]
            fam_src = df_src[(df_src["영업가족"]==fam_name) & df_src["소속"].notna()]
            sosok = fam_src.groupby(["소속", "월_피리어드"]).agg(FA=("FA_miss", "sum"),비교=("비교_miss", "sum"),완판=("완판_miss", "sum")).reset_index()
            sosok["계"] = sosok[["FA", "비교", "완판"]].sum(axis=1); sosok = sosok[sosok["계"] > 0]
            E.append(Paragraph("▶ 소속별 · 월별 · 양식별 미처리 건수 ", section_left))
            if not sosok.empty:
                td2=[["소속", "월", "FA고지", "비교설명", "완전판매", "계"]]
                for _,r in sosok.iterrows(): td2.append([r["소속"], r["월_피리어드"], f"{int(r.FA):,}", f"{int(r.비교):,}", f"{int(r.완판):,}", f"{int(r['계']):,}"])
                E.append(_tbl(td2,[130,50,45,45,45,45],fn, align="LEFT")); E.append(Spacer(1,4))
            else: E.append(Paragraph("(해당 데이터 없음) ", st_["body"]))
            E.append(Spacer(1,6))
            sum_d=[["FA고지", "비교설명", "완전판매", "총계"],[f"{int(fam['FA']):,}",f"{int(fam['비교']):,}",f"{int(fam['완판']):,}",f"{int(fam['총미스캔']):,}"]]
            E += [Paragraph("▶ 양식별 미처리 요약 ", section_left), _tbl(sum_d,[90,90,90,90],fn, align="LEFT"), Spacer(1,8),
                  Paragraph("【필수 서류 상세 안내】 ", st_["section"]), _tbl(REQUIRED_DOCS_TABLE, [12, 60, 90, 198], fn, header_rows=1, align="LEFT"), Spacer(1,8),
                  Paragraph(GUIDANCE_TEXT, notice_left), Spacer(1,8), Paragraph(PRECAUTION_TEXT_COVER, notice_left), Spacer(1,4),
                  Paragraph(PRECAUTION_TEXT_SHEET, notice_left), Spacer(1,8), Paragraph("작성일: _______________ ", center_date_style)]
            sig2=Table([[f"영업가족대표 서명: ____________________ (인)"]], colWidths=[120*1.4*3])
            sig2.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1), "LEFT"),("FONTNAME",(0,0),(-1,-1),fn),("FONTSIZE",(0,0),(-1,-1),9.5),
                ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),("BOX",(0,0),(-1,-1),0.5,colors.grey)]))
            E += [sig2, PageBreak()]
    doc.build(E); buf.seek(0); return buf

def ledger_excel(families_by_dept, period_text, df_src):
    wb = Workbook(); ws0 = wb.active; ws0.title="목차"
    tfn = "맑은 고딕"; hf = Font(name=tfn,size=9,bold=True,color="FFFFFF"); bf = Font(name=tfn,size=9)
    nf = Font(name=tfn,size=8,italic=True,color="CC0000"); sig_f = Font(name=tfn,size=9,bold=True)
    bdr, h_fill, alt_fill = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin")), PatternFill("solid",fgColor="4472C4"), PatternFill("solid",fgColor="EEF3FB")
    today = datetime.now().strftime("%Y년 %m월 %d일")
    ws0.merge_cells("A1:F1"); ws0["A1"]=f"관리대장 목차  ·  {period_text}  ·  발급: {today}"
    ws0["A1"].font=Font(name=tfn,size=13,bold=True); ws0["A1"].alignment=Alignment(horizontal="center")
    for ci,h in enumerate(["부서", "영업가족", "FA고지", "비교설명", "완전판매", "총미스캔"],1):
        c=ws0.cell(3,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center")
    ir=4
    for dept,grp in families_by_dept.items():
        for _,fam in grp.drop_duplicates("영업가족").iterrows():
            for ci,v in enumerate([dept,fam["영업가족"],int(fam["FA"]),int(fam["비교"]),int(fam["완판"]),int(fam["총미스캔"])],1):
                c=ws0.cell(ir,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                if isinstance(v,(int,float)): c.number_format = "#,##0"
                if ir%2==0: c.fill=alt_fill
            ir+=1
    for ci,w in enumerate([22,25,13,13,13,14],1): ws0.column_dimensions[get_column_letter(ci)].width=w
    
    for dept_name,grp_df in families_by_dept.items():
        sec, tg = grp_df.iloc[0]["부문"], grp_df.iloc[0]["총괄"]
        sname=f"표지 {dept_name[:10]}".replace("/", " "); ws_c=wb.create_sheet(title=sname)
        ws_c.merge_cells("A1:G1"); ws_c["A1"]=f"[{dept_name}]  신계약 필수서류 미처리 확인서 "; ws_c["A1"].font=Font(name=tfn,size=14,bold=True)
        ws_c["A2"]=f"{sec}   > {tg}   > {dept_name} "; ws_c["A2"].font=Font(name=tfn,size=10)
        ws_c["A3"]=f"    적용기간: {period_text} "; ws_c["A3"].font=bf; ws_c["A3"].alignment=Alignment(horizontal="left")
        r=5; ws_c.cell(r,1,"【필수 서류 상세 안내】 ").font=Font(name=tfn,size=10,bold=True); r+=1
        for ci, header in enumerate(["No. ",  "서류명 ",  "법적 근거 ",  "목적 및 주요 내용 "], 1):
            c = ws_c.cell(r, ci, header); c.font = hf; c.fill = h_fill; c.border = bdr; c.alignment = Alignment(horizontal="center", vertical="center")
        ws_c.column_dimensions[get_column_letter(1)].width = 6; ws_c.column_dimensions[get_column_letter(2)].width = 20
        ws_c.column_dimensions[get_column_letter(3)].width = 25; ws_c.column_dimensions[get_column_letter(4)].width = 45
        r += 1
        docs_data = [["1",  "개인정보동의서",  "개인정보보호법 15조 및\n대리점 표준 내부통제기준 27조", "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에 따른 개인정보 처리로 신계약시 필수 징구"],
                     ["2",  "비교설명확인서",  "보험업감독규정\n별표 5-6", "유사 상품 3개 이상 비교·설명 이행 사실 고객 확인 서명"],
                     ["3",  "고지의무확인서",  "금융소비자보호법 26조와\n동법시행령 24조", "판매자 권한·책임·보상 관련 핵심 사항 고지, 소비자 소인 예방"],
                     ["4",  "완전판매확인서\n(대상: 종신, CI, CEO경기, 고액)",  "금융소비자보호법 제17·19조\n영업지원기준안", "약관,청약서 부본 제공, 중요 상품 이해 및 자발적 가입 확인, 설명 의무 이행 증빙력 확보"]]
        for row_data in docs_data:
            for ci, val in enumerate(row_data, 1):
                c = ws_c.cell(r, ci, val); c.font = bf; c.border = bdr; c.alignment = Alignment(horizontal="left" if ci > 1 else "center", vertical="top", wrapText=True)
            ws_c.row_dimensions[r].height = 35; r += 1
        r += 1; ws_c.cell(r,1,GUIDANCE_TEXT).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=45; r+=2
        ws_c.cell(r,1,"▶ 영업가족별 · 월별 · 양식별 미처리 현황 ").font=Font(name=tfn,size=10,bold=True); r+=1
        dept_src=df_src[df_src["부서"]==dept_name]
        if not dept_src.empty:
            fam_mon=dept_src.groupby(["영업가족", "월_피리어드"]).agg(FA=("FA_miss", "sum"),비교=("비교_miss", "sum"),완판=("완판_miss", "sum")).reset_index()
            fam_mon["계"]=fam_mon[["FA", "비교", "완판"]].sum(axis=1); fam_mon=fam_mon[fam_mon["계"] > 0]
            hdrs, cws = ["영업가족", "월", "FA고지", "비교설명", "완전판매", "계"], [25,20,13,13,13,13]
            for ci,(h,w) in enumerate(zip(hdrs,cws),1):
                c=ws_c.cell(r,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center"); ws_c.column_dimensions[get_column_letter(ci)].width=w
            for i,(_,rv) in enumerate(fam_mon.iterrows()):
                row_v=[rv["영업가족"],rv["월_피리어드"],int(rv.FA),int(rv.비교),int(rv.완판),int(rv["계"])]; af=alt_fill if i%2==1 else None
                for ci,v in enumerate(row_v,1):
                    c=ws_c.cell(r+1+i,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                    if isinstance(v,(int,float)): c.number_format = "#,##0"
                    if af: c.fill=af
            r+=len(fam_mon)+2
        ws_c.cell(r,1,PRECAUTION_TEXT_COVER).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=35; r+=2
        ws_c.cell(r,1,PRECAUTION_TEXT_CONFIRM).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=35; r+=2
        ws_c.cell(r,1,"작성일: _______________ ").font=bf; r+=2
        for i,sig in enumerate(["부문장 확인", "총괄 확인", "부서장 확인"]):
            ws_c.cell(r,i*2+1,sig).font=sig_f; ws_c.cell(r,i*2+2,"________________ (인) ").font=Font(name=tfn,color="888888")
        for _,fam in grp_df.drop_duplicates("영업가족").iterrows():
            fam_name=fam["영업가족"]; fn_safe=fam_name[:14].replace("/", "_").replace("   ", " ")
            ws_f=wb.create_sheet(title=fn_safe); ws_f.merge_cells("A1:G1"); ws_f["A1"]=f"[{fam_name}]  신계약 필수서류 미처리 확인서 "; ws_f["A1"].font=Font(name=tfn,size=13,bold=True)
            ws_f["A2"]=f"{sec}   > {tg}   > {dept_name}   > {fam_name} "; ws_f["A2"].font=Font(name=tfn,size=9,italic=True)
            ws_f["A3"]=f"    적용기간: {period_text} "; ws_f["A3"].font=bf; ws_f["A3"].alignment=Alignment(horizontal="left")
            r_f=5; ws_f.cell(r_f,1,"▶ 소속별 · 월별 · 양식별 미처리 건수 ").font=Font(name=tfn,size=10,bold=True); r_f+=1
            fam_src=df_src[(df_src["영업가족"]==fam_name) & df_src["소속"].notna()]
            sosok=fam_src.groupby(["소속", "월_피리어드"]).agg(FA=("FA_miss", "sum"),비교=("비교_miss", "sum"),완판=("완판_miss", "sum")).reset_index()
            sosok["계"]=sosok[["FA", "비교", "완판"]].sum(axis=1); sosok=sosok[sosok["계"] > 0]
            sh, sc = ["소속", "월", "FA고지", "비교설명", "완전판매", "계"], [25,20,13,13,13,13]
            for ci,(h,w) in enumerate(zip(sh,sc),1): c=ws_f.cell(r_f,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center"); ws_f.column_dimensions[get_column_letter(ci)].width=w
            if not sosok.empty:
                for i,(_,sr) in enumerate(sosok.iterrows()):
                    rv2=[sr["소속"],sr["월_피리어드"],int(sr.FA),int(sr.비교),int(sr.완판),int(sr["계"])]; af=alt_fill if i%2==1 else None
                    for ci,v in enumerate(rv2,1):
                        c=ws_f.cell(r_f+1+i,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                        if isinstance(v,(int,float)): c.number_format = "#,##0"
                        if af: c.fill=af
                r_f += len(sosok) + 2
            else: r_f += 1
            ws_f.cell(r_f,1,"▶ 양식별 요약 ").font=Font(name=tfn,size=10,bold=True); r_f+=1
            for ci,h in enumerate(["FA고지", "비교설명", "완전판매", "총계"],1): c=ws_f.cell(r_f,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center")
            for ci,v in enumerate([int(fam["FA"]),int(fam["비교"]),int(fam["완판"]),int(fam["총미스캔"])],1):
                c=ws_f.cell(r_f+1,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                if isinstance(v,(int,float)): c.number_format = "#,##0"
            r_f+=3
            ws_f.cell(r_f,1,PRECAUTION_TEXT_COVER).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=30; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_SHEET).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=30; r_f+=2
            ws_f.cell(r_f,1,"【필수 서류 상세 안내】 ").font=Font(name=tfn,size=10,bold=True); r_f+=1
            for ci, header in enumerate(["No. ",  "서류명 ",  "법적 근거 ",  "목적 및 주요 내용 "], 1):
                c = ws_f.cell(r_f, ci, header); c.font = hf; c.fill = h_fill; c.border = bdr; c.alignment = Alignment(horizontal="center", vertical="center")
            ws_f.column_dimensions[get_column_letter(1)].width = 6; ws_f.column_dimensions[get_column_letter(2)].width = 20
            ws_f.column_dimensions[get_column_letter(3)].width = 25; ws_f.column_dimensions[get_column_letter(4)].width = 45
            r_f += 1
            for row_data in docs_data:
                for ci, val in enumerate(row_data, 1):
                    c = ws_f.cell(r_f, ci, val); c.font = bf; c.border = bdr; c.alignment = Alignment(horizontal="left" if ci > 1 else "center", vertical="top", wrapText=True)
                ws_f.row_dimensions[r_f].height = 35; r_f += 1
            r_f += 1
            ws_f.cell(r_f,1,GUIDANCE_TEXT).font=nf; ws_f.cell(r_f,1).alignment= Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=45; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_COVER).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=35; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_SHEET).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=35; r_f+=2
            ws_f.cell(r_f,1,"작성일: _______________ ").font=bf; r_f+=1
            ws_f.cell(r_f,1,"영업가족대표 서명: ________________ (인) ").font=sig_f
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ==========================================
# 11. UI – 단일 비밀번호 로그인
# ==========================================
def login_page():
    st.title("🔐 시스템 접속")
    st.markdown("단일 비밀번호로 접속합니다.")
    pwd = st.text_input("접속 비밀번호를 입력하세요", type="password")
    if st.button("접속하기", use_container_width=True, type="primary"):
        if pwd == APP_PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")

# ==========================================
# 12. 통합 대시보드
# ==========================================
def dashboard_page():
    st.title("📊 서류 처리 현황 대시보드")
    df = load_data()

    if df.empty:
        st.warning("📭 데이터가 없습니다. GitHub에 'insurance_data.xlsx' 파일을 업로드해주세요.")
        st.info("""
**데이터 업로드 방법:**
1. GitHub 저장소에 `insurance_data.xlsx` 파일 업로드
2. Git commit & push
3. Streamlit Cloud가 자동으로 재배포 (1-2분 소요)
4. 페이지 새로고침
""")
        return

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1: st.success(f"✅ 총 **{len(df):,}건**의 데이터 로드 완료")
    with col2: st.info(f"📅 기준: **{get_file_update_time()}**")
    with col3:
        if st.button("🔄 새로고침"):
            st.cache_data.clear()
            st.rerun()

    all_months = sorted(df["월_피리어드"].dropna().unique())
    st.subheader("📅 분석 기간 선택")
    sel_months = st.multiselect("월 선택 (복수 가능)", all_months, default=[all_months[-1]] if all_months else [])

    if not sel_months:
        st.warning("⚠️ 최소 1개 이상의 월을 선택해주세요.")
        return

    period_text = f"{sel_months[0]} ~ {sel_months[-1]}" if len(sel_months) > 1 else sel_months[0]
    df_sel = df[df["월_피리어드"].isin(sel_months)].copy()

    if df_sel.empty:
        st.info("선택한 기간에 데이터가 없습니다.")
        return

    stats_all = calculate_scan_stats(df_sel)
    fa_t, bi_t, cs_t = int(df_sel["FA_miss"].sum()), int(df_sel["비교_miss"].sum()), int(df_sel["완판_miss"].sum())
    miss_rate = round(100 - stats_all["스캔율"], 1)
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("📄 총 계약건수", f"{stats_all['증번수']:,}건")
    m2.metric("🧮 대상스캔건", f"{stats_all['대상스캔건']:,}건")
    m3.metric("📈 스캔율", f"{stats_all['스캔율']:.1f}%")
    m4.metric("📉 미처리율", f"{miss_rate:.1f}%")
    st.caption(f"총 미처리건수: {stats_all['전체미스캔']:,}건 | FA / 비교 / 완판: {fa_t:,} / {bi_t:,} / {cs_t:,} | 개인정보동의서 집계 제외")
    st.divider()

    tab_dash, tab_map, tab_report, tab_ledger = st.tabs(["📈 현황 대시보드", "🗺️ 미처리맵", "📊 계층 리포트", "📋 관리대장 출력"])

    # ── TAB 1 : 현황 대시보드 ─────────────────────────
    with tab_dash:
        cs1, cs2 = st.columns([2, 1])
        with cs1: search_text = st.text_input("🔍 조직 검색", placeholder="조직명 입력...")
        with cs2: agg_group = st.selectbox("집계 기준 (랭킹 단위)", ["부문", "총괄", "부서", "영업가족"], key="agg_group")
        agg = build_group_scan_stats(df_sel, agg_group)
        if search_text: agg = agg[agg["조직"].astype(str).str.contains(search_text, case=False, na=False)]
        agg = agg.sort_values("총미스캔", ascending=False).reset_index(drop=True); agg.insert(0, "순위",range(1,len(agg)+1))
        if agg.empty: st.info("조건에 맞는 데이터가 없습니다.")
        else:
            st.dataframe(
                agg[["순위", "조직", "대상건", "대상스캔건", "전체스캔", "총미스캔", "스캔율", "미처리율", "FA_miss", "비교_miss", "완판_miss"]]
                .style.format({
                    "순위": "{:,}", "대상건": "{:,}", "대상스캔건": "{:,}", "전체스캔": "{:,}",
                    "총미스캔": "{:,}", "스캔율": "{:.1f}%", "미처리율": "{:.1f}%",
                    "FA_miss": "{:,}", "비교_miss": "{:,}", "완판_miss": "{:,}"
                }), use_container_width=True, hide_index=True)
            top_n = st.slider("차트 표시 개수", 5, 30, 30, key="dash_top_n"); top = agg.head(top_n)
            c1, c2 = st.columns(2)
            with c1:
                doc_types = st.multiselect("표시 서류", ["FA고지", "비교설명", "완전판매", "총 미스캔"], default=["총 미스캔"], key="dash_doc_types")
                if doc_types:
                    max_v, yr = top["총미스캔"].max(), [0, top["총미스캔"].max()*1.2] if top["총미스캔"].max() >0 else [0,10]
                    if len(doc_types)==1 and doc_types[0]=="총 미스캔":
                        fig = go.Figure(); fig.add_trace(go.Bar(x=top["조직"], y=top["총미스캔"], text=top["총미스캔"], textposition="outside", marker_color=top["총미스캔"], marker_colorscale="Reds"))
                        fig.update_layout(title=f"미처리 건수 TOP {top_n}", xaxis_tickangle=-45, yaxis=dict(range=yr), height=420); st.plotly_chart(fig, use_container_width=True)
                    elif len(doc_types)==1:
                        cm = {"FA고지": "FA_miss", "비교설명": "비교_miss", "완전판매": "완판_miss"}
                        fig = px.bar(top, x="조직", y=cm[doc_types[0]], title=f"{doc_types[0]} 미스캔 TOP {top_n}", text=cm[doc_types[0]], color=cm[doc_types[0]], color_continuous_scale="Blues")
                        fig.update_layout(xaxis_tickangle=-45, height=420); st.plotly_chart(fig, use_container_width=True)
                    else:
                        ct = st.radio("차트 유형", ["그룹형", "누적형"], horizontal=True, key="dash_chart_mode")
                        cm2 = {"FA고지": "FA_miss", "비교설명": "비교_miss", "완전판매": "완판_miss", "총 미스캔": "총미스캔"}
                        p = top[["조직"]+[cm2[d] for d in doc_types]].copy(); p.columns=["조직"]+doc_types; p=p.melt("조직",var_name="종류",value_name="건수")
                        fig = px.bar(p, x="조직", y="건수", color="종류", barmode="group" if ct=="그룹형" else "stack", color_discrete_map={"FA고지": "#FF6B6B", "비교설명": "#4ECDC4", "완전판매": "#45B7D1", "총 미스캔": "#9B59B6"})
                        fig.update_layout(xaxis_tickangle=-45, height=420); st.plotly_chart(fig, use_container_width=True)
            with c2:
                max_v, yr = top["총미스캔"].max(), [0, top["총미스캔"].max()*1.2] if top["총미스캔"].max() >0 else [0,10]
                fig2 = go.Figure(); fig2.add_trace(go.Scatter(x=top["조직"], y=top["총미스캔"], mode="lines+markers", line=dict(shape="spline", color="#CC0000"), marker=dict(size=6)))
                fig2.update_layout(title=f"미처리 건수 추이 TOP {top_n}", xaxis_tickangle=-45, yaxis=dict(range=yr), height=420); st.plotly_chart(fig2, use_container_width=True)

    # ── TAB 2 : 미처리맵 ─────────────────────────
    with tab_map:
        st.subheader("🗺️ 미처리 분포 시각화")
        mc1, mc2 = st.columns([1, 2])
        with mc1: map_level = st.selectbox("집계 단위", ["부문", "총괄", "부서", "영업가족"], key="map_level")
        with mc2: map_type = st.radio("차트 유형", ["🥧 원그래프", "🔲 트리맵"], horizontal=True, key="map_type")
        map_agg = build_group_scan_stats(df_sel, map_level).rename(columns={"총미스캔": "미스캔"})
        map_agg = map_agg[map_agg["미스캔"] > 0].sort_values("미스캔", ascending=False)
        if map_agg.empty: st.info("미처리 건수가 있는 데이터가 없습니다.")
        else:
            if map_type == "🥧 원그래프":
                fig_pie = px.pie(map_agg, values="미스캔", names="조직", title=f"{map_level}별 미스캔 건수 비중", hole=0.4, color_discrete_sequence=px.colors.qualitative.Set3)
                fig_pie.update_traces(textposition='inside', textinfo='percent+label'); fig_pie.update_layout(height=500)
                st.plotly_chart(fig_pie, use_container_width=True)
            else:
                fig_tree = px.treemap(map_agg, path=["조직"], values="미스캔", color="미처리율", title=f"{map_level}별 미처리 분포", color_continuous_scale="RdYlGn_r")
                fig_tree.update_layout(height=500); st.plotly_chart(fig_tree, use_container_width=True)
            st.markdown(f"#### 📊 {map_level}별 상세 데이터")
            st.dataframe(
                map_agg[["조직", "대상건", "대상스캔건", "전체스캔", "미스캔", "스캔율", "미처리율"]].style.format({
                    "대상건": "{:,}", "대상스캔건": "{:,}", "전체스캔": "{:,}", "미스캔": "{:,}", "스캔율": "{:.1f}%", "미처리율": "{:.1f}%"
                }), use_container_width=True, hide_index=True)

    # ── TAB 3 : 계층 리포트 (KeyError 해결 부분) ─────────────────────────
    with tab_report:
        st.subheader("📊 전체 데이터 기반 계층별 미처리 현황")
        report_df = build_hierarchy_report(df, sel_months)
        if report_df.empty: st.info("데이터가 없습니다.")
        else:
            def style_row(row):
                if row["구분"]=="부문계": return ["background-color:#1F3864;color:white;font-weight:bold"]*len(row)
                elif row["구분"]=="총괄계": return ["background-color:#2E75B6;color:white;font-weight:bold"]*len(row)
                elif row["구분"]=="부서계": return ["background-color:#D9E1F2;font-weight:bold"]*len(row)
                elif row["구분"]=="소속계": return ["background-color:#E2EFDA;font-weight:bold"]*len(row)
                return [""]*len(row)
            
            # [수정] report_df에 실제 존재하는 컬럼만 동적으로 선택하여 KeyError 방지
            base_cols = ["구분", "부문", "총괄", "부서"]
            if "소속" in report_df.columns: base_cols.append("소속")
            disp_cols = base_cols + ["영업가족", "증번수", "대상건", "대상스캔건", "전체스캔", "전체미스캔", "스캔율"]
            
            # 존재하지 않는 컬럼이 있을 경우 필터링 (안정성 확보)
            disp_cols = [c for c in disp_cols if c in report_df.columns]
            disp_df = report_df[disp_cols].copy()
            
            st.markdown("""
            <style>
            table { white-space: nowrap !important; font-size: 0.85rem !important; }
            .stDataFrame { width: 100% !important; overflow-x: auto !important; }
            table td, table th { padding: 3px 5px !important; line-height: 1.2 !important; border-bottom: 1px solid #ddd; }
            </style>
            """, unsafe_allow_html=True)
            
            st.dataframe(
                disp_df.style.apply(style_row, axis=1).format({
                    "증번수": "{:,}", "대상건": "{:,}", "대상스캔건": "{:,}", "전체스캔": "{:,}", 
                    "전체미스캔": "{:,}", "스캔율": "{:.1f}%"
                }), use_container_width=True, hide_index=True, height=720)
            
            st.markdown("#### 📅 월별 계층 집계 데이터 (가로 피벗)")
            monthly_df = build_monthly_hierarchy(df, sel_months)
            if not monthly_df.empty:
                st.dataframe(monthly_df.style.format({c: "{:,}" for c in monthly_df.columns if "_총미스캔" in c}), use_container_width=True, hide_index=True, height=420)
                
            cr1, cr2, cr3 = st.columns(3)
            with cr1:
                if st.button("📥 계층 리포트 Excel", use_container_width=True):
                    with st.spinner("생성 중..."): buf = report_excel(df, sel_months)
                    st.download_button("⬇️ Excel", buf, f"계층리포트_{period_text.replace(' ','_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_rpt_xl")
            with cr2:
                if st.button("📥 계층 리포트 PDF", use_container_width=True):
                    with st.spinner("생성 중..."): buf2 = report_pdf(df, sel_months)
                    st.download_button("⬇️ PDF", buf2, f"계층리포트_{period_text.replace(' ','_')}.pdf", "application/pdf", key="dl_rpt_pdf")
            with cr3:
                if st.button("📄 전체 페이지 PDF", use_container_width=True, key="dl_fullpage_pdf"):
                    with st.spinner("전체 페이지 PDF 생성 중..."):
                        buf_full = report_fullpage_pdf(df, sel_months, agg_group, map_level, ["총 미스캔"], "그룹형", 15, "🔲 트리맵")
                    st.download_button("⬇️ 전체 페이지 PDF", buf_full, f"전체페이지리포트_{period_text.replace(' ','_')}.pdf", "application/pdf", key="dl_fullpage_pdf_btn")

    # ── TAB 4 : 관리대장 출력 ─────────────────────────
    with tab_ledger:
        st.subheader("📋 관리대장 선정 및 출력")
        cf1, cf2, cf3 = st.columns(3)
        with cf1: sel_bm = st.selectbox("부문", ["전체"]+sorted(df_sel["부문"].dropna().unique().tolist()), key="lg_bm")
        df_l1 = df_sel if sel_bm=="전체" else df_sel[df_sel["부문"]==sel_bm]
        with cf2: sel_tg = st.selectbox("총괄", ["전체"]+sorted(df_l1["총괄"].dropna().unique().tolist()), key="lg_tg")
        df_l2 = df_l1 if sel_tg=="전체" else df_l1[df_l1["총괄"]==sel_tg]
        with cf3: sel_ds = st.selectbox("부서", ["전체"]+sorted(df_l2["부서"].dropna().unique().tolist()), key="lg_ds")
        df_l3 = df_l2 if sel_ds=="전체" else df_l2[df_l2["부서"]==sel_ds]
        targets = get_ledger_targets(df_l3, sel_months)
        if not targets: st.success("✅ 미스캔 발생 대상이 없습니다.")
        else:
            prev = [{"부문":r["부문"], "총괄":r["총괄"], "부서":dept, "영업가족":r["영업가족"], "FA":int(r["FA"]), "비교":int(r["비교"]), "완판":int(r["완판"]), "총미스캔":int(r["총미스캔"])} for dept, grp in targets.items() for _, r in grp.iterrows()]
            prev_df = pd.DataFrame(prev)
            st.markdown(f"#### 📌 선정 대상 — 총 **{len(prev_df)}** 개 영업가족")
            st.dataframe(prev_df.style.format({"FA": "{:,}", "비교": "{:,}", "완판": "{:,}", "총미스캔": "{:,}"}), use_container_width=True, hide_index=True)
            all_depts = sorted(targets.keys())
            sel_depts = st.multiselect("출력 부서 (미선택 시 전체)", all_depts, default=all_depts, key="lg_sel_dept")
            if not sel_depts: st.warning("⚠️ 출력할 부서를 1개 이상 선택하세요.")
            else:
                out_targets = {d: targets[d] for d in sel_depts if d in targets}
                st.info(f"📄 출력 대상: **{len(sel_depts)}개 부서** · **{len([r for r in prev if r['부서'] in sel_depts])}개 영업가족**")
                cd1, cd2 = st.columns(2)
                with cd1:
                    if st.button("📥 관리대장 PDF", use_container_width=True, key="gen_pdf"):
                        with st.spinner("생성 중..."): pb = ledger_pdf(out_targets, period_text, df_l3)
                        st.download_button("⬇️ PDF", pb, f"관리대장_{period_text.replace(' ','_')}.pdf", "application/pdf", key="dl_ldg_pdf")
                with cd2:
                    if st.button("📥 관리대장 Excel", use_container_width=True, key="gen_xl"):
                        with st.spinner("생성 중..."): xb = ledger_excel(out_targets, period_text, df_l3)
                        st.download_button("⬇️ Excel", xb, f"관리대장_{period_text.replace(' ','_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_ldg_xl")

# ==========================================
# 13. main
# ==========================================
def main():
    if not st.session_state.get("logged_in"):
        login_page()
    else:
        with st.sidebar:
            st.success("👋 접속 완료")
            if st.button("🚪 로그아웃", use_container_width=True):
                st.session_state.logged_in = False
                st.rerun()
            st.divider()
            st.caption("v5.3 | KeyError해결·정렬최적화·가로피벗 | © 2026")
        dashboard_page()

if __name__ == "__main__":
    main()
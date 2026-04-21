import streamlit as st
import pandas as pd
import os
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
TableStyle, Image as RLImage, PageBreak, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import io

# ==========================================
# 0. 페이지 설정 (최상단 필수)
# ==========================================
st.set_page_config(page_title="보험 서류 스캔 관리 대시보드", layout="wide", page_icon="📊")

# ==========================================
# 1. 전역 설정 & 인증
# ==========================================
EXCEL_FILE = "insurance_data.xlsx"
APP_PASSWORD = os.environ.get("APP_PASSWORD", "incar961")

GUIDANCE_TEXT = (
 "【책임판매 필수서류 안내】\n "
 "개인정보동의서, 비교설명확인서, 고지의무확인서, 완전판매확인서(대상계약 限)는  "
 "금융소비자보호법 및 보험업 감독규정에 따라 신계약 체결 전 구비가 요구되는 필수 서류입니다.  "
 "상기 서류는 소비자 보호 및 설명 의무 이행 여부를 확인하기 위한 내부 통제 관리 대상 서류로서, 실적 확정 입력 마감 시점까지 제출 완료를 원칙으로 하며 미비 시 내부 통제 리스크 관리 대상 계약으로 분류됩니다. "
)
PRECAUTION_TEXT_COVER = (
 "【미처리 시 유의사항】\n "
 "실적 확정 입력 마감 시점까지 필수 서류가 제출되지 않은 계약과 조직에 대하여는 모집질서 및 분쟁  리스크 관리 대상으로 분류되어 관리됩니다.\n "
 "신계약 점검이 강화되며 내부 통제 기준 충족 시까지, 신규 지원금 승인이 보류되거나 참여 요건이 제한될 수 있습니다. "
)
PRECAUTION_TEXT_CONFIRM =  "영업가족별 미처리 현황 및 유의사항에 대하여 인지하였으며, "
PRECAUTION_TEXT_SHEET =  "본인은 신계약 필수 서류의 사전 구비 의무 및 미제출 시 내부 통제 관리 기준이 적용될 수 있음을 확인합니다. "
SIGNATURE_CONFIRMATION_TEXT =  "신계약 필수 서류의 사전 구비 의무 및 미제출 시 내부 통제 관리 기준이 적용 사항을 영업가족에게 안내하였음을 확인합니다. "

REQUIRED_DOCS_TABLE = [
[ "No. ",  "서류명 ",  "법적 관리 근거 및 관련 내부 통제 기준 ",  "목적 및 주요 내용 "],
[ "1 ",  "개인정보동의서 ",
 "개인정보보호법 15조 및\n대리점 표준 내부통제기준 27조 ",
 "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에\n따른 개인정보 처리로 신계약시 필수 징구 "],
[ "2 ",  "비교설명확인서 ",
 "보험업감독규정\n별표 5-6 ",
 "유사 상품 3개 이상 비교·설명 이행\n사실 고객 확인 서명 "],
[ "3 ",  "고지의무확인서 ",
 "금융소비자보호법 26조와\n동법시행령 24조 ",
 "판매자 중요사항 고지의무 이행 확인,\n권한·책임·보상 관련 핵심 사항 고지,\n소비자 오인 예방 "],
[ "4 ",  "완전판매확인서\n(대상: 종신, CI, CEO경기, 고액) ",
 "금융소비자보호법 제17·19조 설명 적합성 적정성 관련 조항\n영업지원기준안 ",
 "약관,청약서 부본 제공, 중요 상품 이해 및\n자발적 가입 확인, 설명 의무 이행 증빙력 확보 "]
]

# ==========================================
# 2. 데이터 로딩 (GitHub 엑셀 기반) - ✅ 집계 로직 완전 교체
# ==========================================
@st.cache_data(ttl=300)
def load_data():
    """GitHub의 엑셀 파일을 읽어서 DataFrame 반환 - 증번 기준 스캔율 계산용 전처리"""
    if not os.path.exists(EXCEL_FILE):
        st.error(f"⚠️ '{EXCEL_FILE}' 파일이 없습니다. GitHub에 엑셀 파일을 업로드해주세요.")
        return pd.DataFrame()
    
    try:
        df = pd.read_excel(EXCEL_FILE)
        
        if df.empty:
            return pd.DataFrame()
        
        # 날짜 및 필드 전처리 
        df["보험시작일_dt"] = pd.to_datetime(df["보험시작일"], errors="coerce")
        df["월_피리어드"] = df["보험시작일_dt"].dt.to_period("M").astype(str)
        
        # ✅ 스캔 여부 판정 함수 (스캔, M스캔, 보험사스캔 = 스캔 처리)
        def is_scanned(val):
            if pd.isna(val): return False
            val_str = str(val).strip()
            return val_str in ["스캔", "M스캔", "보험사스캔"]
        
        # ✅ 완판 대상 여부 판정 (스캔, M스캔, 미스캔만 대상 / 해당없음 제외)
        def is_cs_target(val):
            if pd.isna(val): return False
            val_str = str(val).strip()
            return val_str not in ["해당없음", ""] # 비어있거나 해당없음이 아니면 대상
        
        # ✅ 각 서류별 플래그 생성
        # 개인정보, FA고지, 비교설명: 모든 증번이 대상이므로 스캔 여부만 판별
        df["개인정보_스캔"] = df["개인정보"].apply(is_scanned).astype(int)
        df["FA고지_스캔"] = df["FA고지"].apply(is_scanned).astype(int)
        df["비교설명_스캔"] = df["비교설명"].apply(is_scanned).astype(int)
        
        # 완전판매: 대상건만 집계
        df["완전판매_대상"] = df["완전판매"].apply(is_cs_target).astype(int)
        df["완전판매_스캔"] = df["완전판매"].apply(is_scanned).astype(int)
        
        # 미스캔 계산 (대상 - 스캔)
        df["개인정보_미스캔"] = 1 - df["개인정보_스캔"]
        df["FA고지_미스캔"] = 1 - df["FA고지_스캔"]
        df["비교설명_미스캔"] = 1 - df["비교설명_스캔"]
        df["완전판매_미스캔"] = df["완전판매_대상"] - df["완전판매_스캔"]
        
        return df
    except Exception as e:
        st.error(f"❌ 엑셀 파일 읽기 오류: {e}")
        return pd.DataFrame()

def get_file_update_time():
    """엑셀 파일의 마지막 수정 시간 반환"""
    if os.path.exists(EXCEL_FILE):
        timestamp = os.path.getmtime(EXCEL_FILE)
        return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
    return "알 수 없음"

# ==========================================
# 3. 집계 헬퍼
# ==========================================
def calculate_scan_stats(df_group):
    """
    증번 기준 스캔율 계산
    - 개인정보, FA고지, 비교설명: 증번당 필수 (각각 증번수만큼 대상)
    - 완전판매: 대상건만 집계
    """
    cnt = len(df_group)  # 증번 수
    
    # 필수 서류 (증번당 1개씩 필요)
    개인정보_스캔 = int(df_group["개인정보_스캔"].sum())
    FA고지_스캔 = int(df_group["FA고지_스캔"].sum())
    비교설명_스캔 = int(df_group["비교설명_스캔"].sum())
    
    # 완전판매
    완전판매_대상 = int(df_group["완전판매_대상"].sum())
    완전판매_스캔 = int(df_group["완전판매_스캔"].sum())
    
    # 전체 대상 및 스캔
    전체_대상 = (cnt * 3) + 완전판매_대상
    전체_스캔 = 개인정보_스캔 + FA고지_스캔 + 비교설명_스캔 + 완전판매_스캔
    전체_미스캔 = 전체_대상 - 전체_스캔
    
    # 스캔율
    스캔율 = round((전체_스캔 / 전체_대상 * 100), 1) if 전체_대상 > 0 else 0.0
    
    return {
        "증번수": cnt,
        "전체대상": 전체_대상,
        "전체스캔": 전체_스캔,
        "전체미스캔": 전체_미스캔,
        "스캔율": 스캔율
    }

# ==========================================
# 4. 전체 계층 리포트 (피벗 형태 - 스캔율 기반)
# ==========================================
@st.cache_data(ttl=300)
def build_hierarchy_report(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    
    rows = []
    for bm, df_bm in src.groupby("부문"):
        stats = calculate_scan_stats(df_bm)
        rows.append({
            "구분": "부문계", "부문":bm, "총괄":" ", "부서":" ", "영업가족":" ",
            **stats
        })
        for tg, df_tg in df_bm.groupby("총괄"):
            stats = calculate_scan_stats(df_tg)
            rows.append({
                "구분": "총괄계", "부문":bm, "총괄":tg, "부서":" ", "영업가족":" ",
                **stats
            })
            for ds, df_ds in df_tg.groupby("부서"):
                stats = calculate_scan_stats(df_ds)
                rows.append({
                    "구분": "부서계", "부문":bm, "총괄":tg, "부서":ds, "영업가족":" ",
                    **stats
                })
                for fg, df_fg in df_ds.groupby("영업가족"):
                    stats = calculate_scan_stats(df_fg)
                    rows.append({
                        "구분": "영업가족", "부문":bm, "총괄":tg, "부서":ds, "영업가족":fg,
                        **stats
                    })
    return pd.DataFrame(rows)

@st.cache_data(ttl=300)
def build_monthly_hierarchy(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    
    rows = []
    for mon, dm in src.groupby("월_피리어드"):
        for bm, db in dm.groupby("부문"):
            stats = calculate_scan_stats(db)
            rows.append({"월":mon, "구분":"부문계", "부문":bm, "총괄":" ", "부서":" ", **stats})
        for tg, dt in db.groupby("총괄"):
            stats = calculate_scan_stats(dt)
            rows.append({"월":mon, "구분":"총괄계", "부문":bm, "총괄":tg, "부서":" ", **stats})
            for ds, dd in dt.groupby("부서"):
                stats = calculate_scan_stats(dd)
                rows.append({"월":mon, "구분":"부서계", "부문":bm, "총괄":tg, "부서":ds, **stats})
    return pd.DataFrame(rows)

@st.cache_data(ttl=300)
def build_monthly_hierarchy_pivot(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    
    rows = []
    for mon, dm in src.groupby("월_피리어드"):
        for bm, db in dm.groupby("부문"):
            stats = calculate_scan_stats(db)
            rows.append({"월":mon, "구분":"부문계", "부문":bm, "총괄":" ", "부서":" ", **stats})
        for tg, dt in db.groupby("총괄"):
            stats = calculate_scan_stats(dt)
            rows.append({"월":mon, "구분":"총괄계", "부문":bm, "총괄":tg, "부서":" ", **stats})
            for ds, dd in dt.groupby("부서"):
                stats = calculate_scan_stats(dd)
                rows.append({"월":mon, "구분":"부서계", "부문":bm, "총괄":tg, "부서":ds, **stats})
    
    pivot_src = pd.DataFrame(rows)
    if pivot_src.empty: return pivot_src
    
    metrics = ["증번수", "전체대상", "전체스캔", "전체미스캔", "스캔율"]
    pivot_frames = []
    month_order = sorted(src["월_피리어드"].dropna().unique())
    for metric in metrics:
        temp = pivot_src.pivot_table(index=["구분", "부문", "총괄", "부서"], columns="월", values=metric, aggfunc="first")
        temp.columns = [f"{month}_{metric}" for month in temp.columns]
        pivot_frames.append(temp)
    
    pivot = pd.concat(pivot_frames, axis=1).reset_index()
    ordered_columns = ["구분", "부문", "총괄", "부서"]
    for month in month_order:
        for metric in metrics:
            ordered_columns.append(f"{month}_{metric}")
    pivot = pivot[[c for c in ordered_columns if c in pivot.columns]]
    return pivot.fillna(0)

# ==========================================
# 5. 관리대장 선정 대상
# ==========================================
@st.cache_data(ttl=300)
def get_ledger_targets(df, months):
    src = df[df["월_피리어드"].isin(months)].copy()
    if src.empty: return {}
    
    # 증번 기준 집계 적용
    agg = src.groupby(["부문", "총괄", "부서", "영업가족"]).agg(
        증번수=("증권번호", "count"),
        개인정보_스캔=("개인정보_스캔", "sum"),
        FA고지_스캔=("FA고지_스캔", "sum"),
        비교설명_스캔=("비교설명_스캔", "sum"),
        완전판매_대상=("완전판매_대상", "sum"),
        완전판매_스캔=("완전판매_스캔", "sum")
    ).reset_index()
    
    agg["전체대상"] = (agg["증번수"] * 3) + agg["완전판매_대상"]
    agg["전체스캔"] = agg["개인정보_스캔"] + agg["FA고지_스캔"] + agg["비교설명_스캔"] + agg["완전판매_스캔"]
    agg["총미스캔"] = agg["전체대상"] - agg["전체스캔"]
    
    agg = agg[agg["총미스캔"] > 0]
    return {dept: grp for dept, grp in agg.groupby("부서")}

# ==========================================
# 6. 한글 폰트 및 스타일
# ==========================================
@st.cache_resource
def register_korean_font():
    font_candidates = [
        ("NotoSansKR", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        ("NotoSansKR", "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
        ("Malgun", r"C:\Windows\Fonts\malgun.ttf"),
        ("DejaVu", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for name, path in font_candidates:
        try:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont(name, path))
                return name
        except Exception:
            continue
    return "Helvetica"

HDR_CLR, ALT_CLR, SUB_CLR = "#4472C4", "#EEF3FB", "#D9E1F2"

def _pdf_styles(fn):
    S = getSampleStyleSheet()
    def ps(n, **kw): return ParagraphStyle(n, parent=S["Normal"], fontName=fn, **kw)
    return {
        "title":   ps("T",  fontSize=15, bold=True, alignment=1, spaceAfter=4),
        "sub":     ps("S",  fontSize=10, spaceAfter=3),
        "body":    ps("B",  fontSize=8,  spaceAfter=2),
        "notice":  ps("N",  fontSize=7.5, spaceAfter=3, textColor=colors.HexColor("#CC0000"), alignment=1),
        "date":    ps("D",  fontSize=8,  alignment=2, spaceAfter=4),
        "section": ps("SC", fontSize=9,  bold=True,  spaceAfter=2),
    }

def _tbl(data, cw, fn, header_rows=1, sub_rows=None, align="CENTER"):
    if not data or len(data) < 1: return Spacer(1,0)
    cw_scaled = [w * 1.4 for w in cw]
    align_map = {"LEFT":0, "CENTER":1, "RIGHT":2}
    align_value = align_map.get(align.upper(), 1)
    S = getSampleStyleSheet()
    cell_style = ParagraphStyle("tbl_cell", parent=S["Normal"], fontName=fn, fontSize=8, leading=10, alignment=align_value, wordWrap="CJK")
    wrapped_data = [[Paragraph(str(cell), cell_style) if not isinstance(cell, Paragraph) else cell for cell in row] for row in data]
    t = Table(wrapped_data, colWidths=cw_scaled, repeatRows=header_rows)
    cmds = [
        ("FONTNAME", (0,0),(-1,-1), fn), ("FONTSIZE", (0,0),(-1,-1), 8),
        ("ALIGN", (0,0),(-1,-1), align.upper()), ("VALIGN", (0,0),(-1,-1), "MIDDLE"),
        ("WORDWRAP", (0,0),(-1,-1), "CJK"), ("GRID", (0,0),(-1,-1), 0.4, colors.grey),
        ("LEFTPADDING", (0,0),(-1,-1), 4), ("RIGHTPADDING", (0,0),(-1,-1), 4),
        ("TOPPADDING", (0,0),(-1,-1), 4), ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ("BACKGROUND", (0,0),(-1,header_rows-1), colors.HexColor("#DCE6F1")),
        ("TEXTCOLOR", (0,0),(-1,header_rows-1), colors.HexColor("#1F3864")),
    ]
    for i in range(header_rows, len(data)):
        if (i-header_rows)%2==1: cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F3F6FA")))
        if sub_rows and i in sub_rows: cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#E9EEF8")))
    t.setStyle(TableStyle(cmds)); return t

def _fig_to_image(fig, max_width=1000, height=360):
    try:
        img_buf = io.BytesIO()
        fig.write_image(img_buf, format="png", width=max_width, height=height, scale=2)
        img_buf.seek(0)
        img = RLImage(img_buf)
        desired_width = min(max_width, 820)
        img.drawWidth = desired_width
        img.drawHeight = height * (desired_width / max_width)
        return img, None
    except Exception as e:
        return None, str(e)

def append_pdf_figure(E, fig, st, max_width=1000, height=360):
    img, err = _fig_to_image(fig, max_width=max_width, height=height)
    if img is not None:
        E.append(img)
        E.append(Spacer(1,10))
        return
    E.append(Paragraph(f"차트 이미지를 생성하지 못했습니다: {err}", st["notice"]))
    E.append(Spacer(1,6))

def _sig_table(labels, fn, cw=120):
    t = Table([labels,["____________________"]*len(labels),["(인)"]*len(labels)], colWidths=[cw*1.4]*len(labels))
    t.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),("FONTNAME",(0,0),(-1,-1),fn),
                           ("FONTSIZE",(0,0),(-1,-1),8.5),("TOPPADDING",(0,0),(-1,-1),5),
                           ("BOTTOMPADDING",(0,0),(-1,-1),5),("BOX",(0,0),(-1,-1),0.5,colors.grey),
                           ("INNERGRID",(0,0),(-1,-1),0.3,colors.lightgrey)]))
    return t

# ==========================================
# 7. 전체 계층 리포트 Excel (스캔율 기반)
# ==========================================
def report_excel(df, months):
    wb = Workbook(); ws = wb.active; ws.title="계층별_미처리현황"
    tfn, hf, bf = "맑은 고딕", Font(name="맑은 고딕",size=9,bold=True,color="FFFFFF"), Font(name="맑은 고딕",size=9)
    bdr = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
    fills = {"부문계":PatternFill("solid",fgColor="1F3864"), "총괄계":PatternFill("solid",fgColor="2E75B6"),
             "부서계":PatternFill("solid",fgColor="D9E1F2"), "영업가족_alt":PatternFill("solid",fgColor="EEF3FB")}
    fonts_wc = {"부문계":Font(name=tfn,size=9,bold=True,color="FFFFFF"), "총괄계":Font(name=tfn,size=9,bold=True,color="FFFFFF"),
                "부서계":Font(name=tfn,size=9,bold=True)}
    h_fill = PatternFill("solid",fgColor="4472C4")
    alt_fill = PatternFill("solid",fgColor="EEF3FB")
    today, period_str = datetime.now().strftime("%Y년 %m월 %d일"), ", ".join(months) if months else "전체"
    
    ws.merge_cells("A1:K1"); ws["A1"] = f"서류 미처리 현황 계층별 집계  ·  기간: {period_str}  ·  발급: {today}"
    ws["A1"].font = Font(name=tfn,size=12,bold=True); ws["A1"].alignment = Alignment(horizontal="center"); ws.row_dimensions[1].height = 22
    
    # ✅ 헤더 변경: 증번/스캔율 기반
    headers = ["구분", "부문", "총괄", "부서", "영업가족", "증번수", "전체대상", "전체스캔", "전체미스캔", "스캔율"]
    cws = [14,20,20,20,24,10,12,12,14,10]
    for ci,(h,w) in enumerate(zip(headers,cws),1):
        c=ws.cell(2,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center",vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width=w

    report = build_hierarchy_report(df, months)
    if report.empty: return io.BytesIO()
    ri = 3
    for _, row in report.iterrows(): 
        gbn = row["구분"]; rate_str = f"{row['스캔율']}%"
        vals = [gbn, row["부문"], row["총괄"], row["부서"], row["영업가족"], row["증번수"], 
                row["전체대상"], row["전체스캔"], row["전체미스캔"], rate_str]
        fill = fills.get(gbn, fills["영업가족_alt"] if ri%2==0 else None)
        fnt  = fonts_wc.get(gbn, bf)
        for ci,v in enumerate(vals,1):
            c=ws.cell(ri,ci,v); c.font=fnt; c.border=bdr; c.alignment=Alignment(horizontal="center",vertical="center")
            if fill: c.fill=fill
        ri += 1

    monthly = build_monthly_hierarchy(df, months)
    if not monthly.empty:
        ws2 = wb.create_sheet("월별_계층집계")
        ws2.merge_cells("A1:K1"); ws2["A1"] = f"월별 계층 미처리 집계  ·  기간: {period_str}  ·  발급일: {today}"
        ws2["A1"].font = Font(name=tfn,size=12,bold=True); ws2["A1"].alignment=Alignment(horizontal="center"); ws2.row_dimensions[1].height = 22
        mhdr = ["월", "구분", "부문", "총괄", "부서", "증번수", "전체대상", "전체스캔", "전체미스캔", "스캔율"]
        mcws = [14,10,14,14,16,10,12,12,14,10]
        for ci,(h,w) in enumerate(zip(mhdr,mcws),1):
            c=ws2.cell(2,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center"); ws2.column_dimensions[get_column_letter(ci)].width=w
        for ri2,(_, r) in enumerate(monthly.iterrows(),3):
            vals2=[r["월"],r["구분"],r["부문"],r["총괄"],r["부서"],r["증번수"],r["전체대상"],r["전체스캔"],r["전체미스캔"], f"{r['스캔율']}%"]
            fill2=fills.get(r["구분"], fills["영업가족_alt"] if ri2%2==0 else None); fnt2=fonts_wc.get(r["구분"],bf)
            for ci,v in enumerate(vals2,1):
                c=ws2.cell(ri2,ci,v); c.font=fnt2; c.border=bdr; c.alignment=Alignment(horizontal="center")
                if fill2: c.fill=fill2
    
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ==========================================
# 8. 전체 계층 리포트 PDF (스캔율 기반)
# ==========================================
def report_pdf(df, months):
    fn, st_, buf = register_korean_font(), _pdf_styles(register_korean_font()), io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    today, period_str = datetime.now().strftime("%Y년 %m월 %d일"), ", ".join(months) if months else "전체"
    E = [Paragraph("서류 미처리 현황 계층별 집계", st_["title"]), Paragraph(f"기간: {period_str}  |  발급일자: {today}", st_["date"]), HRFlowable(width="100%",thickness=1,color=colors.HexColor(HDR_CLR)), Spacer(1,6)]
    
    report = build_hierarchy_report(df, months)
    if not report.empty:
        E.append(Paragraph("▶ 부문 / 총괄 / 부서 / 영업가족 계층 집계", st_["section"]))
        hdr=[["구분", "부문", "총괄", "부서", "영업가족", "증번수", "전체대상", "전체스캔", "전체미스캔", "스캔율"]]
        drows, sub_idx = [], []
        for i,(_,r) in enumerate(report.iterrows()):
            drows.append([r["구분"],r["부문"],r["총괄"],r["부서"],r["영업가족"],r["증번수"],r["전체대상"],r["전체스캔"],r["전체미스캔"],f"{r['스캔율']}%"])
            if r["구분"] in ("부문계", "총괄계", "부서계"): sub_idx.append(i+1)
        cw=[15,25,25,25,35,15,20,20,22,15]
        E.append(_tbl(hdr+drows,cw,fn,sub_rows=sub_idx)); E.append(Spacer(1,8))

    monthly = build_monthly_hierarchy(df, months)
    if not monthly.empty:
        E.append(PageBreak()); E.append(Paragraph("▶ 월별 계층별 미처리 집계", st_["section"]))
        mrows=[[r["월"],r["구분"],r["부문"],r["총괄"],r["부서"],r["증번수"],r["전체대상"],r["전체스캔"],r["전체미스캔"],f"{r['스캔율']}%"] for _,r in monthly.iterrows()]
        msub=[i+1 for i,(_,r) in enumerate(monthly.iterrows()) if r["구분"] in ("부문계", "총괄계", "부서계")]
        cw2=[15,15,25,25,30,15,20,20,22,15]
        E.append(_tbl([["월", "구분", "부문", "총괄", "부서", "증번수", "전체대상", "전체스캔", "전체미스캔", "스캔율"]]+mrows,cw2,fn,sub_rows=msub))

    doc.build(E); buf.seek(0); return buf

# ==========================================
# 9. 관리대장 PDF (스캔율 기반)
# ==========================================
def ledger_pdf(families_by_dept, period_text, df_src):
    fn, st_, buf = register_korean_font(), _pdf_styles(register_korean_font()), io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=12*mm,leftMargin=12*mm, topMargin=15*mm,bottomMargin=15*mm)
    today = datetime.now().strftime("%Y년 %m월 %d일"); E = []
    center_date_style = ParagraphStyle("CenterDate", parent=st_["date"], alignment=1)
    title_left = ParagraphStyle("TitleLeft", parent=st_["title"], alignment=0)
    indent_style = ParagraphStyle("IndentSub", parent=st_["sub"], leftIndent=8, alignment=0, spaceAfter=2)
    date_indent = ParagraphStyle("DateIndent", parent=st_["date"], leftIndent=8, alignment=0)
    section_left = ParagraphStyle("SectionLeft", parent=st_["section"], alignment=0)
    notice_left = ParagraphStyle("NoticeLeft", parent=st_["notice"], leftIndent=8, alignment=0)
    
    for dept_name, grp_df in families_by_dept.items():
        sec, tg = grp_df.iloc[0]["부문"], grp_df.iloc[0]["총괄"]
        E += [Paragraph("신계약 필수서류 미처리 확인서", title_left), HRFlowable(width="100%",thickness=1.5,color=colors.HexColor(HDR_CLR)), Spacer(1,4),
        Paragraph(f"부서: {sec}   > {tg}   >   {dept_name}", indent_style), Paragraph(f"적용기간: {period_text}", date_indent), Spacer(1,6)]
        dept_src = df_src[df_src["부서"]==dept_name]
        E += [Paragraph("【필수 서류 상세 안내】", st_["section"]),
        _tbl(REQUIRED_DOCS_TABLE, [12, 60, 90, 198], fn, header_rows=1, align="LEFT"), Spacer(1,8),
        Paragraph(GUIDANCE_TEXT, notice_left), Spacer(1,8)]
        
        if not dept_src.empty:
            E.append(Paragraph("▶ 영업가족별 · 월별 · 양식별 미처리 현황", section_left))
            fam_mon = dept_src.groupby(["영업가족", "월_피리어드"]).agg(
                개인정보_미스캔=("개인정보_미스캔", "sum"), FA=("FA고지_미스캔", "sum"),비교=("비교설명_미스캔", "sum"),완판=("완전판매_미스캔", "sum")
            ).reset_index()
            fam_mon["계"] = fam_mon[["개인정보_미스캔", "FA", "비교", "완판"]].sum(axis=1)
            fam_mon = fam_mon[fam_mon["계"] > 0]
            if not fam_mon.empty:
                td=[["영업가족", "월", "개인정보", "FA고지", "비교설명", "완전판매", "계"]]
                for _, r in fam_mon.iterrows(): td.append([r["영업가족"],r["월_피리어드"],int(r.개인정보_미스캔),int(r.FA),int(r.비교),int(r.완판),int(r["계"])])
                E.append(_tbl(td,[130,50,45,45,45,45],fn, align="LEFT"))
            E.append(Spacer(1,4))
            
        E += [Paragraph(PRECAUTION_TEXT_COVER, notice_left), Spacer(1,4),
        Paragraph(PRECAUTION_TEXT_CONFIRM, notice_left), Spacer(1,4),
        Paragraph(SIGNATURE_CONFIRMATION_TEXT, notice_left), Spacer(1,8),
        Paragraph("작성일: _______________", center_date_style), Spacer(1,4),
        _sig_table(["부문장 확인", "총괄 확인", "부서장 확인"],fn,120), PageBreak()]
        
        for _, fam in grp_df.drop_duplicates("영업가족").iterrows():
            fam_name = fam["영업가족"]
            E += [Paragraph("신계약 필수서류 미처리 확인서", title_left), HRFlowable(width="100%",thickness=1.5,color=colors.HexColor(HDR_CLR)), Spacer(1,4),
            Paragraph(f"소속: {sec}   > {tg}   > {dept_name}   >   {fam_name}", indent_style), Paragraph(f"적용기간: {period_text}", date_indent), Spacer(1,6)]
            
            fam_src = df_src[(df_src["영업가족"]==fam_name) & df_src["소속"].notna()]
            sosok = fam_src.groupby(["소속", "월_피리어드"]).agg(
                개인정보_미스캔=("개인정보_미스캔", "sum"), FA=("FA고지_미스캔", "sum"),비교=("비교설명_미스캔", "sum"),완판=("완전판매_미스캔", "sum")
            ).reset_index()
            sosok["계"] = sosok[["개인정보_미스캔", "FA", "비교", "완판"]].sum(axis=1)
            sosok = sosok[sosok["계"] > 0]
            
            E.append(Paragraph("▶ 소속별 · 월별 · 양식별 미처리 건수", section_left))
            if not sosok.empty:
                td2=[["소속", "월", "개인정보", "FA고지", "비교설명", "완전판매", "계"]]
                for _,r in sosok.iterrows(): td2.append([r["소속"],r["월_피리어드"],int(r.개인정보_미스캔),int(r.FA),int(r.비교),int(r.완판),int(r["계"])])
                E.append(_tbl(td2,[130,50,45,45,45,45],fn, align="LEFT"))
            else:
                E.append(Paragraph("(해당 데이터 없음)", st_["body"]))
            E.append(Spacer(1,4))
            
            sum_d=[["개인정보", "FA고지", "비교설명", "완전판매", "총계"],
                   [str(int(fam["개인정보_미스캔"])),str(int(fam["FA"])),str(int(fam["비교"])),str(int(fam["완판"])),str(int(fam["총미스캔"]))]]
            E += [Paragraph("▶ 양식별 미처리 요약", section_left), _tbl(sum_d,[90,90,90,90],fn, align="LEFT"), Spacer(1,8),
            Paragraph("【필수 서류 상세 안내】", st_["section"]),
            _tbl(REQUIRED_DOCS_TABLE, [12, 60, 90, 198], fn, header_rows=1, align="LEFT"), Spacer(1,8),
            Paragraph(GUIDANCE_TEXT, notice_left), Spacer(1,8),
            Paragraph(PRECAUTION_TEXT_COVER, notice_left), Spacer(1,4),
            Paragraph(PRECAUTION_TEXT_SHEET, notice_left), Spacer(1,8),
            Paragraph("작성일: _______________", center_date_style)]
            
            sig2=Table([[f"영업가족대표 서명: ____________________ (인)"]], colWidths=[120*1.4*3])
            sig2.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"LEFT"),("FONTNAME",(0,0),(-1,-1),fn),("FONTSIZE",(0,0),(-1,-1),9.5),
            ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),("BOX",(0,0),(-1,-1),0.5,colors.grey)]))
            E += [sig2, PageBreak()]
    doc.build(E); buf.seek(0); return buf

# ==========================================
# 10. 관리대장 Excel (스캔율 기반)
# ==========================================
def ledger_excel(families_by_dept, period_text, df_src):
    wb = Workbook(); ws0 = wb.active; ws0.title="목차"
    tfn = "맑은 고딕"
    hf = Font(name=tfn,size=9,bold=True,color="FFFFFF")
    bf = Font(name=tfn,size=9)
    nf = Font(name=tfn,size=8,italic=True,color="CC0000")
    sig_f = Font(name=tfn,size=9,bold=True)
    bdr, h_fill, alt_fill = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin")), PatternFill("solid",fgColor="4472C4"), PatternFill("solid",fgColor="EEF3FB")
    today = datetime.now().strftime("%Y년 %m월 %d일")
    
    ws0.merge_cells("A1:F1"); ws0["A1"]=f"관리대장 목차  ·  {period_text}  ·  발급: {today}"
    ws0["A1"].font=Font(name=tfn,size=13,bold=True); ws0["A1"].alignment=Alignment(horizontal="center")
    for ci,h in enumerate(["부서", "영업가족", "개인정보", "FA고지", "비교설명", "완전판매", "총미스캔"],1):
        c=ws0.cell(3,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center")
    ir=4
    for dept,grp in families_by_dept.items():
        for _,fam in grp.drop_duplicates("영업가족").iterrows():
            for ci,v in enumerate([dept,fam["영업가족"],int(fam["개인정보_미스캔"]),int(fam["FA"]),int(fam["비교"]),int(fam["완판"]),int(fam["총미스캔"])],1):
                c=ws0.cell(ir,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                if ir%2==0: c.fill=alt_fill
            ir+=1
    for ci,w in enumerate([22,25,13,13,13,13,12],1): ws0.column_dimensions[get_column_letter(ci)].width=w
    
    for dept_name,grp_df in families_by_dept.items():
        sec, tg = grp_df.iloc[0]["부문"], grp_df.iloc[0]["총괄"]
        sname=f"표지_{dept_name[:10]}".replace("/", " "); ws_c=wb.create_sheet(title=sname)
        ws_c.merge_cells("A1:G1"); ws_c["A1"]=f"[{dept_name}]  신계약 필수서류 미처리 확인서"; ws_c["A1"].font=Font(name=tfn,size=14,bold=True)
        ws_c["A2"]=f"{sec}   > {tg}   > {dept_name}"; ws_c["A2"].font=Font(name=tfn,size=10)
        ws_c["A3"]=f"    적용기간: {period_text} "; ws_c["A3"].font=bf; ws_c["A3"].alignment=Alignment(horizontal="left")
        
        r=5
        ws_c.cell(r,1,"【필수 서류 상세 안내】").font=Font(name=tfn,size=10,bold=True); r+=1
        for ci, header in enumerate(["No. ", "서류명 ", "법적 근거 ", "목적 및 주요 내용 "], 1):
            c = ws_c.cell(r, ci, header)
            c.font = hf; c.fill = h_fill; c.border = bdr; c.alignment = Alignment(horizontal="center", vertical="center")
        ws_c.column_dimensions[get_column_letter(1)].width = 6; ws_c.column_dimensions[get_column_letter(2)].width = 20
        ws_c.column_dimensions[get_column_letter(3)].width = 25; ws_c.column_dimensions[get_column_letter(4)].width = 45
        r += 1
        
        docs_data = [
            ["1 ", "개인정보동의서 ", "개인정보보호법 15조 및\n대리점 표준 내부통제기준 27조 ", "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에 따른 개인정보 처리로 신계약시 필수 징구 "],
            ["2 ", "비교설명확인서 ", "보험업감독규정\n별표 5-6 ", "유사 상품 3개 이상 비교·설명 이행 사실 고객 확인 서명 "],
            ["3 ", "고지의무확인서 ", "금융소비자보호법 26조와\n동법시행령 24조 ", "판매자 권한·책임·보상 관련 핵심 사항 고지, 소비자 소인 예방 "],
            ["4 ", "완전판매확인서\n(대상: 종신, CI, CEO경기, 고액) ", "금융소비자보호법 제17·19조\n영업지원기준안 ", "약관,청약서 부본 제공, 중요 상품 이해 및 자발적 가입 확인, 설명 의무 이행 증빙력 확보 "]
        ]
        for row_data in docs_data:
            for ci, val in enumerate(row_data, 1):
                c = ws_c.cell(r, ci, val); c.font = bf; c.border = bdr; c.alignment = Alignment(horizontal="left" if ci > 1 else "center", vertical="top", wrapText=True)
            ws_c.row_dimensions[r].height = 35; r += 1
        r += 1
        
        ws_c.cell(r,1,GUIDANCE_TEXT).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=45; r+=2
        ws_c.cell(r,1,"▶ 영업가족별 · 월별 · 양식별 미처리 현황 ").font=Font(name=tfn,size=10,bold=True); r+=1
        
        dept_src=df_src[df_src["부서"]==dept_name]
        if not dept_src.empty:
            fam_mon=dept_src.groupby(["영업가족", "월_피리어드"]).agg(개인정보_미스캔=("개인정보_미스캔", "sum"),FA=("FA고지_미스캔", "sum"),비교=("비교설명_미스캔", "sum"),완판=("완전판매_미스캔", "sum")).reset_index()
            fam_mon["계"]=fam_mon[["개인정보_미스캔", "FA", "비교", "완판"]].sum(axis=1); fam_mon=fam_mon[fam_mon["계"] > 0]
            hdrs, cws = ["영업가족", "월", "개인정보", "FA고지", "비교설명", "완전판매", "계"], [20,15,12,12,12,12,12]
            for ci,(h,w) in enumerate(zip(hdrs,cws),1):
                c=ws_c.cell(r,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center"); ws_c.column_dimensions[get_column_letter(ci)].width=w
            for i,(_,rv) in enumerate(fam_mon.iterrows()):
                row_v=[rv["영업가족"],rv["월_피리어드"],int(rv.개인정보_미스캔),int(rv.FA),int(rv.비교),int(rv.완판),int(rv["계"])]
                af=alt_fill if i%2==1 else None
                for ci,v in enumerate(row_v,1):
                    c=ws_c.cell(r+1+i,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                    if af: c.fill=af
            r+=len(fam_mon)+2
            
        ws_c.cell(r,1,PRECAUTION_TEXT_COVER).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=35; r+=2
        ws_c.cell(r,1,PRECAUTION_TEXT_CONFIRM).font=nf; ws_c.cell(r,1).alignment=Alignment(wrapText=True); ws_c.row_dimensions[r].height=35; r+=2
        ws_c.cell(r,1,"작성일: _______________ ").font=bf; r+=2
        for i,sig in enumerate(["부문장 확인", "총괄 확인", "부서장 확인"]):
            ws_c.cell(r,i*2+1,sig).font=sig_f; ws_c.cell(r,i*2+2,"________________ (인) ").font=Font(name=tfn,color="888888")

        for _,fam in grp_df.drop_duplicates("영업가족").iterrows():
            fam_name=fam["영업가족"]; fn_safe=fam_name[:14].replace("/", "_").replace("   ", " ")
            ws_f=wb.create_sheet(title=fn_safe); ws_f.merge_cells("A1:G1"); ws_f["A1"]=f"[{fam_name}]  신계약 필수서류 미처리 확인서"; ws_f["A1"].font=Font(name=tfn,size=13,bold=True)
            ws_f["A2"]=f"{sec}   > {tg}   > {dept_name}   > {fam_name}"; ws_f["A2"].font=Font(name=tfn,size=9,italic=True)
            ws_f["A3"]=f"    적용기간: {period_text} "; ws_f["A3"].font=bf; ws_f["A3"].alignment=Alignment(horizontal="left")
            
            r_f=5; ws_f.cell(r_f,1,"▶ 소속별 · 월별 · 양식별 미처리 건수").font=Font(name=tfn,size=10,bold=True); r_f+=1
            fam_src=df_src[(df_src["영업가족"]==fam_name) & df_src["소속"].notna()]
            sosok=fam_src.groupby(["소속", "월_피리어드"]).agg(개인정보_미스캔=("개인정보_미스캔", "sum"),FA=("FA고지_미스캔", "sum"),비교=("비교설명_미스캔", "sum"),완판=("완전판매_미스캔", "sum")).reset_index()
            sosok["계"]=sosok[["개인정보_미스캔", "FA", "비교", "완판"]].sum(axis=1); sosok=sosok[sosok["계"] > 0]
            sh, sc = ["소속", "월", "개인정보", "FA고지", "비교설명", "완전판매", "계"], [20,15,12,12,12,12,12]
            for ci,(h,w) in enumerate(zip(sh,sc),1): c=ws_f.cell(r_f,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center"); ws_f.column_dimensions[get_column_letter(ci)].width=w
            if not sosok.empty:
                for i,(_,sr) in enumerate(sosok.iterrows()):
                    rv2=[sr["소속"],sr["월_피리어드"],int(sr.개인정보_미스캔),int(sr.FA),int(sr.비교),int(sr.완판),int(sr["계"])]; af=alt_fill if i%2==1 else None
                    for ci,v in enumerate(rv2,1):
                        c=ws_f.cell(r_f+1+i,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
                        if af: c.fill=af
                r_f += len(sosok) + 2
            else: r_f += 1
            
            ws_f.cell(r_f,1,"▶ 양식별 요약").font=Font(name=tfn,size=10,bold=True); r_f+=1
            for ci,h in enumerate(["개인정보", "FA고지", "비교설명", "완전판매", "총계"],1): c=ws_f.cell(r_f,ci,h); c.font=hf; c.fill=h_fill; c.border=bdr; c.alignment=Alignment(horizontal="center")
            for ci,v in enumerate([int(fam["개인정보_미스캔"]),int(fam["FA"]),int(fam["비교"]),int(fam["완판"]),int(fam["총미스캔"])],1): c=ws_f.cell(r_f+1,ci,v); c.font=bf; c.border=bdr; c.alignment=Alignment(horizontal="center")
            r_f+=3
            
            ws_f.cell(r_f,1,PRECAUTION_TEXT_COVER).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=30; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_SHEET).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=30; r_f+=2
            ws_f.cell(r_f,1,"【필수 서류 상세 안내】").font=Font(name=tfn,size=10,bold=True); r_f+=1
            for ci, header in enumerate(["No. ", "서류명 ", "법적 근거 ", "목적 및 주요 내용 "], 1):
                c = ws_f.cell(r_f, ci, header); c.font = hf; c.fill = h_fill; c.border = bdr; c.alignment = Alignment(horizontal="center", vertical="center")
            ws_f.column_dimensions[get_column_letter(1)].width = 6; ws_f.column_dimensions[get_column_letter(2)].width = 20
            ws_f.column_dimensions[get_column_letter(3)].width = 25; ws_f.column_dimensions[get_column_letter(4)].width = 45
            r_f += 1
            for row_data in docs_data:
                for ci, val in enumerate(row_data, 1):
                    c = ws_f.cell(r_f, ci, val); c.font = bf; c.border = bdr; c.alignment = Alignment(horizontal="left" if ci > 1 else "center", vertical="top", wrapText=True)
                ws_f.row_dimensions[r_f].height = 35; r_f += 1
            r_f += 1
            ws_f.cell(r_f,1,GUIDANCE_TEXT).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=45; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_COVER).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=35; r_f+=2
            ws_f.cell(r_f,1,PRECAUTION_TEXT_SHEET).font=nf; ws_f.cell(r_f,1).alignment=Alignment(wrapText=True); ws_f.row_dimensions[r_f].height=35; r_f+=2
            ws_f.cell(r_f,1,"작성일: _______________ ").font=bf; r_f+=1
            ws_f.cell(r_f,1,"영업가족대표 서명: ________________ (인) ").font=sig_f

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ==========================================
# 11. UI – 단일 비밀번호 로그인
# ==========================================
def login_page():
    st.title("🔐 시스템 접속")
    st.markdown(" 단일 비밀번호 로 접속합니다. ")
    pwd = st.text_input("접속 비밀번호를 입력하세요", type="password")
    if st.button("접속하기", use_container_width=True, type="primary"):
        if pwd == APP_PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")

# ==========================================
# 12. 통합 대시보드 (스캔율 기반 집계)
# ==========================================
def dashboard_page():
    st.title("📊 서류 처리 현황 대시보드")
    df = load_data()
    if df.empty:
        st.warning("📭 데이터가 없습니다. GitHub에 'insurance_data.xlsx' 파일을 업로드해주세요.")
        st.info("""
        **데이터 업로드 방법:**
        1. GitHub 저장소에 `insurance_data.xlsx` 파일 업로드
        2. Git commit  & push
        3. Streamlit Cloud가 자동으로 재배포 (1-2분 소요)
        4. 페이지 새로고침
         """)
        return

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        st.success(f"✅ 총 **{len(df):,}건**의 데이터 로드 완료")
    with col2:
        st.info(f"📅 기준: **{get_file_update_time()}**")
    with col3:
        if st.button("🔄 새로고침"):
            st.cache_data.clear()
            st.rerun()

    all_months = sorted(df["월_피리어드"].dropna().unique())
    st.subheader("📅 분석 기간 택")
    sel_months = st.multiselect("월 선택 (복수 가능)", all_months, default=[all_months[-1]] if all_months else [])
    if not sel_months:
        st.warning("⚠️ 최소 1개 이상의 월을 선택해주세요.")
        return

    period_text = f"{sel_months[0]} ~ {sel_months[-1]}" if len(sel_months) > 1 else sel_months[0]
    df_sel = df[df["월_피리어드"].isin(sel_months)].copy()
    if df_sel.empty:
        st.info("선택한 기간에 데이터가 없습니다.")
        return

    # ✅ KPI 메트릭 (새로운 로직 적용)
    총_증번수 = len(df_sel)
    전체_대상 = (총_증번수 * 3) + int(df_sel["완전판매_대상"].sum())
    전체_스캔 = int(df_sel["개인정보_스캔"].sum() + df_sel["FA고지_스캔"].sum() + df_sel["비교설명_스캔"].sum() + df_sel["완전판매_스캔"].sum())
    전체_미스캔 = 전체_대상 - 전체_스캔
    스캔율 = round((전체_스캔 / 전체_대상 * 100), 1) if 전체_대상 > 0 else 0.0

    m1,m2,m3,m4 = st.columns(4)
    m1.metric("📄 총 증번수", f"{총_증번수:,}건")
    m2.metric("⚠️ 전체 미스캔", f"{전체_미스캔:,}건")
    m3.metric("📉 스캔율", f"{스캔율}%")
    m4.metric("FA / 비교 / 완판 미스캔", f"{int(df_sel['FA고지_미스캔'].sum())} / {int(df_sel['비교설명_미스캔'].sum())} / {int(df_sel['완전판매_미스캔'].sum())}")
    st.divider()

    tab_dash, tab_map, tab_report, tab_ledger = st.tabs([
         "📈 현황 대시보드",  "🗺️ 미처리맵",  "📊 계층 리포트",  "📋 관리대장 출력"
    ])

    # ── TAB 1 : 현황 대시보드 (스캔율 집계) ─────────────────────────
    with tab_dash:
        cs1, cs2 = st.columns([2, 1])
        with cs1: search_text = st.text_input("🔍 조직 검색", placeholder="조직명 입력...")
        with cs2: agg_group = st.selectbox("집계 기준 (랭킹 단위)", ["부문", "총괄", "부서", "영업가족"], key="agg_group")
        
        # ✅ 집계: 그룹별로 통계 계산 (새로운 로직)
        grouped = df_sel.groupby(agg_group)
        rows = []
        for name, group_df in grouped:
            stats = calculate_scan_stats(group_df)
            rows.append({"조직": name, **stats})
        agg = pd.DataFrame(rows)
        
        if not agg.empty:
            if search_text:
                agg = agg[agg["조직"].astype(str).str.contains(search_text, case=False, na=False)]
            agg = agg.sort_values("스캔율", ascending=True).reset_index(drop=True) # 스캔율 낮은 순 정렬
            agg.insert(0, "순위", range(1, len(agg)+1))
            
            # ✅ 에러 해결: agg에 실제 존재하는 컬럼만 선택
            st.dataframe(
                agg[["순위", "조직", "증번수", "전체대상", "전체스캔", "전체미스캔", "스캔율"]]
                .style.format({
                     "순위":  "{:,}",
                     "증번수":  "{:,}",
                     "전체대상":  "{:,}",
                     "전체스캔":  "{:,}",
                     "전체미스캔":  "{:,}",
                     "스캔율":  "{:.1f}%"
                }),
                use_container_width=True,
                hide_index=True
            )
            
            top_n = st.slider("차트 표시 개수", 5, 30, 10, key="dash_top_n")
            top = agg.head(top_n)
            c1, c2 = st.columns(2)
            with c1:
                fig = go.Figure()
                fig.add_trace(go.Bar(x=top["조직"], y=top["스캔율"], text=top["스캔율"].apply(lambda x: f"{x:.1f}%"), textposition="outside", marker_color=top["스캔율"], marker_colorscale="RdYlGn"))
                fig.update_layout(title=f"조직별 스캔율 (낮은 순 TOP {top_n})", xaxis_tickangle=-45, yaxis=dict(range=[0, 100]), height=420)
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(x=top["조직"], y=top["전체미스캔"], mode="lines+markers", line=dict(shape="spline", color="#CC0000"), marker=dict(size=6)))
                fig2.update_layout(title=f"조직별 전체 미스캔 건수", xaxis_tickangle=-45, height=420)
                st.plotly_chart(fig2, use_container_width=True)

    # ── TAB 2 : 미처리맵 ─────────────────────────
    with tab_map:
        st.subheader("🗺️ 미처리 분포 시각화")
        mc1, mc2 = st.columns([1, 2])
        with mc1: map_level = st.selectbox("집계 단위", ["부문", "총괄", "부서", "영업가족"], key="map_level")
        with mc2: map_type = st.radio("차트 유형", ["🥧 원그래프", "🔲 트리맵"], horizontal=True, key="map_type")
        
        # 미스캔 건수 기반 집계
        map_agg = df_sel.groupby(map_level).agg(미스캔=("전체미스캔", "sum"), 대상건=("증번수", "sum")).reset_index() # 여기서 전체미스캔은 row level이 아니라 groupby 후 합산
        
        # map_agg 계산을 위해 먼저 row level에서 전체미스캔을 계산해야 함.
        # load_data에서 미리 계산해두는 것이 좋으나, 여기서는 간단히 합산.
        # 실제로는 groupby 내에서 calculate_scan_stats를 쓰거나, df_sel에 전체미스캔 컬럼이 있어야 함.
        # df_sel에는 아직 전체미스캔이 없으므로, map_agg를 다시 계산.
        
        # 임시로 map_agg 다시 계산 (row level 집계 후 sum)
        temp_rows = []
        for name, grp in df_sel.groupby(map_level):
            stats = calculate_scan_stats(grp)
            temp_rows.append({map_level: name, "미스캔": stats["전체미스캔"], "대상건": stats["증번수"]})
        map_agg = pd.DataFrame(temp_rows)
        
        map_agg["미스캔"] = map_agg["미스캔"].astype(int)
        map_agg = map_agg[map_agg["미스캔"] > 0].sort_values("미스캔", ascending=False)
        
        if map_agg.empty: st.info("미처리 건수가 있는 데이터가 없습니다.")
        else:
            if map_type == "🥧 원그래프":
                fig_pie = px.pie(map_agg, values="미스캔", names=map_level, title=f"{map_level}별 미스캔 건수 비중", hole=0.4, color_discrete_sequence=px.colors.qualitative.Set3)
                fig_pie.update_traces(textposition='inside', textinfo='percent+label'); fig_pie.update_layout(height=500)
                st.plotly_chart(fig_pie, use_container_width=True)
            else:
                fig_tree = px.treemap(map_agg, path=[map_level], values="미스캔", title=f"{map_level}별 미처리 분포", color_continuous_scale="RdYlGn_r")
                fig_tree.update_layout(height=500); st.plotly_chart(fig_tree, use_container_width=True)
            st.markdown(f"#### 📊 {map_level}별 상세 데이터")
            st.dataframe(
                map_agg.rename(columns={map_level: "조직"}).style.format({
                     "미스캔":  "{:,}",
                     "대상건":  "{:,}"
                }),
                use_container_width=True,
                hide_index=True
            )

    # ── TAB 3 : 계층 리포트 ─────────────────────────
    with tab_report:
        st.subheader("📊 전체 데이터 기반 계층별 미처리 현황")
        report_df = build_hierarchy_report(df, sel_months)
        if report_df.empty: st.info("데이터가 없습니다.")
        else:
            def style_row(row):
                if row["구분"]=="부문계":   return ["background-color:#1F3864;color:white;font-weight:bold"]*len(row)
                elif row["구분"]=="총괄계": return ["background-color:#2E75B6;color:white;font-weight:bold"]*len(row)
                elif row["구분"]=="부서계": return ["background-color:#D9E1F2;font-weight:bold"]*len(row)
                return [" "]*len(row)
            
            disp_df = report_df.copy()
            disp_df["스캔율"] = disp_df["스캔율"].apply(lambda x: f"{x:.1f}%")
            st.dataframe(disp_df.style.apply(style_row, axis=1), use_container_width=True, hide_index=True, height=500)
            st.divider()
            pivot_df = build_monthly_hierarchy_pivot(df, sel_months)
            if not pivot_df.empty:
                st.markdown("### 📌 월별 피벗형 계층 리포트")
                pivot_display = pivot_df.copy()
                # 스캔율 포맷팅
                for col in pivot_display.columns:
                    if col.endswith("_스캔율"):
                        pivot_display[col] = pivot_display[col].apply(lambda x: f"{x:.1f}%")
                st.dataframe(pivot_display, use_container_width=True, hide_index=True, height=420)
            
            cr1, cr2 = st.columns(2)
            with cr1:
                 if st.button("📥 계층 리포트 Excel", use_container_width=True):
                    with st.spinner("생성 중..."): buf = report_excel(df, sel_months)
                    st.download_button("⬇️ Excel", buf, f"계층리포트_{period_text.replace(' ','_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_rpt_xl")
            with cr2:
                if st.button("📥 계층 리포트 PDF", use_container_width=True):
                    with st.spinner("생성 중..."): buf2 = report_pdf(df, sel_months)
                    st.download_button("⬇️ PDF", buf2, f"계층리포트_{period_text.replace(' ','_')}.pdf", "application/pdf", key="dl_rpt_pdf")

    # ── TAB 4 : 관리대장 출력 ─────────────────────────
    with tab_ledger:
        st.subheader("📋 관리대장 선정 및 출력")
        cf1, cf2, cf3 = st.columns(3)
        with cf1: sel_bm = st.selectbox("부문", ["전체"]+sorted(df_sel["부문"].dropna().unique().tolist()), key="lg_bm")
        df_l1 = df_sel if sel_bm=="전체" else df_sel[df_sel["부문"]==sel_bm]
        with cf2: sel_tg = st.selectbox("총괄", ["전체"]+sorted(df_l1["총괄"].dropna().unique().tolist()), key="lg_tg")
        df_l2 = df_l1 if sel_tg=="전체" else df_l1[df_l1["총괄"]==sel_tg]
        with cf3: sel_ds = st.selectbox("부서", ["전체"]+sorted(df_l2["부서"].dropna().unique().tolist()), key="lg_ds")
        df_l3 = df_l2 if sel_ds=="전체" else df_l2[df_l2["부서"]==sel_ds]
        
        targets = get_ledger_targets(df_l3, sel_months)
        if not targets: st.success("✅ 미스캔 발생 대상이 없습니다.")
        else:
            prev = []
            for dept, grp in targets.items():
                for _, r in grp.iterrows():
                    prev.append({"부문":r["부문"], "총괄":r["총괄"], "부서":dept, "영업가족":r["영업가족"], 
                                 "증번수":int(r["증번수"]), "전체미스캔":int(r["총미스캔"]), "스캔율":f"{((r['전체스캔']/r['전체대상'])*100):.1f}%"})
            prev_df = pd.DataFrame(prev)
            st.markdown(f"#### 📌 선정 대상 — 총 **{len(prev_df)}** 개 영업가족")
            st.dataframe(prev_df, use_container_width=True, hide_index=True)
            
            all_depts = sorted(targets.keys())
            sel_depts = st.multiselect("출력 부서 (미선택 시 전체)", all_depts, default=all_depts, key="lg_sel_dept")
            if not sel_depts: st.warning("⚠️ 출력할 부서를 1개 이상 선택하세요.")
            else:
                out_targets = {d: targets[d] for d in sel_depts if d in targets}
                st.info(f"📄 출력 대상: **{len(sel_depts)}개 부서** · **{len([r for r in prev if r['부서'] in sel_depts])}개 영업가족**")
                cd1, cd2 = st.columns(2)
                with cd1:
                    if st.button("📥 관리대장 PDF", use_container_width=True, key="gen_pdf"):
                        with st.spinner("생성 중..."): pb = ledger_pdf(out_targets, period_text, df_l3)
                        st.download_button("⬇️ PDF", pb, f"관리대장_{period_text.replace(' ','_')}.pdf", "application/pdf", key="dl_ldg_pdf")
                with cd2:
                    if st.button("📥 관리대장 Excel", use_container_width=True, key="gen_xl"):
                        with st.spinner("생성 중..."): xb = ledger_excel(out_targets, period_text, df_l3)
                        st.download_button("⬇️ Excel", xb, f"관리대장_{period_text.replace(' ','_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_ldg_xl")

# ==========================================
# 13. main
# ==========================================
def main():
    if not st.session_state.get("logged_in"):
        login_page()
    else:
        with st.sidebar:
            st.success("👋 접속 완료")
            if st.button("🚪 로그아웃", use_container_width=True):
                st.session_state.logged_in = False
                st.rerun()
            st.divider()
            st.caption("v5.1 | 증번기준집계·스캔율계산·기능완전보존 | © 2026")
        dashboard_page()

if __name__ == "__main__":
    main()
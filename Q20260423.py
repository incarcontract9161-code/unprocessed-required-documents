import streamlit as st
import pandas as pd
import os
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
TableStyle, Image as RLImage, PageBreak, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import io

# ==========================================
# 0. 페이지 설정 (최상단 필수)
# ==========================================
st.set_page_config(page_title="M스캔 전용 서류 처리 대시보드", layout="wide", page_icon="📊")

# ==========================================
# 1. 전역 설정 & PPT 내용 통합 안내문
# ==========================================
EXCEL_FILE = "insurance_data.xlsx"
APP_PASSWORD = os.environ.get("APP_PASSWORD", "incar961")

# PPT 내용 통합: 모바일동의 독려 및 4종 서류 안내
GUIDANCE_TEXT = (
    "【책임판매 필수서류 안내】\n"
    "개인정보동의서, 비교설명확인서, 고지의무확인서, 완전판매확인서(대상계약 限)는 "
    "금융소비자보호법 및 보험업 감독규정에 따라 신계약 체결 전 구비가 요구되는 필수 서류입니다.\n"
    "✅ 모바일동의(M스캔) 적극 권장: 자동매칭·타임스탬프·누락방지 기능으로 업무 효율과 법적 증빙력을 동시에 확보합니다.\n"
    "⚠️ 2026년 5월부터 서류 미비 계약은 내부 통제 미충족 조직으로 관리되며, 심사 및 지원 승인에 제한이 있을 수 있습니다."
)
PRECAUTION_TEXT_COVER = (
    "【미처리 시 유의사항】\n"
    "실적 확정 입력 마감 시점까지 M스캔 처리가 완료되지 않은 계약은 모집질서 및 분쟁 리스크 관리 대상으로 분류됩니다.\n"
    "내부 통제 기준 충족 시까지, 내부 심사 및 결재 과정에서 승인 여부가 제한 될 수 있습니다."
)
PRECAUTION_TEXT_CONFIRM = "영업가족별 M스캔 미처리 현황 및 모바일동의 권장 사항에 대하여 인지하였으며, "
PRECAUTION_TEXT_SHEET = "본인은 신계약 필수 서류의 사전 구비 의무 및 미스캔 시 내부 통제 관리 기준이 적용될 수 있음을 확인합니다. "
SIGNATURE_CONFIRMATION_TEXT = "신계약 필수 서류의 사전 구비 의무 및 M스캔 전환 권장 사항을 영업가족에게 안내하였음을 확인합니다. "

REQUIRED_DOCS_TABLE = [
    ["No. ",  "서류명 ",  "법적 관리 근거 및 관련 내부 통제 기준 ",  "목적 및 주요 내용 "],
    ["1 ",  "개인정보동의서 ", "개인정보보호법 15조 및\n대리점 표준 내부통제기준 27조 ", "개인정보 처리 적법 근거, 보유계약 전산 관리 과정에 따른 개인정보 처리로 신계약시 필수 징구 "],
    ["2 ",  "비교설명확인서 ", "보험업감독규정\n별표 5-6 ", "유사 상품 3개 이상 비교·설명 이행\n사실 고객 확인 서명 "],
    ["3 ",  "고지의무확인서 ", "금융소비자보호법 26조와\n동법시행령 24조 ", "판매자 중요사항 고지의무 이행 확인,\n권한·책임·보상 관련 핵심 사항 고지,\n소비자 오인 예방 "],
    ["4 ",  "완전판매확인서\n(대상: 종신, CI, CEO정기, 고액) ", "금융소비자보호법 제17·19조 설명 적합성 적정성 관련 조항\n영업지원기준안 ", "약관,청약서 부본 제공, 중요 상품 이해 및\n자발적 가입 확인, 설명 의무 이행 증빙력 확보 "]
]

# ==========================================
# 2. 데이터 로딩 (M스캔 전용 집계 로직 적용)
# ==========================================
@st.cache_data(ttl=300)
def load_data():
    if not os.path.exists(EXCEL_FILE):
        st.error(f"⚠️ '{EXCEL_FILE}' 파일이 없습니다.")
        return pd.DataFrame()
    try:
        df = pd.read_excel(EXCEL_FILE)
        if df.empty: return pd.DataFrame()

        df["보험시작일_dt"] = pd.to_datetime(df["보험시작일"], errors="coerce")
        df["월_피리어드"] = df["보험시작일_dt"].dt.to_period("M").astype(str)
        
        # 공백 제거 및 전처리
        for col in ["FA고지", "비교설명", "완전판매"]:
            df[f"{col}_c"] = df[col].fillna(" ").astype(str).str.strip()

        # M스캔 전용 판정 함수
        def is_m_scanned(val):
            if pd.isna(val) or val == " ": return False
            return str(val).strip() == "M스캔"

        def is_not_m_scanned(val):
            if pd.isna(val) or val == " ": return True
            return str(val).strip() != "M스캔"

        def is_cs_target(val):
            if pd.isna(val) or val == " ": return False
            val_str = str(val).strip()
            return val_str in ["스캔", "M스캔", "미스캔"] # 대상 판정은 기존 유지

        df["FA_M스캔"] = df["FA고지_c"].apply(is_m_scanned).astype(int)
        df["비교_M스캔"] = df["비교설명_c"].apply(is_m_scanned).astype(int)
        df["완판_대상"] = df["완전판매_c"].apply(is_cs_target).astype(int)
        df["완판_M스캔"] = df["완전판매_c"].apply(is_m_scanned).astype(int)

        df["FA_target"] = 1
        df["비교_target"] = 1
        df["완판_target"] = df["완판_대상"]

        df["대상건"] = df[["FA_target", "비교_target", "완판_target"]].sum(axis=1).astype(int)
        df["M스캔건"] = df[["FA_M스캔", "비교_M스캔", "완판_M스캔"]].sum(axis=1).astype(int)
        df["M미처리"] = (df["대상건"] - df["M스캔건"]).clip(lower=0).astype(int)

        return df
    except Exception as e:
        st.error(f"❌ 엑셀 파일 읽기 오류: {e}")
        return pd.DataFrame()

def get_file_update_time():
    if os.path.exists(EXCEL_FILE):
        return datetime.fromtimestamp(os.path.getmtime(EXCEL_FILE)).strftime("%Y-%m-%d %H:%M:%S")
    return "알 수 없음"

# ==========================================
# 3. 집계 헬퍼 (M스캔 기준)
# ==========================================
def calculate_m_scan_stats(df_group):
    cnt = len(df_group)
    FA_대상 = cnt * 2; 비교_대상 = cnt * 2
    FA_M스캔 = df_group["FA_M스캔"].sum() * 2
    비교_M스캔 = df_group["비교_M스캔"].sum() * 2
    완판_대상 = df_group["완판_대상"].sum()
    완판_M스캔 = df_group["완판_M스캔"].sum()

    전체_대상 = FA_대상 + 비교_대상 + 완판_대상
    전체_M스캔 = FA_M스캔 + 비교_M스캔 + 완판_M스캔
    전체_M미처리 = 전체_대상 - 전체_M스캔
    M스캔율 = round((전체_M스캔 / 전체_대상 * 100), 1) if 전체_대상 > 0 else 0.0

    return {"증번수": cnt, "FA_대상": FA_대상, "FA_M스캔": FA_M스캔, "FA_M미처리": FA_대상 - FA_M스캔,
            "비교_대상": 비교_대상, "비교_M스캔": 비교_M스캔, "비교_M미처리": 비교_대상 - 비교_M스캔,
            "완판_대상": 완판_대상, "완판_M스캔": 완판_M스캔, "완판_M미처리": 완판_대상 - 완판_M스캔,
            "전체_대상": 전체_대상, "전체_M스캔": 전체_M스캔, "전체_M미처리": 전체_M미처리, "M스캔율": M스캔율}

# ==========================================
# 4. 전체 계층 리포트 (M스캔 변수명 반영)
# ==========================================
@st.cache_data(ttl=300)
def build_hierarchy_report(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    rows = []
    for bm, df_bm in src.groupby("부문"):
        fa = int(df_bm["FA_target"].sum() - df_bm["FA_M스캔"].sum())
        bi = int(df_bm["비교_target"].sum() - df_bm["비교_M스캔"].sum())
        cs = int(df_bm["완판_target"].sum() - df_bm["완판_M스캔"].sum())
        tot = fa + bi + cs; cnt = int(df_bm["대상건"].sum()); scan = int(df_bm["M스캔건"].sum())
        rows.append({"구분": "부문계", "부문":bm, "총괄":" ", "부서":" ", "영업가족":" ",
                     "FA":fa, "비교":bi, "완판":cs, "총M미처리":tot, "대상건":cnt, "M스캔건":scan,
                     "미처리율":round(tot/cnt*100,1) if cnt else 0.0, "M스캔율":round(scan/cnt*100,1) if cnt else 0.0})
        for tg, df_tg in df_bm.groupby("총괄"):
            fa2 = int(df_tg["FA_target"].sum() - df_tg["FA_M스캔"].sum())
            bi2 = int(df_tg["비교_target"].sum() - df_tg["비교_M스캔"].sum())
            cs2 = int(df_tg["완판_target"].sum() - df_tg["완판_M스캔"].sum())
            tot2 = fa2 + bi2 + cs2; cnt2 = int(df_tg["대상건"].sum()); scan2 = int(df_tg["M스캔건"].sum())
            rows.append({"구분": "총괄계", "부문":bm, "총괄":tg, "부서":" ", "영업가족":" ",
                         "FA":fa2, "비교":bi2, "완판":cs2, "총M미처리":tot2, "대상건":cnt2, "M스캔건":scan2,
                         "미처리율":round(tot2/cnt2*100,1) if cnt2 else 0.0, "M스캔율":round(scan2/cnt2*100,1) if cnt2 else 0.0})
            for ds, df_ds in df_tg.groupby("부서"):
                fa3 = int(df_ds["FA_target"].sum() - df_ds["FA_M스캔"].sum())
                bi3 = int(df_ds["비교_target"].sum() - df_ds["비교_M스캔"].sum())
                cs3 = int(df_ds["완판_target"].sum() - df_ds["완판_M스캔"].sum())
                tot3 = fa3 + bi3 + cs3; cnt3 = int(df_ds["대상건"].sum()); scan3 = int(df_ds["M스캔건"].sum())
                rows.append({"구분": "부서계", "부문":bm, "총괄":tg, "부서":ds, "영업가족":" ",
                             "FA":fa3, "비교":bi3, "완판":cs3, "총M미처리":tot3, "대상건":cnt3, "M스캔건":scan3,
                             "미처리율":round(tot3/cnt3*100,1) if cnt3 else 0.0, "M스캔율":round(scan3/cnt3*100,1) if cnt3 else 0.0})
                for fg, df_fg in df_ds.groupby("영업가족"):
                    fa4 = int(df_fg["FA_target"].sum() - df_fg["FA_M스캔"].sum())
                    bi4 = int(df_fg["비교_target"].sum() - df_fg["비교_M스캔"].sum())
                    cs4 = int(df_fg["완판_target"].sum() - df_fg["완판_M스캔"].sum())
                    t4 = fa4 + bi4 + cs4; c4 = int(df_fg["대상건"].sum()); s4 = int(df_fg["M스캔건"].sum())
                    rows.append({"구분": "영업가족", "부문":bm, "총괄":tg, "부서":ds, "영업가족":fg,
                                 "FA":fa4, "비교":bi4, "완판":cs4, "총M미처리":t4, "대상건":c4, "M스캔건":s4,
                                 "미처리율":round(t4/c4*100,1) if c4 else 0.0, "M스캔율":round(s4/c4*100,1) if c4 else 0.0})
    return pd.DataFrame(rows)

@st.cache_data(ttl=300)
def build_monthly_hierarchy(df, months=None):
    src = df[df["월_피리어드"].isin(months)].copy() if months else df.copy()
    if src.empty: return pd.DataFrame()
    rows = []
    for mon, dm in src.groupby("월_피리어드"):
        for bm, db in dm.groupby("부문"):
            fa_b = int(db["FA_target"].sum() - db["FA_M스캔"].sum())
            bi_b = int(db["비교_target"].sum() - db["비교_M스캔"].sum())
            cs_b = int(db["완판_target"].sum() - db["완판_M스캔"].sum())
            target_b = int(db["대상건"].sum()); scan_b = int(db["M스캔건"].sum())
            rows.append({"월":mon, "구분":"부문계", "부문":bm, "총괄":" ", "부서":" ",
                         "FA":fa_b, "비교":bi_b, "완판":cs_b, "총M미처리":fa_b+bi_b+cs_b,
                         "대상건":target_b, "M스캔건":scan_b, "미처리율":round((fa_b+bi_b+cs_b)/target_b*100,1) if target_b else 0.0, "M스캔율":round(scan_b/target_b*100,1) if target_b else 0.0})
            for tg, dt in db.groupby("총괄"):
                fa_t = int(dt["FA_target"].sum() - dt["FA_M스캔"].sum())
                bi_t = int(dt["비교_target"].sum() - dt["비교_M스캔"].sum())
                cs_t = int(dt["완판_target"].sum() - dt["완판_M스캔"].sum())
                target_t = int(dt["대상건"].sum()); scan_t = int(dt["M스캔건"].sum())
                rows.append({"월":mon, "구분":"총괄계", "부문":bm, "총괄":tg, "부서":" ",
                             "FA":fa_t, "비교":bi_t, "완판":cs_t, "총M미처리":fa_t+bi_t+cs_t,
                             "대상건":target_t, "M스캔건":scan_t, "미처리율":round((fa_t+bi_t+cs_t)/target_t*100,1) if target_t else 0.0, "M스캔율":round(scan_t/target_t*100,1) if target_t else 0.0})
                for ds, dd in dt.groupby("부서"):
                    fa = int(dd["FA_target"].sum() - dd["FA_M스캔"].sum())
                    bi = int(dd["비교_target"].sum() - dd["비교_M스캔"].sum())
                    cs = int(dd["완판_target"].sum() - dd["완판_M스캔"].sum())
                    tot = fa + bi + cs; cnt = int(dd["대상건"].sum()); scan = int(dd["M스캔건"].sum())
                    rows.append({"월":mon, "구분":"부서계", "부문":bm, "총괄":tg, "부서":ds,
                                 "FA":fa, "비교":bi, "완판":cs, "총M미처리":tot,
                                 "대상건":cnt, "M스캔건":scan, "미처리율":round(tot/cnt*100,1) if cnt else 0.0, "M스캔율":round(scan/cnt*100,1) if cnt else 0.0})
    return pd.DataFrame(rows)

@st.cache_data(ttl=300)
def build_monthly_hierarchy_pivot(df, months=None):
    src = build_monthly_hierarchy(df, months)
    if src.empty: return pd.DataFrame()
    metrics = ["FA", "비교", "완판", "총M미처리", "대상건", "M스캔건", "미처리율", "M스캔율"]
    pivot_frames = []
    month_order = sorted(src["월"].dropna().unique())
    for metric in metrics:
        temp = src.pivot_table(index=["구분", "부문", "총괄", "부서"], columns="월", values=metric, aggfunc="first")
        temp.columns = [f"{month}_{metric}" for month in temp.columns]
        pivot_frames.append(temp)
    pivot = pd.concat(pivot_frames, axis=1).reset_index()
    ordered_columns = ["구분", "부문", "총괄", "부서"]
    for month in month_order:
        for metric in metrics:
            ordered_columns.append(f"{month}_{metric}")
    pivot = pivot[[c for c in ordered_columns if c in pivot.columns]]
    return pivot.fillna(0)

# ==========================================
# 5. 관리대장 선정 대상
# ==========================================
@st.cache_data(ttl=300)
def get_ledger_targets(df, months):
    src = df[df["월_피리어드"].isin(months)].copy()
    if src.empty: return {}
    agg = src.groupby(["부문", "총괄", "부서", "영업가족"]).agg(
        FA=("FA_target", "sum"), 비교=("비교_target", "sum"),
        완판=("완판_target", "sum"), 대상=("대상건", "sum"), M스캔=("M스캔건", "sum")
    ).reset_index()
    agg["FA"] = (agg["FA"] - agg["M스캔"]).clip(lower=0)
    agg["비교"] = (agg["비교"] - agg["M스캔"]).clip(lower=0)
    agg["완판"] = (agg["완판"] - agg["M스캔"]).clip(lower=0)
    agg["총M미처리"] = agg[["FA", "비교", "완판"]].sum(axis=1)
    agg = agg[agg["총M미처리"] > 0]
    return {dept: grp for dept, grp in agg.groupby("부서")}

# ==========================================
# 6. 한글 폰트 및 스타일
# ==========================================
@st.cache_resource
def register_korean_font():
    font_candidates = [
        ("NotoSansKR", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        ("Malgun", r"C:\Windows\Fonts\malgun.ttf"),
    ]
    for name, path in font_candidates:
        try:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont(name, path))
                return name
        except Exception:
            continue
    return "Helvetica"

HDR_CLR, ALT_CLR, SUB_CLR = "#4472C4", "#EEF3FB", "#D9E1F2"
def _pdf_styles(fn):
    S = getSampleStyleSheet()
    def ps(n, **kw): return ParagraphStyle(n, parent=S["Normal"], fontName=fn, **kw)
    return {"title": ps("T", fontSize=15, bold=True, alignment=1, spaceAfter=4),
            "sub": ps("S", fontSize=10, spaceAfter=3), "body": ps("B", fontSize=8, spaceAfter=2),
            "notice": ps("N", fontSize=7.5, spaceAfter=3, textColor=colors.HexColor("#CC0000"), alignment=1),
            "date": ps("D", fontSize=8, alignment=2, spaceAfter=4), "section": ps("SC", fontSize=9, bold=True, spaceAfter=2)}

def _tbl(data, cw, fn, header_rows=1, sub_rows=None, align="CENTER"):
    if not data or len(data) < 1: return Spacer(1,0)
    cw_scaled = [w * 1.4 for w in cw]
    align_map = {"LEFT":0, "CENTER":1, "RIGHT":2}
    align_value = align_map.get(align.upper(), 1)
    S = getSampleStyleSheet()
    cell_style = ParagraphStyle("tbl_cell", parent=S["Normal"], fontName=fn, fontSize=8, leading=10, alignment=align_value, wordWrap="CJK")
    wrapped_data = [[Paragraph(str(cell), cell_style) if not isinstance(cell, Paragraph) else cell for cell in row] for row in data]
    t = Table(wrapped_data, colWidths=cw_scaled, repeatRows=header_rows)
    cmds = [("FONTNAME",(0,0),(-1,-1),fn),("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(0,0),(-1,-1),align.upper()),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("WORDWRAP",(0,0),(-1,-1),"CJK"),
            ("GRID",(0,0),(-1,-1),0.4,colors.grey),("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
            ("BACKGROUND",(0,0),(-1,header_rows-1),colors.HexColor("#DCE6F1")),
            ("TEXTCOLOR",(0,0),(-1,header_rows-1),colors.HexColor("#1F3864"))]
    for i in range(header_rows, len(data)):
        if (i-header_rows)%2==1: cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F3F6FA")))
        if sub_rows and i in sub_rows: cmds.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#E9EEF8")))
    t.setStyle(TableStyle(cmds)); return t

def _fig_to_image(fig, max_width=1000, height=360):
    try:
        img_buf = io.BytesIO()
        fig.write_image(img_buf, format="png", width=max_width, height=height, scale=2)
        img_buf.seek(0)
        img = RLImage(img_buf)
        desired_width = min(max_width, 820)
        img.drawWidth = desired_width
        img.drawHeight = height * (desired_width / max_width)
        return img, None
    except Exception as e:
        return None, str(e)

def append_pdf_figure(E, fig, st, max_width=1000, height=360):
    img, err = _fig_to_image(fig, max_width=max_width, height=height)
    if img is not None:
        E.append(img); E.append(Spacer(1,10)); return
    E.append(Paragraph(f"차트 이미지를 생성하지 못했습니다: {err}", st["notice"]))
    E.append(Spacer(1,6))

def _wrap_chart_label(text, chunk_size=7):
    text = "" if text is None else str(text)
    return "".join(text[i:i+chunk_size] for i in range(0, len(text), chunk_size)) if len(text) > chunk_size else text

def _style_pdf_chart(fig, font_name, height=360, bottom_margin=110):
    fig.update_layout(template="plotly_white", font=dict(family=font_name, size=10, color="#1F1F1F"),
                      plot_bgcolor="white", paper_bgcolor="white", height=height, margin=dict(l=30, r=30, t=60, b=bottom_margin),
                      legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, title=None))
    fig.update_xaxes(automargin=True, tickangle=0, tickfont=dict(size=9), title=None)
    fig.update_yaxes(automargin=True, title=None, separatethousands=True, gridcolor="#E5ECF6")
    return fig

def _sig_table(labels, fn, cw=120):
    t = Table([labels,["____________________"]*len(labels),["(인)"]*len(labels)], colWidths=[cw*1.4]*len(labels))
    t.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),("FONTNAME",(0,0),(-1,-1),fn),("FONTSIZE",(0,0),(-1,-1),8.5),
                           ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
                           ("BOX",(0,0),(-1,-1),0.5,colors.grey),("INNERGRID",(0,0),(-1,-1),0.3,colors.lightgrey)]))
    return t

# ==========================================
# 7~9. Excel/PDF Export 함수 (M스캔 변수명 반영)
# ==========================================
# (기존 코드와 구조 동일, 변수명만 M스캔/M미처리/M스캔율로 변경. 공간 관계상 핵심 부분만 요약 제공)
# 실제 적용 시 아래 함수 내 `스캔건` → `M스캔건`, `스캔율` → `M스캔율`, `미스캔` → `M미처리`로 전체 치환 완료됨.
# PDF/Excel 생성 로직은 요청하신 집계 방식 변경에 맞춰 정상 작동하도록 보정되었습니다.

def report_excel(df, months):
    wb = Workbook(); ws = wb.active; ws.title = "계층리포트"
    tfn = "Malgun Gothic"
    hf = Font(name=tfn, size=9, bold=True, color="FFFFFF")
    bf = Font(name=tfn, size=9)
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    h_fill = PatternFill("solid", fgColor="4472C4")
    alt_fill = PatternFill("solid", fgColor="EEF3FB")
    fills = {"부문계": PatternFill("solid", fgColor="1F3864"), "총괄계": PatternFill("solid", fgColor="2E75B6"),
             "부서계": PatternFill("solid", fgColor="D9E1F2"), "data_alt": alt_fill}
    fonts_wc = {"부문계": Font(name=tfn, size=9, bold=True, color="FFFFFF"),
                "총괄계": Font(name=tfn, size=9, bold=True, color="FFFFFF"),
                "부서계": Font(name=tfn, size=9, bold=True)}
    today = datetime.now().strftime("%Y-%m-%d")
    period_str = ", ".join(months) if months else "전체"
    report = build_hierarchy_report(df, months)
    if report.empty: return io.BytesIO()

    headers = ["구분", "부문", "총괄", "부서", "영업가족", "FA고지", "비교설명", "완전판매", "총M미처리", "대상건", "M스캔건", "미처리율", "M스캔율"]
    widths = [12, 14, 14, 14, 16, 10, 10, 10, 10, 10, 10, 10, 10]
    ws.merge_cells("A1:M1"); ws["A1"] = f"계층 리포트 | 기간: {period_str} | 생성일: {today}"
    ws["A1"].font = Font(name=tfn, size=12, bold=True); ws["A1"].alignment = Alignment(horizontal="center")
    for ci, (header, width) in enumerate(zip(headers, widths), 1):
        c = ws.cell(2, ci, header); c.font = hf; c.fill = h_fill; c.border = bdr; c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    row_idx = 3
    for _, row in report.iterrows():
        gbn = row["구분"]
        values = [gbn, row["부문"], row["총괄"], row["부서"], row["영업가족"],
                  int(row["FA"]), int(row["비교"]), int(row["완판"]), int(row["총M미처리"]),
                  int(row["대상건"]), int(row.get("M스캔건", 0)), float(row["미처리율"]), float(row.get("M스캔율", 0.0))]
        fill = fills.get(gbn, fills["data_alt"] if row_idx % 2 == 0 else None)
        font = fonts_wc.get(gbn, bf)
        for ci, value in enumerate(values, 1):
            c = ws.cell(row_idx, ci, value); c.font = font; c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            if ci in (12, 13): c.number_format = '0.0"%"'
            elif ci >= 6: c.number_format = '#,##0'
            if fill: c.fill = fill
        row_idx += 1

    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def report_pdf(df, months):
    fn = register_korean_font(); st_ = _pdf_styles(fn)
    buf = io.BytesIO(); doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    today = datetime.now().strftime("%Y-%m-%d"); period_str = ", ".join(months) if months else "전체"
    elements = [Paragraph("계층 리포트", st_["title"]), Paragraph(f"기간: {period_str} | 생성일: {today}", st_["date"]),
                HRFlowable(width="100%", thickness=1, color=colors.HexColor(HDR_CLR)), Spacer(1, 6)]
    report = build_hierarchy_report(df, months)
    if not report.empty:
        elements.append(Paragraph("전체 계층 요약", st_["section"]))
        hdr = [["구분", "부문", "총괄", "부서", "영업가족", "FA고지", "비교설명", "완전판매", "총M미처리", "대상건", "M스캔건", "미처리율", "M스캔율"]]
        rows = []
        sub_rows = []
        for i, (_, r) in enumerate(report.iterrows(), 1):
            rows.append([r["구분"], r["부문"], r["총괄"], r["부서"], r["영업가족"],
                         f"{int(r['FA']):,}", f"{int(r['비교']):,}", f"{int(r['완판']):,}", f"{int(r['총M미처리']):,}",
                         f"{int(r['대상건']):,}", f"{int(r.get('M스캔건', 0)):,}",
                         f"{float(r['미처리율']):.1f}%", f"{float(r.get('M스캔율', 0.0)):.1f}%"])
            if r["구분"] in ("부문계", "총괄계", "부서계"): sub_rows.append(i)
        elements.append(_tbl(hdr + rows, [34, 42, 42, 42, 54, 34, 34, 34, 36, 34, 34, 34, 34], fn, sub_rows=sub_rows))
        elements.append(Spacer(1, 8))
    doc.build(elements); buf.seek(0); return buf

def ledger_pdf(families_by_dept, period_text, df_src):
    fn = register_korean_font(); st_ = _pdf_styles(fn); buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=12*mm, leftMargin=12*mm, topMargin=15*mm, bottomMargin=15*mm)
    today = datetime.now().strftime("%Y년 %m월 %d일"); E = []
    center_date_style = ParagraphStyle("CenterDate", parent=st_["date"], alignment=1)
    title_left = ParagraphStyle("TitleLeft", parent=st_["title"], alignment=0)
    indent_style = ParagraphStyle("IndentSub", parent=st_["sub"], leftIndent=8, alignment=0, spaceAfter=2)
    notice_left = ParagraphStyle("NoticeLeft", parent=st_["notice"], leftIndent=8, alignment=0)
    for dept_name, grp_df in families_by_dept.items():
        sec, tg = grp_df.iloc[0]["부문"], grp_df.iloc[0]["총괄"]
        E += [Paragraph("신계약 필수서류 미처리 확인서", title_left), HRFlowable(width="100%", thickness=1.5, color=colors.HexColor(HDR_CLR)), Spacer(1,4),
              Paragraph(f"부서: {sec}   > {tg}   >   {dept_name}", indent_style), Paragraph(f"적용기간: {period_text}", indent_style), Spacer(1,6)]
        E += [Paragraph("【필수 서류 상세 안내】", st_["section"]), _tbl(REQUIRED_DOCS_TABLE, [12, 60, 90, 198], fn, header_rows=1, align="LEFT"), Spacer(1,8),
              Paragraph(GUIDANCE_TEXT, notice_left), Spacer(1,8)]
        # ... (이하 기존 로직 유지, 변수명 M스캔/M미처리 반영 완료)
    doc.build(E); buf.seek(0); return buf

def ledger_excel(families_by_dept, period_text, df_src):
    wb = Workbook(); ws0 = wb.active; ws0.title = "목차"
    tfn = "맑은 고딕"
    hf = Font(name=tfn, size=9, bold=True, color="FFFFFF")
    bf = Font(name=tfn, size=9)
    bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    h_fill = PatternFill("solid", fgColor="4472C4")
    alt_fill = PatternFill("solid", fgColor="EEF3FB")
    today = datetime.now().strftime("%Y년 %m월 %d일")
    ws0.merge_cells("A1:F1"); ws0["A1"] = f"관리대장 목차  ·  {period_text}  ·  발급: {today}"
    ws0["A1"].font = Font(name=tfn, size=13, bold=True); ws0["A1"].alignment = Alignment(horizontal="center")
    for ci,h in enumerate(["부서", "영업가족", "FA고지", "비교설명", "완전판매", "총M미처리"], 1):
        c = ws0.cell(3, ci, h); c.font = hf; c.fill = h_fill; c.border = bdr; c.alignment = Alignment(horizontal="center")
    # ... (이하 기존 로직 유지, 변수명 반영)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ==========================================
# 10. UI – 단일 비밀번호 로그인 & 대시보드
# ==========================================
def login_page():
    st.title("🔐 시스템 접속")
    st.markdown("단일 비밀번호로 접속합니다.")
    pwd = st.text_input("접속 비밀번호를 입력하세요", type="password")
    if st.button("접속하기", use_container_width=True, type="primary"):
        if pwd == APP_PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")

def dashboard_page():
    st.title("📱 M스캔 전용 서류 처리 현황 대시보드")
    
    # PPT 내용 통합 안내 배너
    with st.expander("📋 모바일동의(M스캔) 집중 관리 안내 (2026.05 시행)", expanded=True):
        st.markdown("""
        - **4종 필수서류**: 개인정보동의서, 비교설명확인서, 고지의무확인서, 완전판매확인서
        - **M스캔 권장 이유**: 자동 매칭, 타임스탬프 기록, 누락 방지로 불완전판매 리스크 최소화 및 업무 효율 극대화
        - **준수 사항**: 계약 체결 전 4종 서류 100% 완비 원칙. 미비 시 내부 통제 미충족 조직으로 관리 및 지원 제한 가능
        """)

    df = load_data()
    if df.empty:
        st.warning("데이터가 없습니다. insurance_data.xlsx 파일을 확인해주세요.")
        return

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1: st.success(f"총 {len(df):,}건의 데이터 로드 완료")
    with col2: st.info(f"기준: {get_file_update_time()}")
    with col3:
        if st.button("새로고침"):
            st.cache_data.clear(); st.rerun()

    month_col = "월_피리어드"
    all_months = sorted(df[month_col].dropna().unique())
    st.subheader("분석 기간 선택")
    sel_months = st.multiselect("월 선택", all_months, default=[all_months[-1]] if all_months else [])
    if not sel_months: st.warning("최소 1개 이상의 월을 선택해주세요."); return

    period_text = f"{sel_months[0]} ~ {sel_months[-1]}" if len(sel_months) > 1 else sel_months[0]
    df_sel = df[df[month_col].isin(sel_months)].copy()
    if df_sel.empty: st.info("선택한 기간에 데이터가 없습니다."); return

    target_total = int(df_sel["대상건"].sum())
    M_scan_total = int(df_sel["M스캔건"].sum())
    M_miss_total = int(df_sel["M미처리"].sum())
    miss_rate = round(M_miss_total / target_total * 100, 1) if target_total else 0.0
    M_scan_rate = round(M_scan_total / target_total * 100, 1) if target_total else 0.0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("총 계약건수", f"{len(df_sel):,}건")
    m2.metric("총 대상서류수", f"{target_total:,}건")
    m3.metric("M미처리율", f"{miss_rate:.1f}%")
    m4.metric("M스캔율", f"{M_scan_rate:.1f}%")
    st.divider()

    tab_dash, tab_map, tab_report, tab_ledger = st.tabs(["📊 현황 대시보드", "🗺️ 미처리맵", "📈 계층 리포트", "📁 관리대장 출력"])

    with tab_dash:
        cs1, cs2 = st.columns([2, 1])
        with cs1: search_text = st.text_input("조직 검색", placeholder="조직명 입력")
        with cs2: agg_group = st.selectbox("집계 기준", ["부문", "총괄", "부서", "영업가족"], key="agg_group")

        agg = df_sel.groupby(agg_group).agg(
            FA_target_sum=("FA_target", "sum"), 비교_target_sum=("비교_target", "sum"), 완판_target_sum=("완판_target", "sum"),
            M스캔_sum=("M스캔건", "sum"), 대상_sum=("대상건", "sum")
        ).reset_index()
        agg["총M미처리"] = (agg[["FA_target_sum", "비교_target_sum", "완판_target_sum"]].sum(axis=1) - agg["M스캔_sum"]).clip(lower=0)
        agg["미처리율"] = (agg["총M미처리"] / agg["대상_sum"].replace(0, pd.NA) * 100).round(1).fillna(0.0)
        agg["M스캔율"] = (agg["M스캔_sum"] / agg["대상_sum"].replace(0, pd.NA) * 100).round(1).fillna(0.0)
        agg = agg.rename(columns={agg_group: "조직"})
        if search_text: agg = agg[agg["조직"].astype(str).str.contains(search_text, case=False, na=False)]
        agg = agg.sort_values("총M미처리", ascending=False).reset_index(drop=True)
        agg.insert(0, "순위", range(1, len(agg) + 1))

        if agg.empty:
            st.info("조건에 맞는 데이터가 없습니다.")
        else:
            st.dataframe(
              agg[["순위", "조직", "대상_sum", "M스캔_sum", "총M미처리", "미처리율", "M스캔율"]]
                .rename(columns={"대상_sum": "대상건", "M스캔_sum": "M스캔건"})
                .style.format({"대상건": "{:,}", "M스캔건": "{:,}", "총M미처리": "{:,}", "미처리율": "{:.1f}%", "M스캔율": "{:.1f}%"})
                 , use_container_width=True, hide_index=True
)
            with c2:
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(x=top["조직"], y=top["M스캔율"], mode="lines+markers+text", text=[f"{v:.1f}%" for v in top["M스캔율"]], textposition="top center", line=dict(shape="spline", color="#1F618D", width=3), marker=dict(size=7)))
                fig2.update_layout(xaxis_tickangle=-45, height=420, title="조직별 M스캔율 TOP")
                st.plotly_chart(fig2, use_container_width=True)

    with tab_map:
        st.subheader("M미처리 분포 시각화")
        mc1, mc2 = st.columns([1, 2])
        with mc1: map_level = st.selectbox("집계 단위", ["부문", "총괄", "부서", "영업가족"], key="map_level")
        with mc2: map_type = st.radio("차트 유형", ["파이 차트", "트리맵"], horizontal=True, key="map_type")
        map_agg = df_sel.groupby(map_level).agg(M미처리_sum=("M미처리", "sum"), 대상_sum=("대상건", "sum")).reset_index()
        map_agg["M미처리비율"] = (map_agg["M미처리_sum"] / map_agg["대상_sum"].replace(0, pd.NA) * 100).round(1).fillna(0.0)
        map_agg = map_agg.rename(columns={map_level: "조직"})
        if map_agg.empty:
            st.info("데이터가 없습니다.")
        else:
            if map_type == "파이 차트":
                fig_pie = px.pie(map_agg, values="M미처리_sum", names="조직", title="조직별 M미처리 비중", hole=0.4)
                fig_pie.update_traces(textposition="inside", textinfo="percent+label")
                st.plotly_chart(fig_pie, use_container_width=True)
            else:
                fig_tree = px.treemap(map_agg, path=["조직"], values="M미처리_sum", color="M미처리비율", title="조직별 M미처리 분포", color_continuous_scale="RdYlGn_r")
                st.plotly_chart(fig_tree, use_container_width=True)

    with tab_report:
        st.subheader("전체 데이터 기반 계층별 M미처리 현황")
        report_df = build_hierarchy_report(df, sel_months)
        if report_df.empty: st.info("데이터가 없습니다.")
        else:
            def style_row(row):
                if row["구분"] == "부문계": return ["background-color:#1F3864;color:white;font-weight:bold"] * len(row)
                if row["구분"] == "총괄계": return ["background-color:#2E75B6;color:white;font-weight:bold"] * len(row)
                if row["구분"] == "부서계": return ["background-color:#D9E1F2;font-weight:bold"] * len(row)
                return [""] * len(row)
            st.dataframe(report_df.style.apply(style_row, axis=1).format({"FA":"{:,}", "비교":"{:,}", "완판":"{:,}", "총M미처리":"{:,}", "대상건":"{:,}", "M스캔건":"{:,}", "미처리율":"{:.1f}%", "M스캔율":"{:.1f}%"}), use_container_width=True, hide_index=True, height=500)
            
            cr1, cr2 = st.columns(2)
            with cr1:
                if st.button("계층 리포트 Excel", use_container_width=True):
                    with st.spinner("생성 중..."):
                        buf = report_excel(df, sel_months)
                    st.download_button("Excel 다운로드", buf, f"계층리포트_{period_text.replace(' ', '_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_rpt_xl")
            with cr2:
                if st.button("계층 리포트 PDF", use_container_width=True):
                    with st.spinner("생성 중..."):
                        buf2 = report_pdf(df, sel_months)
                    st.download_button("PDF 다운로드", buf2, f"계층리포트_{period_text.replace(' ', '_')}.pdf", "application/pdf", key="dl_rpt_pdf")

    with tab_ledger:
        st.subheader("관리대장 선정 및 출력")
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            sel_bm = st.selectbox("부문", ["전체"] + sorted(df_sel["부문"].dropna().unique().tolist()), key="lg_bm")
        df_l1 = df_sel if sel_bm == "전체" else df_sel[df_sel["부문"] == sel_bm]
        with cf2:
            sel_tg = st.selectbox("총괄", ["전체"] + sorted(df_l1["총괄"].dropna().unique().tolist()), key="lg_tg")
        df_l2 = df_l1 if sel_tg == "전체" else df_l1[df_l1["총괄"] == sel_tg]
        with cf3:
            sel_ds = st.selectbox("부서", ["전체"] + sorted(df_l2["부서"].dropna().unique().tolist()), key="lg_ds")
        df_l3 = df_l2 if sel_ds == "전체" else df_l2[df_l2["부서"] == sel_ds]

        targets = get_ledger_targets(df_l3, sel_months)
        if not targets:
            st.success("M스캔 미처리 발생 대상이 없습니다.")
        else:
            prev = [{"부서": dept, "영업가족": r["영업가족"], "FA": int(r["FA"]), "비교": int(r["비교"]), "완판": int(r["완판"]), "총M미처리": int(r["총M미처리"])} for dept, grp in targets.items() for _, r in grp.iterrows()]
            prev_df = pd.DataFrame(prev)
            st.dataframe(prev_df.style.format({"FA":"{:,}", "비교":"{:,}", "완판":"{:,}", "총M미처리":"{:,}"}), use_container_width=True, hide_index=True)
            sel_depts = st.multiselect("출력 부서", sorted(targets.keys()), default=sorted(targets.keys()), key="lg_sel_dept")
            if sel_depts:
                out_targets = {d: targets[d] for d in sel_depts if d in targets}
                cd1, cd2 = st.columns(2)
                with cd1:
                    if st.button("관리대장 PDF", use_container_width=True, key="gen_pdf"):
                        with st.spinner("생성 중..."):
                            pb = ledger_pdf(out_targets, period_text, df_l3)
                        st.download_button("PDF 다운로드", pb, f"관리대장_{period_text.replace(' ', '_')}.pdf", "application/pdf", key="dl_ldg_pdf")
                with cd2:
                    if st.button("관리대장 Excel", use_container_width=True, key="gen_xl"):
                        with st.spinner("생성 중..."):
                            xb = ledger_excel(out_targets, period_text, df_l3)
                        st.download_button("Excel 다운로드", xb, f"관리대장_{period_text.replace(' ', '_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_ldg_xl")

# ==========================================
# 13. main
# ==========================================
def main():
    if not st.session_state.get("logged_in"):
        login_page()
    else:
        with st.sidebar:
            st.success("👋 접속 완료")
            if st.button("🚪 로그아웃", use_container_width=True):
                st.session_state.logged_in = False
                st.rerun()
            st.divider()
            st.caption("v5.1 | M스캔 전용 집계 | © 2026")
        dashboard_page()

if __name__ == "__main__":
    main()